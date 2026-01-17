"""
Export views for Django CRM.
Handles CSV and Excel export of leads with filters.
"""

import csv
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from .models import Lead
import logging

logger = logging.getLogger(__name__)


def build_lead_filters(filters: dict) -> Q:
    """
    Build Django Q object from filter dictionary.
    
    Args:
        filters: Dictionary with filter parameters
            - search: str - Search in name, email, phone, company
            - status: str - Lead status
            - source: str - Lead source
            - lead_type: str - Lead type
            - min_score: int - Minimum quality score
            - has_phone: bool - Filter by phone presence
            - date_from: str - Created after date (ISO format)
            - date_to: str - Created before date (ISO format)
    
    Returns:
        Django Q object for filtering
    """
    q = Q()
    
    # Search
    search = filters.get('search', '').strip()
    if search:
        q &= (
            Q(name__icontains=search) |
            Q(email__icontains=search) |
            Q(telefon__icontains=search) |
            Q(company__icontains=search)
        )
    
    # Status
    status_filter = filters.get('status', '').strip()
    if status_filter and status_filter != 'all':
        q &= Q(status=status_filter)
    
    # Source
    source = filters.get('source', '').strip()
    if source and source != 'all':
        q &= Q(source=source)
    
    # Lead type
    lead_type = filters.get('lead_type', '').strip()
    if lead_type and lead_type != 'all':
        q &= Q(lead_type=lead_type)
    
    # Min score
    min_score = filters.get('min_score')
    if min_score:
        try:
            q &= Q(quality_score__gte=int(min_score))
        except (ValueError, TypeError):
            pass
    
    # Has phone
    has_phone = filters.get('has_phone')
    if has_phone in ['true', True, '1', 1]:
        q &= ~Q(telefon__isnull=True) & ~Q(telefon='')
    
    # Date range
    date_from = filters.get('date_from')
    if date_from:
        q &= Q(created_at__gte=date_from)
    
    date_to = filters.get('date_to')
    if date_to:
        q &= Q(created_at__lte=date_to)
    
    return q


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_leads_csv(request):
    """
    Export leads as CSV with optional filters.
    
    Query parameters:
        - search, status, source, lead_type, min_score, has_phone, date_from, date_to
    
    Returns:
        CSV file download
    """
    try:
        # Build filters
        filters = {
            'search': request.GET.get('search', ''),
            'status': request.GET.get('status', ''),
            'source': request.GET.get('source', ''),
            'lead_type': request.GET.get('lead_type', ''),
            'min_score': request.GET.get('min_score'),
            'has_phone': request.GET.get('has_phone'),
            'date_from': request.GET.get('date_from'),
            'date_to': request.GET.get('date_to'),
        }
        
        # Get user role to determine if we need to filter by assigned_to
        user_role = None
        if request.user.groups.exists():
            user_role = request.user.groups.first().name
        
        # Build queryset
        queryset = Lead.objects.all()
        
        # Filter by assigned_to for Telefonist
        if user_role == 'Telefonist':
            queryset = queryset.filter(assigned_to=request.user)
        
        # Apply filters
        q = build_lead_filters(filters)
        queryset = queryset.filter(q).order_by('-quality_score', '-created_at')
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        headers = [
            'ID', 'Name', 'Email', 'Telefon', 'Status', 'Quelle', 
            'Qualitäts-Score', 'Lead-Typ', 'Firma', 'Position', 
            'Standort', 'Interesse-Level', 'Anrufe', 'Erstellt am'
        ]
        writer.writerow(headers)
        
        # Write data
        for lead in queryset:
            writer.writerow([
                lead.id,
                lead.name,
                lead.email or '',
                lead.telefon or '',
                lead.get_status_display(),
                lead.get_source_display(),
                lead.quality_score,
                lead.get_lead_type_display(),
                lead.company or '',
                lead.role or '',
                lead.location or '',
                lead.interest_level,
                lead.call_count,
                lead.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])
        
        # Create response
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="leads_export.csv"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting leads to CSV: {e}", exc_info=True)
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_leads_excel(request):
    """
    Export leads as Excel (XLSX) with optional filters.
    
    Query parameters:
        - search, status, source, lead_type, min_score, has_phone, date_from, date_to
    
    Returns:
        Excel file download
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Build filters
        filters = {
            'search': request.GET.get('search', ''),
            'status': request.GET.get('status', ''),
            'source': request.GET.get('source', ''),
            'lead_type': request.GET.get('lead_type', ''),
            'min_score': request.GET.get('min_score'),
            'has_phone': request.GET.get('has_phone'),
            'date_from': request.GET.get('date_from'),
            'date_to': request.GET.get('date_to'),
        }
        
        # Get user role
        user_role = None
        if request.user.groups.exists():
            user_role = request.user.groups.first().name
        
        # Build queryset
        queryset = Lead.objects.all()
        
        # Filter by assigned_to for Telefonist
        if user_role == 'Telefonist':
            queryset = queryset.filter(assigned_to=request.user)
        
        # Apply filters
        q = build_lead_filters(filters)
        queryset = queryset.filter(q).order_by('-quality_score', '-created_at')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Write header
        headers = [
            'ID', 'Name', 'Email', 'Telefon', 'Status', 'Quelle', 
            'Qualitäts-Score', 'Lead-Typ', 'Firma', 'Position', 
            'Standort', 'Interesse-Level', 'Anrufe', 'Erstellt am'
        ]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color='06b6d4', end_color='06b6d4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for lead in queryset:
            ws.append([
                lead.id,
                lead.name,
                lead.email or '',
                lead.telefon or '',
                lead.get_status_display(),
                lead.get_source_display(),
                lead.quality_score,
                lead.get_lead_type_display(),
                lead.company or '',
                lead.role or '',
                lead.location or '',
                lead.interest_level,
                lead.call_count,
                lead.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="leads_export.xlsx"'
        
        return response
        
    except ImportError:
        logger.error("openpyxl not installed")
        return Response({
            'success': False,
            'error': 'Excel export nicht verfügbar (openpyxl nicht installiert)'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    except Exception as e:
        logger.error(f"Error exporting leads to Excel: {e}", exc_info=True)
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
