"""
Export Service
Exportiert Reports als PDF, Excel, CSV
"""
import io
import csv
from datetime import datetime
from typing import Dict, Any, Optional


class ReportExporter:
    """Exportiert Reports in verschiedene Formate"""
    
    def export_pdf(self, report_data: Dict[str, Any], title: str = "Report") -> bytes:
        """Exportiert Report als PDF"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Titel
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
            )
            story.append(Paragraph(title, title_style))
            
            # Zeitraum
            if 'period' in report_data:
                period = report_data['period']
                story.append(Paragraph(
                    f"Zeitraum: {period.get('start', '')} bis {period.get('end', '')}",
                    styles['Normal']
                ))
                story.append(Spacer(1, 20))
            
            # Summary
            if 'summary' in report_data:
                story.append(Paragraph("Zusammenfassung", styles['Heading2']))
                summary_data = [[k, str(v)] for k, v in report_data['summary'].items()]
                if summary_data:
                    table = Table(summary_data, colWidths=[200, 200])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 1, colors.white),
                    ]))
                    story.append(table)
            
            doc.build(story)
            return buffer.getvalue()
            
        except ImportError:
            # Fallback: Einfaches Text-PDF
            return self._simple_pdf_fallback(report_data, title)
    
    def _simple_pdf_fallback(self, report_data: Dict[str, Any], title: str) -> bytes:
        """Einfacher Text-Export wenn reportlab nicht verfügbar"""
        content = f"# {title}\n\n"
        content += f"Generiert: {datetime.now().isoformat()}\n\n"
        
        if 'summary' in report_data:
            content += "## Zusammenfassung\n"
            for k, v in report_data['summary'].items():
                content += f"- {k}: {v}\n"
        
        return content.encode('utf-8')
    
    def export_excel(self, report_data: Dict[str, Any], title: str = "Report") -> bytes:
        """Exportiert Report als Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = title[:31]  # Excel Limit
            
            # Header Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1E90FF", end_color="1E90FF", fill_type="solid")
            
            row = 1
            
            # Titel
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=14)
            row += 2
            
            # Summary
            if 'summary' in report_data:
                ws.cell(row=row, column=1, value="Zusammenfassung").font = Font(bold=True)
                row += 1
                for k, v in report_data['summary'].items():
                    ws.cell(row=row, column=1, value=str(k))
                    ws.cell(row=row, column=2, value=str(v))
                    row += 1
                row += 1
            
            # Listen-Daten (z.B. sources, funnel, etc.)
            for key in ['sources', 'funnel', 'status_distribution', 'source_distribution']:
                if key in report_data and report_data[key]:
                    ws.cell(row=row, column=1, value=key.replace('_', ' ').title()).font = Font(bold=True)
                    row += 1
                    
                    data_list = report_data[key]
                    if data_list and isinstance(data_list[0], dict):
                        # Header
                        headers = list(data_list[0].keys())
                        for col, header in enumerate(headers, 1):
                            cell = ws.cell(row=row, column=col, value=header)
                            cell.font = header_font
                            cell.fill = header_fill
                        row += 1
                        
                        # Daten
                        for item in data_list:
                            for col, header in enumerate(headers, 1):
                                ws.cell(row=row, column=col, value=item.get(header, ''))
                            row += 1
                    row += 1
            
            # Auto-fit Spaltenbreite
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        # Skip cells that can't be processed
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            return buffer.getvalue()
            
        except ImportError:
            raise ImportError("openpyxl nicht installiert. Bitte 'pip install openpyxl' ausführen.")
    
    def export_csv(self, report_data: Dict[str, Any]) -> str:
        """Exportiert Report als CSV"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Summary
        if 'summary' in report_data:
            writer.writerow(['Zusammenfassung', ''])
            for k, v in report_data['summary'].items():
                writer.writerow([k, v])
            writer.writerow([])
        
        # Listen-Daten
        for key in ['sources', 'funnel', 'status_distribution', 'source_distribution', 'daily_trend']:
            if key in report_data and report_data[key]:
                data_list = report_data[key]
                if data_list and isinstance(data_list[0], dict):
                    writer.writerow([key.replace('_', ' ').title()])
                    headers = list(data_list[0].keys())
                    writer.writerow(headers)
                    for item in data_list:
                        writer.writerow([item.get(h, '') for h in headers])
                    writer.writerow([])
        
        return output.getvalue()
