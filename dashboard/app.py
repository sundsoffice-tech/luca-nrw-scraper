# -*- coding: utf-8 -*-
"""
LUCA Control Center - Main Dashboard Application
"""

import os
import sys
import json
import queue
import threading
import time
from flask import Flask, render_template, jsonify, request, Response, send_file
from flask_cors import CORS

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dashboard.db_schema import (
    ensure_dashboard_schema, 
    initialize_default_search_modes, 
    initialize_default_settings,
    save_performance_settings,
    load_performance_settings
)
from dashboard.api import stats, costs, modes, settings
from dashboard.scraper_control import get_controller
from dashboard.scheduler import get_scheduler

# Third-party imports for performance monitoring
import psutil

# Database path
DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'scraper.db')

# Global log queue for SSE
log_queue = queue.Queue(maxsize=1000)

# Global performance state
_perf_state = {
    'mode': 'balanced',  # eco, balanced, power, custom
    'cpu_limit': 80,
    'ram_limit': 85,
    'auto_throttle': True,
    'auto_pause': True,
    'night_mode': False,
    'custom': {
        'threads': 4,
        'async_limit': 35,
        'batch_size': 20,
        'request_delay': 2.5
    },
    'current_multiplier': 1.0  # Throttle multiplier
}

# Performance mode presets
PERF_MODES = {
    'eco': {
        'threads': 1,
        'async_limit': 10,
        'batch_size': 5,
        'request_delay': 5.0,
        'description': 'Minimal CPU usage, slow but gentle'
    },
    'balanced': {
        'threads': 4,
        'async_limit': 35,
        'batch_size': 20,
        'request_delay': 2.5,
        'description': 'Auto-adjusts based on system load'
    },
    'power': {
        'threads': 8,
        'async_limit': 75,
        'batch_size': 40,
        'request_delay': 1.0,
        'description': 'Maximum speed, high CPU usage'
    }
}


def create_app(db_path: str = None) -> Flask:
    """
    Create and configure the Flask dashboard application.
    
    Args:
        db_path: Path to SQLite database (optional)
        
    Returns:
        Configured Flask application
    """
    global DB_PATH
    if db_path:
        DB_PATH = db_path
    
    app = Flask(__name__, 
                template_folder='templates',
                static_folder='static')
    CORS(app)
    
    # Initialize database schema
    import sqlite3
    con = sqlite3.connect(DB_PATH)
    ensure_dashboard_schema(con)
    initialize_default_search_modes(con)
    initialize_default_settings(con)
    con.close()
    
    # ==================== Routes ====================
    
    @app.route('/')
    def index():
        """Render main dashboard page."""
        return render_template('dashboard.html')
    
    @app.route('/analytics')
    def analytics():
        """Render analytics page."""
        return render_template('analytics.html')
    
    @app.route('/settings')
    def settings_page():
        """Render settings page."""
        return render_template('settings.html')
    
    @app.route('/leads')
    def leads_page():
        """Render leads management page."""
        return render_template('leads.html')
    
    # ==================== API Routes ====================
    
    @app.route('/api/stats')
    def api_stats():
        """Get dashboard statistics/KPIs."""
        try:
            data = stats.get_stats(DB_PATH)
            return jsonify(data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/stats/candidates')
    def api_stats_candidates():
        """Get candidate-specific statistics."""
        try:
            import sqlite3
            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            
            # Total candidates
            cur.execute("SELECT COUNT(*) as count FROM leads WHERE lead_type = 'candidate'")
            total_candidates = cur.fetchone()['count']
            
            # Candidates today
            cur.execute("""
                SELECT COUNT(*) as count FROM leads 
                WHERE lead_type = 'candidate' 
                AND DATE(last_updated) = DATE('now')
            """)
            candidates_today = cur.fetchone()['count']
            
            # Candidates with experience
            cur.execute("""
                SELECT COUNT(*) as count FROM leads 
                WHERE lead_type = 'candidate' 
                AND experience_years IS NOT NULL 
                AND experience_years > 0
            """)
            with_experience = cur.fetchone()['count']
            
            # Average experience years
            cur.execute("""
                SELECT AVG(experience_years) as avg_exp FROM leads 
                WHERE lead_type = 'candidate' 
                AND experience_years IS NOT NULL 
                AND experience_years > 0
            """)
            avg_experience = cur.fetchone()['avg_exp'] or 0
            
            # Candidates by location (top 5)
            cur.execute("""
                SELECT location, COUNT(*) as count 
                FROM leads 
                WHERE lead_type = 'candidate' 
                AND location IS NOT NULL 
                AND location != ''
                GROUP BY location 
                ORDER BY count DESC 
                LIMIT 5
            """)
            by_location = [dict(row) for row in cur.fetchall()]
            
            con.close()
            
            return jsonify({
                'total_candidates': total_candidates,
                'today': candidates_today,
                'with_experience': with_experience,
                'avg_experience_years': round(avg_experience, 1),
                'by_location': by_location
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/costs')
    def api_costs():
        """Get API cost breakdown."""
        try:
            data = costs.get_costs(DB_PATH)
            return jsonify(data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/costs/chart')
    def api_costs_chart():
        """Get cost data formatted for charts."""
        try:
            days = int(request.args.get('days', 7))
            data = costs.get_cost_chart_data(DB_PATH, days)
            return jsonify(data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/charts/leads')
    def api_charts_leads():
        """Get lead history chart data."""
        try:
            days = int(request.args.get('days', 7))
            data = stats.get_leads_history(DB_PATH, days)
            return jsonify(data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/charts/sources')
    def api_charts_sources():
        """Get top sources chart data."""
        try:
            limit = int(request.args.get('limit', 10))
            data = stats.get_top_sources(DB_PATH, limit)
            return jsonify(data)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/mode', methods=['GET', 'POST'])
    def api_mode():
        """Get or set search mode."""
        try:
            if request.method == 'GET':
                current = modes.get_current_mode(DB_PATH)
                available = modes.get_all_modes(DB_PATH)
                return jsonify({
                    'current_mode': current,
                    'available_modes': available
                })
            else:
                data = request.get_json()
                mode_key = data.get('mode')
                if not mode_key:
                    return jsonify({'error': 'mode parameter required'}), 400
                
                success = modes.set_active_mode(DB_PATH, mode_key)
                if success:
                    # Log mode change
                    log_queue.put(json.dumps({
                        'timestamp': time.time(),
                        'level': 'INFO',
                        'message': f'Search mode changed to: {mode_key}'
                    }))
                    return jsonify({'success': True, 'mode': mode_key})
                else:
                    return jsonify({'error': 'Invalid mode or database error'}), 400
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/settings', methods=['GET', 'POST'])
    def api_settings():
        """Get or update settings."""
        try:
            if request.method == 'GET':
                data = settings.get_all_settings(DB_PATH)
                definitions = settings.get_setting_definitions()
                return jsonify({
                    'settings': data,
                    'definitions': definitions
                })
            else:
                data = request.get_json()
                results = settings.update_settings_bulk(DB_PATH, data)
                
                # Log settings change
                log_queue.put(json.dumps({
                    'timestamp': time.time(),
                    'level': 'INFO',
                    'message': f'Settings updated: {", ".join(data.keys())}'
                }))
                
                return jsonify({'success': True, 'results': results})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/learning/stats')
    def api_learning_stats():
        """Get learning statistics for learning mode."""
        con = None
        try:
            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            
            # Check if learning tables exist
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='learning_domains'")
            if not cur.fetchone():
                return jsonify({
                    'top_domains': [],
                    'top_patterns': [],
                    'total_domains_learned': 0,
                    'total_leads_from_learning': 0,
                    'message': 'Learning mode not yet used'
                })
            
            # Top Domains
            cur.execute("""
                SELECT domain, score, leads_found, total_visits 
                FROM learning_domains 
                ORDER BY score DESC LIMIT 10
            """)
            top_domains = [
                {
                    'domain': row['domain'], 
                    'score': round(row['score'], 3), 
                    'leads': row['leads_found'], 
                    'visits': row['total_visits']
                } 
                for row in cur.fetchall()
            ]
            
            # Top Patterns
            cur.execute("""
                SELECT pattern_type, pattern_value, confidence_score, success_count 
                FROM success_patterns 
                ORDER BY confidence_score DESC LIMIT 10
            """)
            top_patterns = [
                {
                    'type': row['pattern_type'], 
                    'value': row['pattern_value'], 
                    'confidence': round(row['confidence_score'], 3), 
                    'count': row['success_count']
                } 
                for row in cur.fetchall()
            ]
            
            # Overall statistics
            cur.execute("SELECT COUNT(*), SUM(leads_found) FROM learning_domains")
            domain_stats = cur.fetchone()
            
            return jsonify({
                'top_domains': top_domains,
                'top_patterns': top_patterns,
                'total_domains_learned': domain_stats[0] or 0,
                'total_leads_from_learning': domain_stats[1] or 0
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
        finally:
            if con:
                con.close()
    
    @app.route('/api/stream/logs')
    def api_stream_logs():
        """Server-Sent Events stream for live logs."""
        def generate():
            while True:
                try:
                    # Get log entry from queue with timeout
                    log_entry = log_queue.get(timeout=1)
                    yield f"data: {log_entry}\n\n"
                except queue.Empty:
                    # Send keepalive comment
                    yield ": keepalive\n\n"
                    time.sleep(0.5)
        
        return Response(generate(), mimetype='text/event-stream')
    
    @app.route('/api/export/csv')
    def api_export_csv():
        """Export leads as CSV."""
        try:
            import csv
            import io
            import sqlite3
            
            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            
            cur.execute("SELECT * FROM leads ORDER BY id DESC")
            rows = cur.fetchall()
            
            if not rows:
                con.close()
                return jsonify({'error': 'No leads to export'}), 404
            
            # Create CSV in memory
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            for row in rows:
                writer.writerow(dict(row))
            
            con.close()
            
            # Create response
            output.seek(0)
            return Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment;filename=leads.csv'}
            )
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    # ==================== Scraper Control API ====================
    
    @app.route('/api/scraper/start', methods=['POST'])
    def api_scraper_start():
        """Start the scraper with given parameters."""
        try:
            params = request.get_json() or {}
            controller = get_controller()
            result = controller.start(params)
            
            if result.get('success'):
                # Log start event
                log_queue.put(json.dumps({
                    'timestamp': time.time(),
                    'level': 'INFO',
                    'message': f'Scraper started with PID {result.get("pid")}'
                }))
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/scraper/stop', methods=['POST'])
    def api_scraper_stop():
        """Stop the running scraper."""
        try:
            controller = get_controller()
            result = controller.stop()
            
            if result.get('success'):
                log_queue.put(json.dumps({
                    'timestamp': time.time(),
                    'level': 'INFO',
                    'message': 'Scraper stopped'
                }))
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/scraper/pause', methods=['POST'])
    def api_scraper_pause():
        """Pause the running scraper."""
        try:
            controller = get_controller()
            result = controller.pause()
            
            if result.get('success'):
                log_queue.put(json.dumps({
                    'timestamp': time.time(),
                    'level': 'INFO',
                    'message': 'Scraper paused'
                }))
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/scraper/resume', methods=['POST'])
    def api_scraper_resume():
        """Resume the paused scraper."""
        try:
            controller = get_controller()
            result = controller.resume()
            
            if result.get('success'):
                log_queue.put(json.dumps({
                    'timestamp': time.time(),
                    'level': 'INFO',
                    'message': 'Scraper resumed'
                }))
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/scraper/reset', methods=['POST'])
    def api_scraper_reset():
        """Reset URL cache and queries."""
        try:
            controller = get_controller()
            result = controller.reset(DB_PATH)
            
            if result.get('success'):
                log_queue.put(json.dumps({
                    'timestamp': time.time(),
                    'level': 'INFO',
                    'message': 'Cache and queries reset'
                }))
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/api/scraper/status', methods=['GET'])
    def api_scraper_status():
        """Get current scraper status."""
        try:
            controller = get_controller()
            status = controller.get_status()
            return jsonify(status)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/scraper/output', methods=['GET'])
    def api_scraper_output():
        """Get recent scraper output."""
        try:
            lines = int(request.args.get('lines', 50))
            controller = get_controller()
            output = controller.get_output(lines)
            return jsonify({'output': output})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    # ==================== Scheduler API ====================
    
    @app.route('/api/scheduler', methods=['GET'])
    def api_scheduler_get():
        """Get scheduler configuration."""
        try:
            def dummy_start(params):
                # This will be called by scheduler
                controller = get_controller()
                return controller.start(params)
            
            scheduler = get_scheduler(DB_PATH, dummy_start)
            config = scheduler.get_config()
            
            # Calculate next run for display
            next_run = scheduler.calculate_next_run(config)
            if next_run:
                config['next_run_display'] = next_run.strftime('%Y-%m-%d %H:%M')
            
            return jsonify(config)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/scheduler', methods=['POST'])
    def api_scheduler_update():
        """Update scheduler configuration."""
        try:
            config = request.get_json() or {}
            
            def dummy_start(params):
                controller = get_controller()
                return controller.start(params)
            
            scheduler = get_scheduler(DB_PATH, dummy_start)
            result = scheduler.update_config(config)
            
            # Log change
            log_queue.put(json.dumps({
                'timestamp': time.time(),
                'level': 'INFO',
                'message': f'Scheduler updated: {"enabled" if config.get("enabled") else "disabled"}'
            }))
            
            return jsonify(result)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    # ==================== Leads API ====================
    
    @app.route('/api/leads', methods=['GET'])
    def api_leads_get():
        """Get leads with advanced filtering, sorting, and pagination."""
        try:
            import sqlite3
            
            # Pagination
            page = int(request.args.get('page', 1))
            per_page = int(request.args.get('per_page', 50))
            
            # Existing filters
            search = request.args.get('search', '')
            status = request.args.get('status', '')
            date_from = request.args.get('date_from', '')
            lead_type = request.args.get('lead_type', '')
            
            # New filters
            phone_filter = request.args.get('phone', 'all')
            email_filter = request.args.get('email', 'all')
            source_filter = request.args.get('source', 'all')
            date_filter = request.args.get('date', 'all')
            date_to = request.args.get('date_to', '')
            industry_filter = request.args.get('industry', 'all')
            quality_filter = request.args.get('quality', 'all')
            
            # Sorting
            sort_by = request.args.get('sort', 'created_at')
            sort_dir = request.args.get('dir', 'desc')
            
            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            
            # Build query
            where_clauses = []
            params = []
            
            # Search filter
            if search:
                where_clauses.append("(name LIKE ? OR company_name LIKE ? OR company LIKE ? OR telefon LIKE ? OR mobile_number LIKE ? OR email LIKE ?)")
                search_pattern = f"%{search}%"
                params.extend([search_pattern] * 6)
            
            # Status filter
            if status and status != 'all':
                where_clauses.append("status = ?")
                params.append(status)
            
            # Lead type filter
            if lead_type and lead_type != 'all':
                where_clauses.append("lead_type = ?")
                params.append(lead_type)
            
            # Phone filter
            if phone_filter == 'has_mobile':
                where_clauses.append("((mobile_number IS NOT NULL AND mobile_number != '') OR (telefon IS NOT NULL AND telefon != ''))")
            elif phone_filter == 'no_mobile':
                where_clauses.append("((mobile_number IS NULL OR mobile_number = '') AND (telefon IS NULL OR telefon = ''))")
            elif phone_filter == 'mobile_only':
                where_clauses.append("((mobile_number LIKE '015%' OR mobile_number LIKE '016%' OR mobile_number LIKE '017%') OR (telefon LIKE '015%' OR telefon LIKE '016%' OR telefon LIKE '017%'))")
            elif phone_filter == 'landline_only':
                where_clauses.append("(((mobile_number IS NOT NULL AND mobile_number != '' AND mobile_number NOT LIKE '01%') OR (telefon IS NOT NULL AND telefon != '' AND telefon NOT LIKE '01%')))")
            
            # Email filter
            if email_filter == 'has_email':
                where_clauses.append("email IS NOT NULL AND email != ''")
            elif email_filter == 'no_email':
                where_clauses.append("(email IS NULL OR email = '')")
            
            # Source filter
            if source_filter != 'all':
                where_clauses.append("(source_url LIKE ? OR quelle LIKE ?)")
                source_pattern = f'%{source_filter}%'
                params.extend([source_pattern, source_pattern])
            
            # Date filter
            if date_filter == 'today':
                where_clauses.append("(DATE(created_at) = DATE('now') OR DATE(last_updated) = DATE('now'))")
            elif date_filter == 'yesterday':
                where_clauses.append("(DATE(created_at) = DATE('now', '-1 day') OR DATE(last_updated) = DATE('now', '-1 day'))")
            elif date_filter == '7days':
                where_clauses.append("((created_at >= DATE('now', '-7 days')) OR (last_updated >= DATE('now', '-7 days')))")
            elif date_filter == '30days':
                where_clauses.append("((created_at >= DATE('now', '-30 days')) OR (last_updated >= DATE('now', '-30 days')))")
            elif date_filter == 'custom' and date_from and date_to:
                where_clauses.append("((DATE(created_at) BETWEEN ? AND ?) OR (DATE(last_updated) BETWEEN ? AND ?))")
                params.extend([date_from, date_to, date_from, date_to])
            elif date_from:
                where_clauses.append("(DATE(created_at) >= ? OR DATE(last_updated) >= ?)")
                params.extend([date_from, date_from])
            
            # Industry filter
            if industry_filter != 'all':
                where_clauses.append("industry = ?")
                params.append(industry_filter)
            
            # Quality filter (confidence score)
            if quality_filter == 'high':
                where_clauses.append("confidence > 0.8")
            elif quality_filter == 'medium':
                where_clauses.append("confidence BETWEEN 0.5 AND 0.8")
            elif quality_filter == 'low':
                where_clauses.append("confidence < 0.5")
            
            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
            
            # Get total count
            count_query = f"SELECT COUNT(*) FROM leads WHERE {where_sql}"
            cur.execute(count_query, params)
            total = cur.fetchone()[0]
            
            # Sorting
            allowed_sorts = ['name', 'mobile_number', 'telefon', 'email', 'company', 'company_name', 'source_url', 'quelle', 'created_at', 'last_updated', 'confidence', 'id']
            if sort_by in allowed_sorts:
                sort_direction = 'DESC' if sort_dir == 'desc' else 'ASC'
                order_by = f"{sort_by} {sort_direction}"
            else:
                order_by = "created_at DESC"
            
            # Get paginated data
            offset = (page - 1) * per_page
            data_query = f"SELECT * FROM leads WHERE {where_sql} ORDER BY {order_by} LIMIT ? OFFSET ?"
            cur.execute(data_query, params + [per_page, offset])
            rows = cur.fetchall()
            
            leads = [dict(row) for row in rows]
            
            con.close()
            
            return jsonify({
                'leads': leads,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': total,
                    'total_pages': (total + per_page - 1) // per_page
                }
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/leads/<int:lead_id>/status', methods=['PUT'])
    def api_lead_update_status(lead_id):
        """Update lead status."""
        try:
            import sqlite3
            
            data = request.get_json() or {}
            new_status = data.get('status')
            
            if not new_status:
                return jsonify({'error': 'status parameter required'}), 400
            
            con = sqlite3.connect(DB_PATH)
            cur = con.cursor()
            
            cur.execute("""
                UPDATE leads SET status = ? WHERE id = ?
            """, (new_status, lead_id))
            
            con.commit()
            con.close()
            
            return jsonify({'success': True, 'status': new_status})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/leads/export/<format>', methods=['GET'])
    def api_leads_export(format):
        """Export leads in various formats."""
        try:
            import sqlite3
            import csv
            import io
            
            # Get filters from query string
            search = request.args.get('search', '')
            status = request.args.get('status', '')
            
            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            
            # Build query with filters
            where_clauses = []
            params = []
            
            if search:
                where_clauses.append("(name LIKE ? OR company LIKE ?)")
                search_pattern = f"%{search}%"
                params.extend([search_pattern, search_pattern])
            
            if status:
                where_clauses.append("status = ?")
                params.append(status)
            
            where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
            query = f"SELECT * FROM leads WHERE {where_sql} ORDER BY id DESC"
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            if not rows:
                con.close()
                return jsonify({'error': 'No leads to export'}), 404
            
            if format == 'csv':
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=rows[0].keys())
                writer.writeheader()
                for row in rows:
                    writer.writerow(dict(row))
                
                con.close()
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=leads.csv'}
                )
            
            elif format == 'xlsx':
                # XLSX export using openpyxl directly
                from openpyxl import Workbook
                from openpyxl.styles import Font
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Leads"
                
                # Write header
                headers = list(rows[0].keys())
                ws.append(headers)
                
                # Style header
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                
                # Write data
                for row in rows:
                    ws.append(list(row))
                
                con.close()
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return Response(
                    output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment;filename=leads.xlsx'}
                )
            
            elif format == 'pdf':
                # Simple PDF export using reportlab
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
                from reportlab.lib.styles import getSampleStyleSheet
                
                output = io.BytesIO()
                doc = SimpleDocTemplate(output, pagesize=A4)
                elements = []
                
                # Title
                styles = getSampleStyleSheet()
                title = Paragraph("LUCA - Lead Export", styles['Title'])
                elements.append(title)
                
                # Prepare data for table
                # Limit to 100 rows for PDF to keep file size reasonable
                PDF_EXPORT_LIMIT = 100
                data = [list(rows[0].keys())]  # Header
                for row in rows[:PDF_EXPORT_LIMIT]:
                    data.append([str(v) if v is not None else '' for v in row])
                
                # Create table
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                doc.build(elements)
                
                con.close()
                output.seek(0)
                return Response(
                    output.getvalue(),
                    mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment;filename=leads.pdf'}
                )
            
            else:
                con.close()
                return jsonify({'error': 'Invalid format'}), 400
            
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/leads/export', methods=['POST'])
    def api_leads_export_bulk():
        """Export selected or filtered leads in various formats."""
        try:
            import sqlite3
            import csv
            import io
            from datetime import datetime
            
            data = request.get_json() or {}
            lead_ids = data.get('ids', [])
            format_type = data.get('format', 'csv')
            filters = data.get('filters', {})
            
            con = sqlite3.connect(DB_PATH)
            con.row_factory = sqlite3.Row
            cur = con.cursor()
            
            # Build query
            if lead_ids:
                # Export specific IDs
                placeholders = ','.join('?' * len(lead_ids))
                query = f"SELECT * FROM leads WHERE id IN ({placeholders})"
                params = lead_ids
            else:
                # Export with filters
                where_clauses = []
                params = []
                
                for key, value in filters.items():
                    if not value or value == 'all':
                        continue
                    
                    if key == 'search' and value:
                        where_clauses.append("(name LIKE ? OR company LIKE ? OR email LIKE ?)")
                        search_pattern = f"%{value}%"
                        params.extend([search_pattern] * 3)
                    elif key == 'status' and value:
                        where_clauses.append("status = ?")
                        params.append(value)
                    elif key == 'lead_type' and value:
                        where_clauses.append("lead_type = ?")
                        params.append(value)
                
                where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
                query = f"SELECT * FROM leads WHERE {where_sql} ORDER BY id DESC"
            
            cur.execute(query, params)
            rows = cur.fetchall()
            
            if not rows:
                con.close()
                return jsonify({'error': 'No leads to export'}), 404
            
            leads = [dict(row) for row in rows]
            con.close()
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            
            if format_type == 'csv':
                output = io.StringIO()
                if leads:
                    writer = csv.DictWriter(output, fieldnames=leads[0].keys())
                    writer.writeheader()
                    writer.writerows(leads)
                
                response = Response(
                    output.getvalue(),
                    mimetype='text/csv; charset=utf-8'
                )
                response.headers['Content-Disposition'] = f'attachment; filename=leads_export_{timestamp}.csv'
                return response
            
            elif format_type == 'excel':
                try:
                    from openpyxl import Workbook
                    from io import BytesIO
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Leads"
                    
                    if leads:
                        # Header
                        headers = list(leads[0].keys())
                        ws.append(headers)
                        
                        # Data
                        for lead in leads:
                            ws.append(list(lead.values()))
                    
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    
                    response = Response(
                        output.getvalue(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    response.headers['Content-Disposition'] = f'attachment; filename=leads_export_{timestamp}.xlsx'
                    return response
                except ImportError:
                    return jsonify({'error': 'openpyxl nicht installiert'}), 500
            
            elif format_type == 'vcard':
                vcards = []
                for lead in leads:
                    vcard = f"""BEGIN:VCARD
VERSION:3.0
FN:{lead.get('name', 'Unknown')}
ORG:{lead.get('company', lead.get('company_name', ''))}
TEL;TYPE=CELL:{lead.get('mobile_number', lead.get('telefon', ''))}
EMAIL:{lead.get('email', '')}
NOTE:Quelle: {lead.get('source_url', lead.get('quelle', ''))}
END:VCARD"""
                    vcards.append(vcard)
                
                response = Response(
                    '\n'.join(vcards),
                    mimetype='text/vcard; charset=utf-8'
                )
                response.headers['Content-Disposition'] = f'attachment; filename=leads_contacts_{timestamp}.vcf'
                return response
            
            else:
                return jsonify({'error': 'Invalid format'}), 400
                
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/leads/delete', methods=['POST'])
    def api_leads_delete_bulk():
        """Delete selected leads."""
        try:
            import sqlite3
            
            data = request.get_json() or {}
            lead_ids = data.get('ids', [])
            
            if not lead_ids:
                return jsonify({'error': 'Keine Leads ausgewählt'}), 400
            
            con = sqlite3.connect(DB_PATH)
            cur = con.cursor()
            
            placeholders = ','.join('?' * len(lead_ids))
            cur.execute(f"DELETE FROM leads WHERE id IN ({placeholders})", lead_ids)
            deleted = cur.rowcount
            
            con.commit()
            con.close()
            
            return jsonify({
                'success': True,
                'deleted': deleted,
                'message': f'{deleted} Lead(s) gelöscht'
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/leads/<int:lead_id>', methods=['DELETE'])
    def api_leads_delete_single(lead_id):
        """Delete a single lead."""
        try:
            import sqlite3
            
            con = sqlite3.connect(DB_PATH)
            cur = con.cursor()
            cur.execute("DELETE FROM leads WHERE id = ?", (lead_id,))
            deleted = cur.rowcount
            con.commit()
            con.close()
            
            if deleted:
                return jsonify({'success': True, 'message': 'Lead gelöscht'})
            return jsonify({'error': 'Lead nicht gefunden'}), 404
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    # ==================== Health API ====================
    
    @app.route('/api/health', methods=['GET'])
    def api_health():
        """Get system health information."""
        try:
            import sqlite3
            import os
            
            health_info = {
                'api_keys': {
                    'openai': 'unknown',
                    'google_cse': 'unknown',
                    'perplexity': 'unknown'
                },
                'circuit_breaker': [],
                'last_errors': [],
                'db_size_mb': 0,
                'uptime_hours': 0
            }
            
            # Check API keys from environment
            if os.getenv('OPENAI_API_KEY'):
                health_info['api_keys']['openai'] = 'configured'
            else:
                health_info['api_keys']['openai'] = 'missing'
            
            if os.getenv('GOOGLE_CSE_API_KEY'):
                health_info['api_keys']['google_cse'] = 'configured'
            else:
                health_info['api_keys']['google_cse'] = 'missing'
            
            if os.getenv('PERPLEXITY_API_KEY'):
                health_info['api_keys']['perplexity'] = 'configured'
            else:
                health_info['api_keys']['perplexity'] = 'missing'
            
            # Get database size
            if os.path.exists(DB_PATH):
                db_size_bytes = os.path.getsize(DB_PATH)
                health_info['db_size_mb'] = round(db_size_bytes / 1024 / 1024, 2)
            
            # Get recent errors from runs table
            try:
                con = sqlite3.connect(DB_PATH)
                con.row_factory = sqlite3.Row
                cur = con.cursor()
                
                cur.execute("""
                    SELECT id, status, error_message, started_at
                    FROM runs
                    WHERE status = 'error' OR error_message IS NOT NULL
                    ORDER BY id DESC
                    LIMIT 5
                """)
                
                error_rows = cur.fetchall()
                health_info['last_errors'] = [dict(row) for row in error_rows]
                
                con.close()
            except:
                pass
            
            return jsonify(health_info)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    # ==================== Performance API ====================
    
    @app.route('/api/performance', methods=['GET'])
    def api_performance_get():
        """Get current system performance metrics."""
        global _perf_state
        try:
            # System metrics
            cpu_percent = psutil.cpu_percent(interval=0.1)
            ram = psutil.virtual_memory()
            
            # Network I/O
            net_io = psutil.net_io_counters()
            
            # Scraper process metrics (if running)
            scraper_cpu = 0
            scraper_ram = 0
            controller = get_controller()
            if controller.pid and controller.is_running():
                try:
                    proc = psutil.Process(controller.pid)
                    scraper_cpu = proc.cpu_percent(interval=0)
                    scraper_ram = proc.memory_info().rss / 1024 / 1024  # MB
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            
            return jsonify({
                'system': {
                    'cpu_percent': cpu_percent,
                    'ram_percent': ram.percent,
                    'ram_used_gb': round(ram.used / 1024 / 1024 / 1024, 2),
                    'ram_total_gb': round(ram.total / 1024 / 1024 / 1024, 2),
                    'net_sent_mb': round(net_io.bytes_sent / 1024 / 1024, 2),
                    'net_recv_mb': round(net_io.bytes_recv / 1024 / 1024, 2)
                },
                'scraper': {
                    'cpu_percent': scraper_cpu,
                    'ram_mb': round(scraper_ram, 2)
                },
                'settings': _perf_state,
                'modes': PERF_MODES,
                'throttled': _perf_state['current_multiplier'] < 1.0
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/performance', methods=['POST'])
    def api_performance_set():
        """Update performance settings."""
        global _perf_state
        try:
            data = request.get_json()
            
            if 'mode' in data:
                _perf_state['mode'] = data['mode']
            if 'cpu_limit' in data:
                _perf_state['cpu_limit'] = int(data['cpu_limit'])
            if 'ram_limit' in data:
                _perf_state['ram_limit'] = int(data['ram_limit'])
            if 'auto_throttle' in data:
                _perf_state['auto_throttle'] = bool(data['auto_throttle'])
            if 'auto_pause' in data:
                _perf_state['auto_pause'] = bool(data['auto_pause'])
            if 'night_mode' in data:
                _perf_state['night_mode'] = bool(data['night_mode'])
            if 'custom' in data:
                _perf_state['custom'].update(data['custom'])
            
            # Save to database
            save_performance_settings(_perf_state)
            
            # Log settings change
            log_queue.put(json.dumps({
                'timestamp': time.time(),
                'level': 'INFO',
                'message': f'Performance settings updated: mode={_perf_state["mode"]}'
            }))
            
            return jsonify({'success': True, 'settings': _perf_state})
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    @app.route('/api/performance/effective', methods=['GET'])
    def api_performance_effective():
        """Get the effective performance parameters based on current mode and system load."""
        global _perf_state
        try:
            from datetime import datetime
            
            mode = _perf_state['mode']
            
            if mode == 'custom':
                params = _perf_state['custom'].copy()
            else:
                params = PERF_MODES.get(mode, PERF_MODES['balanced']).copy()
                # Remove description from params
                params.pop('description', None)
            
            # Apply throttle multiplier if auto-throttle is enabled
            if _perf_state['auto_throttle'] and _perf_state['current_multiplier'] < 1.0:
                multiplier = _perf_state['current_multiplier']
                params['async_limit'] = max(5, int(params['async_limit'] * multiplier))
                params['batch_size'] = max(3, int(params['batch_size'] * multiplier))
                params['request_delay'] = params['request_delay'] / multiplier
            
            # Night mode adjustments
            night_mode_active = False
            if _perf_state['night_mode']:
                hour = datetime.now().hour
                if hour >= 23 or hour < 6:
                    night_mode_active = True
                    params['async_limit'] = max(5, params['async_limit'] // 2)
                    params['request_delay'] = params['request_delay'] * 2
            
            return jsonify({
                'mode': mode,
                'params': params,
                'throttle_multiplier': _perf_state['current_multiplier'],
                'night_mode_active': night_mode_active
            })
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    return app


def start_performance_monitor():
    """Start background thread that monitors system performance and auto-throttles."""
    def monitor_loop():
        global _perf_state
        
        while True:
            try:
                if not _perf_state['auto_throttle']:
                    time.sleep(5)
                    continue
                
                cpu = psutil.cpu_percent(interval=1)
                ram = psutil.virtual_memory().percent
                
                cpu_limit = _perf_state['cpu_limit']
                ram_limit = _perf_state['ram_limit']
                
                # Calculate throttle multiplier
                multiplier = 1.0
                
                if cpu > cpu_limit:
                    # Reduce by how much we're over the limit
                    cpu_excess = (cpu - cpu_limit) / 100
                    multiplier = min(multiplier, 1.0 - cpu_excess)
                
                if ram > ram_limit:
                    ram_excess = (ram - ram_limit) / 100
                    multiplier = min(multiplier, 1.0 - ram_excess)
                
                # Clamp between 0.2 and 1.0
                multiplier = max(0.2, min(1.0, multiplier))
                
                _perf_state['current_multiplier'] = multiplier
                
                # Auto-pause if RAM is critical
                if _perf_state['auto_pause'] and ram > 95:
                    controller = get_controller()
                    if controller.is_running() and not controller.paused:
                        controller.pause()
                        print(f"[PERF] Auto-paused scraper due to high RAM: {ram}%")
                        add_log_entry('WARNING', f'Auto-paused scraper due to high RAM: {ram}%')
                
                time.sleep(3)
                
            except Exception as e:
                print(f"[PERF] Monitor error: {e}")
                time.sleep(5)
    
    thread = threading.Thread(target=monitor_loop, daemon=True)
    thread.start()
    return thread


def add_log_entry(level: str, message: str):
    """
    Add a log entry to the queue for SSE streaming.
    
    Args:
        level: Log level (INFO, WARNING, ERROR)
        message: Log message
    """
    try:
        log_queue.put(json.dumps({
            'timestamp': time.time(),
            'level': level,
            'message': message
        }), block=False)
    except queue.Full:
        pass  # Drop message if queue is full


if __name__ == '__main__':
    import os
    
    # Load performance settings from database
    saved_settings = load_performance_settings()
    _perf_state.update(saved_settings)
    
    # Start performance monitoring thread
    start_performance_monitor()
    
    app = create_app()
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    print(f"🎯 LUCA Control Center starting on http://127.0.0.1:5056")
    print(f"⚠️  Debug mode: {debug_mode}")
    if not debug_mode:
        print(f"⚠️  This is a development server. Use a production WSGI server for production.")
    app.run(host='127.0.0.1', port=5056, debug=debug_mode)
