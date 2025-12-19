# -*- coding: utf-8 -*-
"""
Leads-Scraper (NRW Vertrieb/Callcenter/D2D) — robust + inkrementell + UI

Features:
- Suche: Google CSE (Key/CX-Rotation, Backoff, Pagination) + Fallback Bing + Seed-URLs
- Fetch: Retries, SSL-Verify mit Fallback (Flag), robots tolerant (Cache)
- Filter: Harter Domain-Block (News/Behörden/Verbände), Pfad-Whitelist, Content-Validation tolerant
- Extraktion: OpenAI (JSON) oder Regex (E-Mail/Tel/WhatsApp/wa.me + Namensheuristik); Kleinanzeigen-Extractor
- Scoring: Kontakt-/Branchen-Signale, starker WhatsApp/Telefon/E-Mail-Boost, URL-Hints, dynamischer Schwellenwert
- Enrichment: Firma/Größe/Branche/Region/Frische + Confidence/DataQuality realistisch
- Persistenz: SQLite (runs, queries_done, urls_seen, leads) => keine Doppelten, Wiederaufnahme
- Export: CSV/XLSX (append nur neue Leads)
- Depth: Interne Tiefe + Sitemap-Hints
- UI: Flask (Start/Stop/Force/Reset/Seen-Reset/Queries-Reset + Live-Logs via SSE)
"""

import warnings
warnings.filterwarnings("ignore", message="This package.*renamed to.*ddgs")

# Stdlib
import argparse
import asyncio
import csv
import hashlib
import json
import os
import queue
import random
import re
import sqlite3
import sys
import threading
import time
import traceback
import urllib.parse
import pandas
import pandas as pd
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple, Union
from flask import Response, render_template_string


# Third-party
from curl_cffi.requests import AsyncSession
import tldextract
import urllib3
import aiohttp
import io
from bs4 import BeautifulSoup, XMLParsedAsHTMLWarning
from bs4.element import Comment
from pypdf import PdfReader
import re
try:
    from ddgs import DDGS  # Neues Paket
    HAVE_DDG = True
except ImportError:
    try:
        from duckduckgo_search import DDGS  # Altes Paket
        HAVE_DDG = True
    except ImportError:
        HAVE_DDG = False
try:
    from httpx import ReadError, ConnectTimeout
except Exception:
    class _DDGTimeout(Exception): ...
    ReadError = ConnectTimeout = _DDGTimeout
try:
    from aiohttp.client_exceptions import ClientConnectorError, ServerDisconnectedError
except Exception:
    class _NetErr(Exception): ...
    ClientConnectorError = ServerDisconnectedError = _NetErr
from dotenv import load_dotenv
from urllib.robotparser import RobotFileParser
from stream2_extraction_layer.open_data_resolver import resolve_company_domain
from stream2_extraction_layer.extraction_enhanced import (
    extract_name_enhanced,
    extract_role_with_context,
)
from learning_engine import (
    LearningEngine,
    is_mobile_number,
    is_job_posting,
)

# Suppress the noisy XML warning
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

# =========================
# NRW Städte für städtebasierte Suche
# =========================
NRW_CITIES = [
    "Köln", "Düsseldorf", "Dortmund", "Essen", "Duisburg",
    "Bochum", "Wuppertal", "Bielefeld", "Bonn", "Münster",
    "Gelsenkirchen", "Aachen", "Mönchengladbach"
]
NRW_CITIES_EXTENDED = [
    "Köln", "Düsseldorf", "Dortmund", "Essen", "Duisburg",
    "Bochum", "Wuppertal", "Bielefeld", "Bonn", "Münster",
    "Gelsenkirchen", "Mönchengladbach", "Aachen", "Krefeld", "Oberhausen",
    "Hagen", "Hamm", "Mülheim an der Ruhr", "Leverkusen", "Solingen",
    "Neuss", "Paderborn", "Recklinghausen", "Remscheid", "Moers",
    "Siegen", "Witten", "Iserlohn", "Bergisch Gladbach", "Herne"
]
# Top 40 Städte in NRW für massive Abdeckung
NRW_BIG_CITIES = [
    "Köln", "Düsseldorf", "Dortmund", "Essen", "Duisburg", "Bochum", "Wuppertal", "Bielefeld", "Bonn", "Münster",
    "Gelsenkirchen", "Mönchengladbach", "Aachen", "Chemnitz", "Krefeld", "Oberhausen", "Hagen", "Hamm", "Mülheim",
    "Leverkusen", "Solingen", "Herne", "Neuss", "Paderborn", "Bottrop", "Recklinghausen", "Bergisch Gladbach",
    "Remscheid", "Moers", "Siegen", "Gütersloh", "Witten", "Iserlohn", "Düren", "Ratingen", "Lünen", "Marl",
    "Velbert", "Minden", "Viersen"
]
# Großstädte mit stärkerer Titel-Streuung
METROPOLIS = ["Köln", "Düsseldorf", "Dortmund", "Essen", "Duisburg", "Bochum", "Bonn", "Münster"]

# Professionelle Job-Titel für LinkedIn/Xing
SALES_TITLES = [
    "Vertrieb", "Sales Manager", "Account Manager", "Außendienst", "Telesales",
    "Verkäufer", "Handelsvertreter", "Key Account Manager", "Area Sales Manager"
]
# Regionen statt nur Städte (für breite Suche)
NRW_REGIONS = [
    "nrw", "nordrhein-westfalen", "nordrhein westfalen",
    "düsseldorf", "köln", "dortmund", "essen", "duisburg",
    "bochum", "wuppertal", "bielefeld", "bonn", "münster",
    "gelsenkirchen", "mönchengladbach", "aachen", "krefeld",
    "oberhausen", "hagen", "hamm", "mülheim", "leverkusen",
    "solingen", "herne", "neuss", "paderborn", "bottrop",
    "ruhrgebiet", "rheinland", "sauerland", "münsterland", "owl",
]

# Private Mail-Provider (keine Firmen!)
PRIVATE_MAILS = '("@gmail.com" OR "@gmx.de" OR "@web.de" OR "@t-online.de" OR "@yahoo.de" OR "@outlook.com" OR "@icloud.com")'

# Handy-Nummern Muster (Masse!)
MOBILE_PATTERNS = '("017" OR "016" OR "015" OR "+49 17" OR "+49 16" OR "+49 15" OR "+4917" OR "+4916" OR "+4915")'

# --- Globales Query-Set (wird in __main__ gesetzt) ---
QUERIES: List[str] = []
DEFAULT_QUERIES: List[str] = [
    # --- GROUP A: AGENCY & PARTNER LISTS (B2B) ---
    'site:de "Unsere Vertriebspartner" "PLZ" "Mobil" -jobs',
    'site:de "Unsere Vertretungen" "Industrievertretung" "Kontakt"',
    'site:de "Vertriebsnetz" "Handelsvertretung" "Ansprechpartner"',
    'site:de "Gebietsvertretung" "PLZ" "Telefon" -stepstone -indeed',
    'site:de "Vertragshändler" "Maschinenbau" "Ansprechpartner"',
    'filetype:pdf "Vertreterliste" "Mobil" "PLZ" -bewerbung',
    'filetype:pdf "Preisliste" "Industrievertretung" "Kontakt"',
    'filetype:xlsx "Händlerverzeichnis" "Ansprechpartner" "Mobil"',
    '"Inhaltlich Verantwortlicher" "Handelsvertretung" "Mobil" -GmbH -UG',
    '"Angaben gemäß § 5 TMG" "Handelsvertretung" "Inhaber" "Mobil"',
    'inurl:impressum "Handelsvertretung" "powered by WordPress" "Mobil"',
    'site:cdh.de "Mitglied" "Kontakt"',

    # --- GROUP B: JOBSEEKERS & CVs (B2C) ---
    'filetype:pdf "Lebenslauf" "Vertrieb" "Mobil" "Wohnhaft" -muster -vorlage',
    'filetype:pdf "Curriculum Vitae" "Sales Manager" "Handy" -sample',
    'filetype:pdf "Bewerbung als" "Vertriebsmitarbeiter" "Anlagen" "Mobil"',
    'site:kleinanzeigen.de "Stellengesuch" "Vertrieb" -stellenangebot',
    'site:quoka.de "Stellengesuch" "Vertrieb" "Kontakt"',
    'site:markt.de "Jobgesuch" "Verkauf" "Mobil"',
    'site:linkedin.com/in/ "Suche neue Herausforderung" "Vertrieb" "Kontakt"',
    'site:xing.com/profile "ab sofort verfügbar" "Sales" "Mobil"',
]
NICHE_QUERIES = {
    "construction": [
        '("Werksvertretung" OR "Handelsvertretung") ("SHK" OR "Sanitär" OR "Heizung") "Gebietsleiter" "Mobil"',
        '("Bauelemente" OR "Fenster") ("Werksvertretung" OR "Fachhandelspartner") "PLZ" "Ansprechpartner"',
        'site:de ("Außendienst" OR "Vertriebsaußendienst") ("SHK" OR "Elektro") "Gebietsverkaufsleiter" "Kontakt"',
        'filetype:pdf "Fachpartnerliste" "Heizung" "Vertriebspartner" "Telefon"'
    ],
    "medical": [
        '("Medizinprodukteberater" OR "Klinikreferent") "Gebietsleiter" "Kontakt" "Mobil"',
        '("Außendienst" OR "Key Account Manager") ("Medizintechnik" OR "Medical Devices") "Region" "Deutschland"',
        'filetype:pdf "Lebenslauf" "Medizinprodukteberater" "Geburtsdatum" "Mobil"',
        'site:de "Unsere Außendienstmitarbeiter" "Medizintechnik" "Ansprechpartner"'
    ],
    "food": [
        '("Distributorenliste" OR "Distributor list") ("Food" OR "Lebensmittel") "Deutschland"',
        '("Großhandel" OR "Großhändler") "Gastronomie" "Vertriebspartner" "Kontakt"',
        '("Food Startup" OR "Getränke") "sucht Vertrieb" "Handelspartner"',
        'filetype:pdf "Vertriebspartner" "Getränke" "Kontakt" "Deutschland"'
    ]
}
# Freelancer & Modern Sales
FREELANCE_QUERIES = [
    # 1. Freelancer Portals (Scraping Profiles)
    'site:freelancermap.de "Vertrieb" "Deutschland" "verfügbar" -projekt',
    'site:freelance.de "Vertriebsfreelancer" "Kaltakquise" "Kontakt"',
    'site:freelancermap.de "Telesales" "Muttersprache Deutsch" "Mobil"',

    # 2. Modern Sales / Closer (High Ticket)
    '("High Ticket Closer" OR "Closer") "deutschsprachig" "Provision" "Kontakt"',
    '("Setter" OR "Appointment Setter") "Remote" "Vertrieb" "Gesuch"',
    '"Remote Vertrieb" "SaaS" "deutschsprachig" "Provision"',

    # 3. Trade Fair Lists (Germany)
    'site:.de filetype:pdf "Ausstellerverzeichnis" "Halle" "Kontakt" -2019 -2020',
    'site:.de filetype:xls "Ausstellerliste" "Vertrieb" "Telefon" "Deutschland"',
    '"Ausstellerverzeichnis" "Fachmesse" "Deutschland" "Mobil" filetype:pdf'
]
GUERILLA_QUERIES = [
    # 1. Insolvency & Layoffs (Trigger Events)
    '("Entlassungswelle" OR "Stellenabbau") "Vertrieb" "Deutschland" "2024" -news',
    '("Insolvenz" OR "Geschäftsaufgabe") "Vertriebsteam" "suche Job" site:linkedin.com',
    'site:kununu.com "Kündigung" "Vertrieb" "Bewertung" "2025"',

    # 2. Frustration & Advice (Community Hacking)
    'site:reddit.com ("Vertrieb" OR "Sales") "gekündigt" "was tun"',
    'site:wiwi-treff.de "Vertrieb" "Kündigung" "Wechsel" -stellenmarkt',
    'site:gutefrage.net "Vertrieb" "gekündigt" "neuer Job"',

    # 3. Career Changers (Quereinsteiger)
    '("Suche Job ohne Ausbildung" OR "Quereinstieg") "Vertrieb" "schnell Geld" site:.de',
    'site:urbia.de "Job ohne Ausbildung" "Vertrieb" "Teilzeit"',

    # 4. Social Public Posts (Real-time Signals)
    'site:facebook.com "Ich suche Arbeit" "Vertrieb" "bitte melden"',
    'site:instagram.com "Suche Job" "Vertrieb" "DM me"',
    'site:tiktok.com "Suche Job" "Sales" "Deutschland"',
    '("Job gesucht" OR "Arbeit gesucht") "Vertrieb" "#jobgesucht" site:.de'
]
# Flatten the niche queries into the main list
for industry_dorks in NICHE_QUERIES.values():
    DEFAULT_QUERIES.extend(industry_dorks)
DEFAULT_QUERIES.extend(FREELANCE_QUERIES)
DEFAULT_QUERIES.extend(GUERILLA_QUERIES)

# === Export-Felder (CSV/XLSX) ===
ENH_FIELDS = [
    "name","rolle","email","telefon","quelle","score","tags","region",
    "role_guess","lead_type","salary_hint","commission_hint","opening_line",
    "ssl_insecure","company_name","company_size","hiring_volume",
    "industry","recency_indicator","location_specific",
    "confidence_score","last_updated","data_quality",
    "phone_type","whatsapp_link","private_address","social_profile_url",
    "ai_category","ai_summary",
    "experience_years","skills","availability","current_status","industries","location","profile_text",
    # New candidate-focused fields
    "candidate_status","mobility","industries_experience","source_type",
    "profile_url","cv_url","contact_preference","last_activity","name_validated"
]

def export_xlsx(filename: str, rows=None):
    con = db(); cur = con.cursor()
    cur.execute("SELECT * FROM leads")
    data = cur.fetchall()
    con.close()
    if not data:
        return
    df = pd.DataFrame(data, columns=data[0].keys())
    df.to_excel(filename, index=False)


# Warnungen zu unsicherem SSL dämpfen (bewusstes Fallback via Flag)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =========================
# Konfiguration & Globals
# =========================

load_dotenv(override=True)

USER_AGENT = "Mozilla/5.0 (compatible; VertriebFinder/2.3; +https://example.com)"
DEFAULT_CSV = "vertrieb_kontakte.csv"
DEFAULT_XLSX = "vertrieb_kontakte.xlsx"
DB_PATH = os.getenv("SCRAPER_DB", "scraper.db")

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
GCS_API_KEY    = os.getenv("GCS_API_KEY", "")
GCS_CX_RAW     = os.getenv("GCS_CX", "")
BING_API_KEY   = os.getenv("BING_API_KEY", "")

HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", "10"))  # Reduced to 10s for cost control
MAX_FETCH_SIZE = int(os.getenv("MAX_FETCH_SIZE", str(2 * 1024 * 1024)))  # 2MB default

POOL_SIZE = int(os.getenv("POOL_SIZE", "12"))  # (historisch; wird in Async-Version nicht mehr genutzt)

ALLOW_PDF = (os.getenv("ALLOW_PDF", "0") == "1")
ALLOW_INSECURE_SSL = (os.getenv("ALLOW_INSECURE_SSL", "1") == "1")

# Neue Async-ENV
ASYNC_LIMIT = int(os.getenv("ASYNC_LIMIT", "35"))          # globale max. gleichzeitige Requests
ASYNC_PER_HOST = int(os.getenv("ASYNC_PER_HOST", "3"))     # pro Host
HTTP2_ENABLED = (os.getenv("HTTP2", "1") == "1")
USE_TOR = False

# Proxy environment variables to clear for nuclear cleanup
PROXY_ENV_VARS = [
    "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY",
    "http_proxy", "https_proxy", "all_proxy",
    "FTP_PROXY", "ftp_proxy", "SOCKS_PROXY", "socks_proxy"
]

NRW_CITIES = ["Köln", "Düsseldorf", "Dortmund", "Essen", "Duisburg", "Bochum", "Wuppertal", "Bielefeld", "Bonn", "Münster"]

ENABLE_KLEINANZEIGEN = (os.getenv("ENABLE_KLEINANZEIGEN", "1") == "1")
KLEINANZEIGEN_MAX_RESULTS = int(os.getenv("KLEINANZEIGEN_MAX_RESULTS", "20"))

# Telefonbuch Enrichment Config
TELEFONBUCH_ENRICHMENT_ENABLED = (os.getenv("TELEFONBUCH_ENRICHMENT_ENABLED", "1") == "1")
TELEFONBUCH_STRICT_MODE = (os.getenv("TELEFONBUCH_STRICT_MODE", "1") == "1")
TELEFONBUCH_RATE_LIMIT = float(os.getenv("TELEFONBUCH_RATE_LIMIT", "3.0"))
TELEFONBUCH_CACHE_DAYS = int(os.getenv("TELEFONBUCH_CACHE_DAYS", "7"))
TELEFONBUCH_MOBILE_ONLY = (os.getenv("TELEFONBUCH_MOBILE_ONLY", "1") == "1")
# Direct crawl URLs for Kleinanzeigen Stellengesuche (bypassing Google)
DIRECT_CRAWL_URLS = [
    # Stellengesuche NRW - Vertrieb/Sales
    "https://www.kleinanzeigen.de/s-stellengesuche/nordrhein-westfalen/vertrieb/k0c107l929",
    "https://www.kleinanzeigen.de/s-stellengesuche/nordrhein-westfalen/sales/k0c107l929", 
    "https://www.kleinanzeigen.de/s-stellengesuche/nordrhein-westfalen/verkauf/k0c107l929",
    "https://www.kleinanzeigen.de/s-stellengesuche/nordrhein-westfalen/aussendienst/k0c107l929",
    "https://www.kleinanzeigen.de/s-stellengesuche/nordrhein-westfalen/kundenberater/k0c107l929",
    "https://www.kleinanzeigen.de/s-stellengesuche/nordrhein-westfalen/handelsvertreter/k0c107l929",
    
    # Bundesweit (mehr Volumen)
    "https://www.kleinanzeigen.de/s-stellengesuche/vertrieb/k0c107",
    "https://www.kleinanzeigen.de/s-stellengesuche/sales/k0c107",
    "https://www.kleinanzeigen.de/s-stellengesuche/verkauf/k0c107",
    "https://www.kleinanzeigen.de/s-stellengesuche/handelsvertreter/k0c107",
    "https://www.kleinanzeigen.de/s-stellengesuche/akquise/k0c107",
    "https://www.kleinanzeigen.de/s-stellengesuche/telesales/k0c107",
    "https://www.kleinanzeigen.de/s-stellengesuche/call-center/k0c107",
]

# Markt.de Stellengesuche URLs
MARKT_DE_URLS = [
    # NRW
    "https://www.markt.de/stellengesuche/nordrhein-westfalen/vertrieb/",
    "https://www.markt.de/stellengesuche/nordrhein-westfalen/sales/",
    "https://www.markt.de/stellengesuche/nordrhein-westfalen/verkauf/",
    "https://www.markt.de/stellengesuche/nordrhein-westfalen/kundenberater/",
    # Bundesweit
    "https://www.markt.de/stellengesuche/vertrieb/",
    "https://www.markt.de/stellengesuche/sales/",
    "https://www.markt.de/stellengesuche/handelsvertreter/",
]

# Quoka.de Stellengesuche URLs
QUOKA_DE_URLS = [
    # NRW Städte
    "https://www.quoka.de/stellengesuche/duesseldorf/",
    "https://www.quoka.de/stellengesuche/koeln/",
    "https://www.quoka.de/stellengesuche/dortmund/",
    "https://www.quoka.de/stellengesuche/essen/",
    # Kategorien
    "https://www.quoka.de/stellengesuche/vertrieb-verkauf/",
    "https://www.quoka.de/stellengesuche/kundenservice/",
]

# Kalaydo.de Stellengesuche URLs (NRW-fokussiert!)
KALAYDO_DE_URLS = [
    # Kalaydo ist stark in NRW
    "https://www.kalaydo.de/stellengesuche/nordrhein-westfalen/",
    "https://www.kalaydo.de/stellengesuche/koeln/",
    "https://www.kalaydo.de/stellengesuche/duesseldorf/",
    "https://www.kalaydo.de/stellengesuche/bonn/",
    "https://www.kalaydo.de/stellengesuche/aachen/",
]

# Meinestadt.de Stellengesuche URLs
MEINESTADT_DE_URLS = [
    # Top NRW Städte
    "https://www.meinestadt.de/duesseldorf/stellengesuche",
    "https://www.meinestadt.de/koeln/stellengesuche",
    "https://www.meinestadt.de/dortmund/stellengesuche",
    "https://www.meinestadt.de/essen/stellengesuche",
    "https://www.meinestadt.de/bochum/stellengesuche",
    "https://www.meinestadt.de/wuppertal/stellengesuche",
    "https://www.meinestadt.de/muenster/stellengesuche",
    "https://www.meinestadt.de/bonn/stellengesuche",
]

# Direct crawl source configuration
DIRECT_CRAWL_SOURCES = {
    "kleinanzeigen": True,
    "markt_de": True,
    "quoka": True,
    "kalaydo": True,
    "meinestadt": True,
}

# =========================
# Candidate-Focused Constants
# =========================

# Source Types - Extensive list for candidate-focused scraping
SOURCE_TYPES = {
    # PRIMÄR - Direkte Jobsuche
    "stellengesuch_kleinanzeigen": "Kleinanzeigen.de Stellengesuch",
    "stellengesuch_markt": "Markt.de Stellengesuch",
    "stellengesuch_quoka": "Quoka Stellengesuch",
    "stellengesuch_kalaydo": "Kalaydo Stellengesuch",
    
    # SOCIAL MEDIA - Aktiv suchend
    "linkedin_opentowork": "LinkedIn #OpenToWork",
    "linkedin_seeking": "LinkedIn 'seeking opportunities'",
    "xing_offen": "Xing 'offen für Angebote'",
    "xing_jobsuche": "Xing 'auf Jobsuche'",
    "facebook_jobpost": "Facebook Jobsuche-Post",
    "facebook_gruppe": "Facebook Vertriebs-Gruppe",
    "instagram_bio": "Instagram 'looking for work'",
    "tiktok_karriere": "TikTok Karriere-Content",
    
    # MESSENGER GRUPPEN
    "telegram_vertrieb": "Telegram Vertriebsgruppe",
    "telegram_jobs": "Telegram Jobs-Kanal",
    "whatsapp_gruppe": "WhatsApp Vertriebsgruppe",
    "discord_karriere": "Discord Karriere-Server",
    
    # FOREN & COMMUNITIES
    "reddit_arbeitsleben": "Reddit r/arbeitsleben",
    "reddit_fragreddit": "Reddit Karrierefrage",
    "gutefrage_job": "Gutefrage Jobsuche",
    "wiwi_treff": "WiWi-Treff Forum",
    
    # FREELANCER PORTALE
    "freelancermap": "Freelancermap Profil",
    "freelance_de": "Freelance.de Profil",
    "gulp": "GULP Freiberufler",
    "fiverr_sales": "Fiverr Sales-Services",
    "upwork_german": "Upwork deutschsprachig",
    
    # LEBENSLAUF/CV
    "cv_pdf_public": "Öffentliches PDF (Lebenslauf)",
    "cv_stepstone": "StepStone Kandidatenprofil",
    "cv_indeed": "Indeed Lebenslauf",
    
    # KRISENUNTERNEHMEN (Abwerbung)
    "kununu_krise": "Kununu Bewertung (Krise)",
    "linkedin_layoff": "LinkedIn Entlassungs-Post",
    "insolvenz_news": "Insolvenz-Meldung",
    
    # BRANCHEN-SPEZIFISCH
    "solar_vertrieb": "Solar-Branche Vertrieb",
    "telekom_sales": "Telekom-Branche Sales",
    "versicherung_makler": "Versicherungs-Makler",
    "energie_berater": "Energieberater",
    "immobilien_makler": "Immobilien-Makler",
    
    # QUEREINSTEIGER-POTENZIAL
    "gastro_wechsler": "Gastro → Vertrieb Wechsler",
    "einzelhandel_pro": "Einzelhandel mit Sales-Talent",
    "mlm_aussteiger": "MLM/Network Marketing Aussteiger",
    "callcenter_erfahren": "Call-Center mit Erfahrung",
    "promotion_aktiv": "Promoter mit Ambitionen",
    "door2door_veteran": "D2D Erfahrung",
}

# Required vs Optional Fields
REQUIRED_FIELDS = {
    "name": "Echter menschlicher Name (KI-geprüft!)",
    "telefon": "Handynummer (015x, 016x, 017x)",
}

OPTIONAL_FIELDS = {
    "email": "E-Mail (bevorzugt privat)",
    "whatsapp_link": "WhatsApp-Link",
    "profile_url": "Social-Media Profil",
    "skills": "Fähigkeiten",
    "experience_years": "Berufserfahrung",
    "availability": "Verfügbarkeit",
    "industries_experience": "Branchenerfahrung",
    "location": "Standort/Region",
    "mobility": "Mobilität",
    "cv_url": "Lebenslauf-URL",
}

# Hidden Gems - Quereinsteiger with sales potential
HIDDEN_GEMS_PATTERNS = {
    "gastro_talent": {
        "keywords": ["restaurant", "gastronomie", "kellner", "service"],
        "positive": ["kundenorientiert", "stressresistent", "kommunikativ"],
        "reason": "Gastro-Erfahrung = Kundenkontakt + Belastbarkeit"
    },
    "einzelhandel_pro": {
        "keywords": ["einzelhandel", "verkäufer", "filiale", "retail"],
        "positive": ["umsatz", "beratung", "kasse", "kundenservice"],
        "reason": "Einzelhandel = Verkaufserfahrung + Kundenumgang"
    },
    "callcenter_veteran": {
        "keywords": ["callcenter", "call center", "kundenservice", "hotline"],
        "positive": ["outbound", "telefonverkauf", "telesales"],
        "reason": "Call-Center = Telefonakquise-Erfahrung"
    },
    "mlm_refugee": {
        "keywords": ["network marketing", "mlm", "direktvertrieb", "tupperware"],
        "positive": ["selbstständig", "provision", "akquise"],
        "reason": "MLM-Aussteiger = Kaltakquise + Hunger auf echten Job"
    },
    "promotion_hustler": {
        "keywords": ["promotion", "promoter", "messe", "events"],
        "positive": ["kommunikativ", "überzeugend", "präsentation"],
        "reason": "Promoter = Ansprache-Erfahrung + keine Scheu"
    },
    "door2door_warrior": {
        "keywords": ["d2d", "door to door", "haustür", "außendienst"],
        "positive": ["provision", "abschlussstark", "hartnäckig"],
        "reason": "D2D-Erfahrung = Härtester Vertrieb überhaupt"
    },
    "fitness_coach": {
        "keywords": ["fitness", "personal trainer", "coach", "studio"],
        "positive": ["motivierend", "zielorientiert", "membership"],
        "reason": "Fitness = Verkauf + Überzeugungskraft"
    },
}

# Default mode is now candidates
DEFAULT_MODE = "candidates"

# Hard blocks - never save these
UNIVERSAL_HARD_BLOCKS = [
    r'\(m/w/d\)', r'\(w/m/d\)', r'\(d/m/w\)',
    r'jetzt bewerben', r'bewerbung an:',
    r'stellenanzeige', r'vakanz',
    r'recruiting@', r'bewerbung@', r'karriere@',
    r'hr-manager', r'talent acquisition',
    r'crypto.*gewinn', r'bitcoin.*investier',
    r'casino', r'dating',
]

# Always allow in candidate mode
CANDIDATE_ALWAYS_ALLOW = [
    "suche job", "suche arbeit", "suche stelle",
    "suche neue herausforderung", "stellengesuch",
    "#opentowork", "offen für angebote", "offen für neues",
    "arbeitslos", "arbeitssuchend", "freigestellt", "gekündigt",
    "verfügbar ab", "ab sofort verfügbar", "wechselwillig",
    "auf jobsuche", "biete meine dienste", "biete mich an",
    "freiberuflich verfügbar",
]

# Candidate export fields
CANDIDATE_EXPORT_FIELDS = [
    # Pflicht
    "name", "telefon",
    
    # Kandidaten-spezifisch
    "candidate_status", "source_type", "lead_type",
    "experience_years", "skills", "industries_experience",
    "availability", "mobility",
    
    # Kontakt
    "email", "whatsapp_link", "profile_url", "contact_preference",
    
    # Meta
    "score", "region", "last_activity", "last_updated",
    "name_validated",
]

# =========================
# Mode Configurations
# =========================
MODE_CONFIGS = {
    "standard": {
        "description": "Normaler Betrieb",
        "deep_crawl": True,
        "learning_enabled": False,
        "async_limit": 35,
        "request_delay": 2.5,
        "max_retries": 2,
        "snippet_priority": False,
        "save_patterns": False
    },
    "learning": {
        "description": "Lernt aus erfolgreichen Extraktionen",
        "deep_crawl": True,
        "learning_enabled": True,
        "async_limit": 30,
        "request_delay": 3.0,
        "max_retries": 3,
        "snippet_priority": True,
        "save_patterns": True,
        "pattern_analysis": True,
        "success_tracking": True,
        "domain_scoring": True,
        "query_optimization": True
    },
    "aggressive": {
        "description": "Maximale Geschwindigkeit, mehr Requests",
        "deep_crawl": True,
        "learning_enabled": False,
        "async_limit": 75,
        "request_delay": 1.0,
        "max_retries": 1,
        "snippet_priority": False,
        "save_patterns": False,
        "follow_links": True,
        "crawl_depth": 3
    },
    "snippet_only": {
        "description": "Nur Snippet-Extraktion, kein Deep-Crawl",
        "deep_crawl": False,
        "learning_enabled": False,
        "async_limit": 50,
        "request_delay": 1.5,
        "max_retries": 1,
        "snippet_priority": True,
        "save_patterns": False
    }
}

# Global mode configuration (will be set in main)
ACTIVE_MODE_CONFIG = None

# === Rotation: Proxies & User-Agents ===
def _env_list(val: str, sep: str) -> list[str]:
    return [x.strip() for x in (val or "").split(sep) if x.strip()]

PROXY_POOL = _env_list(os.getenv("PROXY_POOL", ""), ",")  # "http://user:pass@ip1:port, http://ip2:port"
UA_POOL_ENV = _env_list(os.getenv("UA_POOL", ""), "|")

# Realistische Desktop-UAs (falls UA_POOL leer)
UA_POOL_DEFAULT = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_6) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edg/121.0.0.0 Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.6167.140 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:121.0) Gecko/20100101 Firefox/121.0",
]
UA_POOL = UA_POOL_ENV if UA_POOL_ENV else UA_POOL_DEFAULT


# Basis-MinScore aus .env; wird pro Query-Runde dynamisch angepasst
MIN_SCORE_ENV = int(os.getenv("MIN_SCORE", "40"))

MAX_PER_DOMAIN = int(os.getenv("MAX_PER_DOMAIN", "5"))
INTERNAL_DEPTH_PER_DOMAIN = int(os.getenv("INTERNAL_DEPTH_PER_DOMAIN", "10"))

SLEEP_BETWEEN_QUERIES = float(os.getenv("SLEEP_BETWEEN_QUERIES", "2.7"))
SEED_FORCE = (os.getenv("SEED_FORCE", "0") == "1")


def get_performance_params():
    """
    Fetch current performance parameters from dashboard API.
    Returns effective parameters based on performance mode and system load.
    """
    try:
        import requests
        response = requests.get('http://127.0.0.1:5056/api/performance/effective', timeout=2)
        if response.status_code == 200:
            data = response.json()
            return data.get('params', {})
    except Exception:
        pass
    
    # Fallback to environment variables/defaults
    return {
        'threads': int(os.getenv("THREADS", "4")),
        'async_limit': int(os.getenv("ASYNC_LIMIT", "35")),
        'batch_size': int(os.getenv("BATCH_SIZE", "20")),
        'request_delay': float(os.getenv("SLEEP_BETWEEN_QUERIES", "2.7"))
    }

# -------------- Logging --------------
def log(level:str, msg:str, **ctx):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ctx_str = (" " + json.dumps(ctx, ensure_ascii=False)) if ctx else ""
    line = f"[{ts}] [{level.upper():7}] {msg}{ctx_str}"
    print(line, flush=True)
    if UILOGQ is not None:
        try: UILOGQ.put_nowait(line)
        except Exception: pass

def die(msg, **ctx):
    log("fatal", msg, **ctx); sys.exit(1)

# -------------- UI State --------------
UILOGQ: Optional[queue.Queue] = None
RUN_FLAG = {"running": False, "force": False}

# -------------- Dataclass CFG --------------
@dataclass
class ScraperConfig:
    min_score: int = MIN_SCORE_ENV
    max_results_per_domain: int = MAX_PER_DOMAIN
    request_timeout: int = HTTP_TIMEOUT
    allow_pdf: bool = ALLOW_PDF
    allow_insecure_ssl: bool = ALLOW_INSECURE_SSL
    pool_size: int = POOL_SIZE
    internal_depth_per_domain: int = INTERNAL_DEPTH_PER_DOMAIN

CFG = ScraperConfig()

# =========================
# DB-Layer (SQLite)
# =========================

_DB_READY = False  # einmaliges Schema-Setup pro Prozess
_LEARNING_ENGINE: Optional[LearningEngine] = None  # Global learning engine instance

LEAD_FIELDS = [
    "name","rolle","email","telefon","quelle","score","tags","region",
    "role_guess","lead_type","salary_hint","commission_hint","opening_line","ssl_insecure",
    "company_name","company_size","hiring_volume","industry",
    "recency_indicator","location_specific","confidence_score","last_updated",
    "data_quality","phone_type","whatsapp_link","private_address","social_profile_url",
    "ai_category","ai_summary",
    # New candidate-focused fields
    "candidate_status","experience_years","availability","mobility","skills",
    "industries_experience","source_type","profile_url","cv_url",
    "contact_preference","last_activity","name_validated"
]

def _ensure_schema(con: sqlite3.Connection) -> None:
    """
    Stellt sicher, dass alle Tabellen existieren, fehlende Spalten nachgezogen
    und Indizes korrekt angelegt sind. Idempotent.
    """
    cur = con.cursor()
    # Basis-Tabellen
    cur.executescript("""
    PRAGMA journal_mode = WAL;

    CREATE TABLE IF NOT EXISTS leads(
      id INTEGER PRIMARY KEY,
      name TEXT,
      rolle TEXT,
      email TEXT,
      telefon TEXT,
      quelle TEXT,
      score INT,
      tags TEXT,
      region TEXT,
      role_guess TEXT,
      lead_type TEXT,
      salary_hint TEXT,
      commission_hint TEXT,
      opening_line TEXT,
      ssl_insecure TEXT,
      company_name TEXT,
      company_size TEXT,
      hiring_volume TEXT,
      industry TEXT,
      recency_indicator TEXT,
      location_specific TEXT,
      confidence_score INT,
      last_updated TEXT,
      data_quality INT,
      phone_type TEXT,
      whatsapp_link TEXT,
      private_address TEXT,
      social_profile_url TEXT,
      ai_category TEXT,
      ai_summary TEXT
      -- neue Spalten werden unten per ALTER TABLE nachgezogen
    );

    CREATE TABLE IF NOT EXISTS runs(
      id INTEGER PRIMARY KEY,
      started_at TEXT,
      finished_at TEXT,
      status TEXT,
      links_checked INTEGER,
      leads_new INTEGER
    );

    CREATE TABLE IF NOT EXISTS queries_done(
      q TEXT PRIMARY KEY,
      last_run_id INTEGER,
      ts TEXT
    );

    CREATE TABLE IF NOT EXISTS urls_seen(
      url TEXT PRIMARY KEY,
      first_run_id INTEGER,
      ts TEXT
    );

    CREATE TABLE IF NOT EXISTS telefonbuch_cache(
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      name TEXT NOT NULL,
      city TEXT NOT NULL,
      query_hash TEXT UNIQUE,
      results_json TEXT,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    """)
    con.commit()

    # Fehlende Spalten in leads nachziehen (z. B. nach Update)
    cur.execute("PRAGMA table_info(leads)")
    existing_cols = {row[1] for row in cur.fetchall()}  # 2. Feld = Spaltenname

    # phone_type / whatsapp_link sicherstellen
    if "phone_type" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN phone_type TEXT")
    if "whatsapp_link" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN whatsapp_link TEXT")
    try:
        cur.execute("ALTER TABLE leads ADD COLUMN lead_type TEXT")
    except Exception:
        pass
    if "private_address" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN private_address TEXT")
    if "social_profile_url" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN social_profile_url TEXT")
    if "ai_category" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN ai_category TEXT")
    if "ai_summary" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN ai_summary TEXT")
    
    # Candidate-specific columns (existing)
    if "experience_years" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN experience_years INTEGER")
    if "skills" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN skills TEXT")  # JSON array
    if "availability" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN availability TEXT")
    if "current_status" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN current_status TEXT")
    if "industries" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN industries TEXT")  # JSON array
    if "location" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN location TEXT")
    if "profile_text" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN profile_text TEXT")
    
    # New candidate-focused columns
    if "candidate_status" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN candidate_status TEXT")  # aktiv_suchend, passiv_offen, etc.
    if "mobility" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN mobility TEXT")  # remote, hybrid, vor_ort, reisebereit
    if "industries_experience" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN industries_experience TEXT")  # JSON array
    if "source_type" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN source_type TEXT")  # Detailed source type
    if "profile_url" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN profile_url TEXT")  # Social media profile
    if "cv_url" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN cv_url TEXT")  # CV/Resume PDF URL
    if "contact_preference" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN contact_preference TEXT")  # whatsapp, telefon, email, telegram
    if "last_activity" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN last_activity TEXT")  # When candidate was last active
    if "name_validated" not in existing_cols:
        cur.execute("ALTER TABLE leads ADD COLUMN name_validated INTEGER")  # 1 = AI-validated real name
    con.commit()

    # Indizes (partielle UNIQUE nur wenn Werte vorhanden)
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS ux_leads_email
        ON leads(email) WHERE email IS NOT NULL AND email <> ''
    """)
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS ux_leads_tel
        ON leads(telefon) WHERE telefon IS NOT NULL AND telefon <> ''
    """)
    con.commit()

def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    global _DB_READY, _LEARNING_ENGINE
    if not _DB_READY:
        _ensure_schema(con)
        # Initialize dashboard schema
        try:
            from dashboard.db_schema import ensure_dashboard_schema, initialize_default_search_modes, initialize_default_settings
            ensure_dashboard_schema(con)
            initialize_default_search_modes(con)
            initialize_default_settings(con)
        except ImportError as e:
            # Dashboard module not available - this is expected if dashboard deps not installed
            import sys
            if '--verbose' in sys.argv or os.getenv('DEBUG'):
                print(f"Note: Dashboard module not available: {e}")
        except Exception as e:
            # Log other errors but don't fail
            print(f"Warning: Could not initialize dashboard schema: {e}")
        _DB_READY = True
    # Initialize learning engine on first DB access
    if _LEARNING_ENGINE is None:
        _LEARNING_ENGINE = LearningEngine(DB_PATH)
    return con

def init_db():
    # bleibt als expliziter Initialisierer erhalten (macht intern dasselbe)
    con = db()
    con.close()

def get_learning_engine() -> Optional[LearningEngine]:
    """Get the global learning engine instance."""
    global _LEARNING_ENGINE
    if _LEARNING_ENGINE is None:
        _LEARNING_ENGINE = LearningEngine(DB_PATH)
    return _LEARNING_ENGINE

def init_mode(mode: str) -> Dict[str, Any]:
    """
    Initialize the operating mode and apply its configuration.
    
    IMPORTANT: This function must be called during startup before any async operations begin,
    as it modifies global configuration variables.
    
    Args:
        mode: Mode name (standard, learning, aggressive, snippet_only)
    
    Returns:
        Mode configuration dictionary
    """
    global ACTIVE_MODE_CONFIG, ASYNC_LIMIT, _LEARNING_ENGINE
    
    config = MODE_CONFIGS.get(mode, MODE_CONFIGS["standard"])
    ACTIVE_MODE_CONFIG = config
    
    # Apply mode-specific settings (must be done before any async operations start)
    ASYNC_LIMIT = config.get("async_limit", 35)
    
    # Initialize learning engine if learning mode is enabled
    if config.get("learning_enabled"):
        if _LEARNING_ENGINE is None:
            _LEARNING_ENGINE = LearningEngine(DB_PATH)
        log("INFO", "Learning-Modus aktiviert", stats=_LEARNING_ENGINE.get_pattern_stats())
    
    log("INFO", "Betriebsmodus initialisiert",
        mode=mode,
        description=config.get("description"),
        async_limit=config.get("async_limit"),
        learning_enabled=config.get("learning_enabled", False)
    )
    
    return config

def migrate_db_unique_indexes():
    """
    Fallback für sehr alte Schemas mit harten UNIQUE-Constraints.
    Nur ausführen, wenn Einfügen weiterhin scheitert.
    """
    con = db(); cur = con.cursor()
    try:
        cur.execute("INSERT OR IGNORE INTO leads (name,email,telefon) VALUES (?,?,?)",
                    ("_probe_", "", ""))
        con.commit()
    except Exception:
        cur.executescript("""
        BEGIN TRANSACTION;
        CREATE TABLE leads_new(
          id INTEGER PRIMARY KEY,
          name TEXT, rolle TEXT, email TEXT, telefon TEXT, quelle TEXT,
          score INT, tags TEXT, region TEXT, role_guess TEXT, salary_hint TEXT,
          commission_hint TEXT, opening_line TEXT, ssl_insecure TEXT,
          company_name TEXT, company_size TEXT, hiring_volume TEXT, industry TEXT,
          recency_indicator TEXT, location_specific TEXT, confidence_score INT,
          last_updated TEXT, data_quality INT,
          phone_type TEXT, whatsapp_link TEXT,
          lead_type TEXT,
          private_address TEXT, social_profile_url TEXT,
          experience_years INTEGER, skills TEXT, availability TEXT,
          current_status TEXT, industries TEXT, location TEXT, profile_text TEXT
        );

        INSERT INTO leads_new (
          id,name,rolle,email,telefon,quelle,score,tags,region,role_guess,salary_hint,
          commission_hint,opening_line,ssl_insecure,company_name,company_size,hiring_volume,
          industry,recency_indicator,location_specific,confidence_score,last_updated,data_quality,
          phone_type,whatsapp_link,lead_type,private_address,social_profile_url,
          experience_years,skills,availability,current_status,industries,location,profile_text
        )
        SELECT
          id,name,rolle,email,telefon,quelle,score,tags,region,role_guess,salary_hint,
          commission_hint,opening_line,ssl_insecure,company_name,company_size,hiring_volume,
          industry,recency_indicator,location_specific,confidence_score,last_updated,data_quality,
          '' AS phone_type, '' AS whatsapp_link, '' AS lead_type,
          '' AS private_address, '' AS social_profile_url,
          NULL AS experience_years, '' AS skills, '' AS availability,
          '' AS current_status, '' AS industries, '' AS location, '' AS profile_text
        FROM leads;

        DROP TABLE leads;
        ALTER TABLE leads_new RENAME TO leads;
        COMMIT;
        """)
        con.commit()
        # Indizes nachziehen
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_leads_email
            ON leads(email) WHERE email IS NOT NULL AND email <> ''
        """)
        cur.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS ux_leads_tel
            ON leads(telefon) WHERE telefon IS NOT NULL AND telefon <> ''
        """)
        con.commit()
    finally:
        con.close()

def is_query_done(q: str) -> bool:
    con = db(); cur = con.cursor()
    cur.execute("SELECT 1 FROM queries_done WHERE q=?", (q,))
    hit = cur.fetchone()
    con.close()
    return bool(hit)

def mark_query_done(q: str, run_id: int):
    con = db(); cur = con.cursor()
    cur.execute(
        "INSERT OR REPLACE INTO queries_done(q,last_run_id,ts) VALUES(?,?,datetime('now'))",
        (q, run_id)
    )
    con.commit(); con.close()

_seen_urls_cache: set[str] = set()

def mark_url_seen(url: str, run_id: int):
    global _seen_urls_cache
    con = db(); cur = con.cursor()
    cur.execute(
        "INSERT OR IGNORE INTO urls_seen(url,first_run_id,ts) VALUES(?,?,datetime('now'))",
        (url, run_id)
    )
    con.commit(); con.close()
    _seen_urls_cache.add(_normalize_for_dedupe(url))

def url_seen(url: str) -> bool:
    norm = _normalize_for_dedupe(url)
    if norm in _seen_urls_cache:
        return True
    con = db(); cur = con.cursor()
    cur.execute("SELECT 1 FROM urls_seen WHERE url=?", (url,))
    hit = cur.fetchone()
    con.close()
    if hit:
        _seen_urls_cache.add(norm)
    return bool(hit)

def _url_seen_fast(url: str) -> bool:
    return _normalize_for_dedupe(url) in _seen_urls_cache


# =========================
# Telefonbuch Enrichment
# =========================

async def get_cached_telefonbuch_result(name: str, city: str) -> Optional[List[Dict]]:
    """Prüft ob Ergebnis im Cache ist (max 7 Tage alt)"""
    if not name or not city:
        return None
    
    query_hash = hashlib.md5(f"{name.lower()}:{city.lower()}".encode()).hexdigest()
    
    con = db()
    cur = con.cursor()
    cur.execute("""
        SELECT results_json, created_at 
        FROM telefonbuch_cache 
        WHERE query_hash = ?
    """, (query_hash,))
    row = cur.fetchone()
    con.close()
    
    if not row:
        return None
    
    results_json, created_at = row
    
    # Check if cache is still valid (max 7 days)
    try:
        from datetime import datetime, timedelta
        cache_time = datetime.fromisoformat(created_at)
        age = datetime.now() - cache_time
        if age > timedelta(days=TELEFONBUCH_CACHE_DAYS):
            return None  # Cache expired
    except Exception:
        return None
    
    try:
        results = json.loads(results_json) if results_json else []
        log("debug", "Telefonbuch-Cache Hit", name=name, city=city)
        return results
    except Exception:
        return None


async def cache_telefonbuch_result(name: str, city: str, results: List[Dict]):
    """Speichert Ergebnis im Cache"""
    if not name or not city:
        return
    
    query_hash = hashlib.md5(f"{name.lower()}:{city.lower()}".encode()).hexdigest()
    results_json = json.dumps(results, ensure_ascii=False)
    
    con = db()
    cur = con.cursor()
    cur.execute("""
        INSERT OR REPLACE INTO telefonbuch_cache (name, city, query_hash, results_json, created_at)
        VALUES (?, ?, ?, ?, datetime('now'))
    """, (name, city, query_hash, results_json))
    con.commit()
    con.close()


# Rate limiter for telefonbuch requests
class _TelefonbuchRateLimiter:
    def __init__(self, interval: float = 3.0):
        self.interval = interval
        self.last_request = 0.0
        self.lock = asyncio.Lock()
    
    async def __aenter__(self):
        async with self.lock:
            now = time.time()
            elapsed = now - self.last_request
            if elapsed < self.interval:
                wait_time = self.interval - elapsed
                await asyncio.sleep(wait_time)
            self.last_request = time.time()
    
    async def __aexit__(self, *args):
        pass

_telefonbuch_rate = _TelefonbuchRateLimiter(interval=TELEFONBUCH_RATE_LIMIT)


async def query_dasoertliche(name: str, city: str) -> List[Dict]:
    """
    Führt eine Suche auf dasoertliche.de durch.
    
    URL-Format: https://www.dasoertliche.de/?kw={name}&ci={city}
    
    Extrahiert aus HTML:
        - name: Name des Eintrags
        - phone: Telefonnummer
        - address: Adresse
        - city: Stadt
    
    Returns:
        Liste von Treffern als Dicts
    """
    if not name or not city:
        return []
    
    # Build search URL
    params = urllib.parse.urlencode({"kw": name, "ci": city})
    url = f"https://www.dasoertliche.de/?{params}"
    
    log("debug", "Telefonbuch-Query", name=name, city=city)
    
    # Rate limiting
    async with _telefonbuch_rate:
        try:
            # Fetch the page using _make_client
            async with _make_client(True, USER_AGENT, None, False, HTTP_TIMEOUT) as client:
                resp = await client.get(url, timeout=HTTP_TIMEOUT)
                
                if not resp or resp.status_code != 200:
                    log("warn", "Telefonbuch-Query fehlgeschlagen", status=resp.status_code if resp else -1)
                    return []
                
                html = resp.text if hasattr(resp, 'text') else resp.content.decode('utf-8', errors='ignore')
                
        except Exception as e:
            log("warn", "Telefonbuch-Query Exception", error=str(e))
            return []
    
    # Parse HTML
    results = []
    parsing_strategy = None
    try:
        soup = BeautifulSoup(html, "html.parser")
        
        # Try to find entries - dasoertliche.de structure may vary
        # Strategy 1: Look for article elements with entry/treffer/hit classes
        entries = soup.find_all("article", class_=re.compile(r"entry|treffer|hit", re.I))
        if entries:
            parsing_strategy = "article_elements"
        
        # Strategy 2: Look for schema.org Person entries
        if not entries:
            entries = soup.find_all("div", itemtype=re.compile(r"Person", re.I))
            if entries:
                parsing_strategy = "schema_org_person"
        
        # Strategy 3: Fallback - look for any container with phone numbers
        if not entries:
            entries = soup.find_all("div", class_=re.compile(r"treffer|result|entry", re.I))
            if entries:
                parsing_strategy = "fallback_divs"
        
        if parsing_strategy:
            log("debug", "Telefonbuch-Parse-Strategie", strategy=parsing_strategy, entries_found=len(entries))
        else:
            log("debug", "Telefonbuch-Parse: Keine Einträge gefunden")
        
        for entry in entries[:5]:  # Limit to first 5 results
            result = {}
            
            # Extract name
            name_elem = entry.find(itemprop="name") or entry.find("h2") or entry.find(class_=re.compile(r"name", re.I))
            if name_elem:
                result["name"] = name_elem.get_text(strip=True)
            
            # Extract phone
            phone_elem = (
                entry.find(itemprop="telephone") or 
                entry.find(class_=re.compile(r"phone|telefon|tel", re.I)) or
                entry.find("a", href=re.compile(r"^tel:"))
            )
            if phone_elem:
                phone_text = phone_elem.get_text(strip=True)
                # Clean phone number
                phone_clean = re.sub(r'[^\d\+]', '', phone_text.replace(" ", ""))
                result["phone"] = phone_clean
            
            # Extract address
            address_elem = entry.find(itemprop="streetAddress") or entry.find(class_=re.compile(r"street|address|strasse", re.I))
            if address_elem:
                result["address"] = address_elem.get_text(strip=True)
            
            # Extract city
            city_elem = entry.find(itemprop="addressLocality") or entry.find(class_=re.compile(r"city|ort", re.I))
            if city_elem:
                result["city"] = city_elem.get_text(strip=True)
            
            # Only add if we have at least name and phone
            if result.get("name") and result.get("phone"):
                results.append(result)
        
    except Exception as e:
        log("warn", "Telefonbuch-Parse Exception", error=str(e))
        return []
    
    log("info", "Telefonbuch: Treffer gefunden", count=len(results))
    return results


def should_accept_enrichment(
    original_name: str,
    original_city: str,
    results: List[Dict]
) -> Tuple[bool, Optional[Dict], str]:
    """
    Prüft ob Enrichment akzeptiert werden soll.
    
    Checks:
        1. Genau 1 Treffer
        2. Name-Match >= 90% (fuzzy matching)
        3. Stadt-Match (enthält oder gleich)
        4. Telefonnummer ist Mobilnummer (015/016/017)
    
    Returns:
        (accept: bool, result: Dict or None, reason: str)
    """
    if not results:
        return False, None, "Keine Treffer"
    
    if TELEFONBUCH_STRICT_MODE and len(results) != 1:
        return False, None, f"Mehrere Treffer ({len(results)})"
    
    result = results[0]
    
    # Check name match (simple fuzzy matching)
    result_name = result.get("name", "").lower().strip()
    original_name_clean = original_name.lower().strip()
    
    # Simple similarity check - normalize and compare
    def normalize_name(n):
        # Remove common German academic and professional titles
        # Comprehensive list of titles to handle variations
        titles = [
            r'dr\.?\s*med\.?',  # Dr. med., Dr.med
            r'dr\.?\s*phil\.?',  # Dr. phil.
            r'dr\.?\s*ing\.?',   # Dr. Ing.
            r'dr\.?',            # Dr.
            r'prof\.?\s*dr\.?',  # Prof. Dr.
            r'prof\.?',          # Prof.
            r'dipl\.?-ing\.?',   # Dipl.-Ing.
            r'dipl\.?',          # Dipl.
            r'ing\.?',           # Ing.
            r'mag\.?',           # Mag.
            r'msc\.?',           # MSc.
            r'mba\.?',           # MBA.
            r'b\.?\s*sc\.?',     # B.Sc.
            r'm\.?\s*sc\.?',     # M.Sc.
        ]
        pattern = r'\b(' + '|'.join(titles) + r')\b'
        n = re.sub(pattern, '', n, flags=re.I)
        return ' '.join(n.split())
    
    result_name_norm = normalize_name(result_name)
    original_name_norm = normalize_name(original_name_clean)
    
    # Check if names are similar enough
    # Simple approach: check if one is contained in the other or vice versa
    name_match = (
        result_name_norm in original_name_norm or
        original_name_norm in result_name_norm or
        result_name_norm == original_name_norm
    )
    
    if not name_match:
        # Try word-by-word match (at least 2 words should match)
        result_words = set(result_name_norm.split())
        original_words = set(original_name_norm.split())
        common_words = result_words & original_words
        if len(common_words) < 2:
            return False, None, f"Name-Mismatch: '{result_name}' vs '{original_name}'"
    
    # Check city match
    result_city = result.get("city", "").lower().strip()
    original_city_clean = original_city.lower().strip()
    
    city_match = (
        result_city in original_city_clean or
        original_city_clean in result_city or
        result_city == original_city_clean
    )
    
    if not city_match and result_city:  # Allow missing city in result
        return False, None, f"Stadt-Mismatch: '{result_city}' vs '{original_city}'"
    
    # Check phone is mobile number
    phone = result.get("phone", "")
    if not phone:
        return False, None, "Keine Telefonnummer"
    
    # Normalize phone
    normalized_phone = normalize_phone(phone)
    
    # Check if mobile
    if TELEFONBUCH_MOBILE_ONLY:
        if not is_mobile_number(normalized_phone):
            return False, None, "Keine Mobilnummer"
    
    return True, result, "Akzeptiert"


async def enrich_phone_from_telefonbuch(
    name: str,
    city: str,
    strict: bool = True
) -> Optional[Dict[str, str]]:
    """
    Sucht Telefonnummer in dasoertliche.de
    
    Args:
        name: Vor- und Nachname (z.B. "Max Mustermann")
        city: Stadt (z.B. "Düsseldorf")
        strict: Nur bei exakt 1 Treffer (100% Genauigkeit)
    
    Returns:
        Dict mit {"phone": "0176...", "address": "..."} oder None
    
    Regeln:
        - Nur wenn Name UND Stadt vorhanden
        - Nur bei GENAU 1 Treffer (wenn strict=True)
        - Nur Mobilnummern (015x, 016x, 017x) werden akzeptiert
        - Rate-Limiting: Max 1 Request / 3 Sekunden
    """
    if not TELEFONBUCH_ENRICHMENT_ENABLED:
        return None
    
    if not name or not city:
        return None
    
    log("info", "Telefonbuch-Enrichment gestartet", name=name, city=city)
    
    # Check cache first
    cached = await get_cached_telefonbuch_result(name, city)
    if cached is not None:
        results = cached
    else:
        # Query dasoertliche.de
        results = await query_dasoertliche(name, city)
        
        # Cache the results
        await cache_telefonbuch_result(name, city, results)
    
    # Validate and accept enrichment
    accept, result, reason = should_accept_enrichment(name, city, results)
    
    if not accept:
        log("warn", "Telefonbuch-Enrichment abgelehnt", reason=reason, name=name, city=city)
        return None
    
    phone = result.get("phone", "")
    address = result.get("address", "")
    
    log("info", "Telefonbuch-Enrichment erfolgreich", 
        name=name, city=city, phone=phone[:8]+"..." if len(phone) > 8 else phone)
    
    return {
        "phone": phone,
        "address": address
    }


async def enrich_leads_with_telefonbuch(leads: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Enriches leads without phone numbers using telefonbuch lookup.
    Returns the enriched leads list.
    """
    if not TELEFONBUCH_ENRICHMENT_ENABLED or not leads:
        return leads
    
    enriched_leads = []
    for lead in leads:
        # Only enrich if NO phone but has name + city
        if not lead.get("telefon") and lead.get("name") and lead.get("region"):
            name = lead["name"]
            city = lead["region"]
            
            # Skip if name looks like a company
            if _looks_like_company_name(name):
                log("debug", "Telefonbuch-Enrichment übersprungen (Firmenname)", name=name)
                enriched_leads.append(lead)
                continue
            
            # Try to enrich
            enrichment = await enrich_phone_from_telefonbuch(name, city)
            
            if enrichment and enrichment.get("phone"):
                # Normalize and validate phone before assignment
                normalized_phone = normalize_phone(enrichment["phone"])
                if normalized_phone:  # Only proceed if normalization succeeded
                    is_valid, phone_type = validate_phone(normalized_phone)
                    if is_valid and phone_type == "mobile":
                        lead["telefon"] = normalized_phone
                        lead["phone_type"] = "mobile"
                        
                        # Add address if available
                        if enrichment.get("address"):
                            lead["private_address"] = enrichment["address"]
                        
                        # Tag as enriched
                        tags = lead.get("tags", "")
                        if tags:
                            lead["tags"] = tags + ",telefonbuch_enriched"
                        else:
                            lead["tags"] = "telefonbuch_enriched"
                        
                        log("info", "Telefonbuch-Enrichment erfolgreich", 
                            name=name, city=city, phone=normalized_phone[:8]+"..." if len(normalized_phone) > 8 else normalized_phone)
                    else:
                        log("debug", "Telefonbuch-Enrichment: Ungültige Nummer", 
                            name=name, phone=normalized_phone, phone_type=phone_type)
                else:
                    log("debug", "Telefonbuch-Enrichment: Normalisierung fehlgeschlagen", 
                        name=name, original_phone=enrichment["phone"])
        
        enriched_leads.append(lead)
    
    return enriched_leads


def insert_leads(leads: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Führt INSERT OR IGNORE aus. Zieht Schema automatisch nach (fehlende Spalten).
    Phone hardfilter: Re-validates phone before insert to ensure no invalid phones slip through.
    STRICT RULE: Only mobile numbers allowed - landline numbers are rejected.
    """
    if not leads:
        return []

    con = db(); cur = con.cursor()

    cols = ",".join(LEAD_FIELDS)
    placeholders = ",".join(["?"] * len(LEAD_FIELDS))
    sql = f"INSERT OR IGNORE INTO leads ({cols}) VALUES ({placeholders})"

    new_rows = []
    learning_engine = get_learning_engine()
    
    try:
        for r in leads:
            # CRITICAL: Check for job postings first - NEVER save job ads as leads
            source_url = r.get("quelle", "")
            if is_job_posting(url=source_url, title=r.get("name", ""), 
                             snippet=r.get("opening_line", ""), content=r.get("tags", "")):
                log("debug", "Lead dropped at insert (job posting)", url=source_url)
                continue
            
            # Name validation (heuristic only - async AI validation would be too slow here)
            name = (r.get("name") or "").strip()
            if name:
                is_real, confidence, reason = _validate_name_heuristic(name)
                if not is_real:
                    log("debug", "Lead dropped at insert (invalid name)", name=name, reason=reason, url=source_url)
                    continue
                r["name_validated"] = 1  # Mark as validated
            else:
                # No name - skip lead (name is mandatory)
                log("debug", "Lead dropped at insert (no name)", url=source_url)
                continue
            
            # Phone hardfilter: Re-check phone validity before insert
            phone = (r.get("telefon") or "").strip()
            if phone:
                is_valid, phone_type = validate_phone(phone)
                if not is_valid:
                    log("debug", "Lead dropped at insert (invalid phone)", phone=phone, url=source_url)
                    continue
                
                # STRICT: Only mobile numbers allowed
                # Normalize phone first before checking if it's mobile
                normalized_phone = normalize_phone(phone)
                if not is_mobile_number(normalized_phone):
                    log("debug", "Lead dropped at insert (not mobile number)", phone=phone, url=source_url)
                    continue
                
                # Update phone_type to mobile (since we validated it's mobile)
                r["phone_type"] = "mobile"
            else:
                # No phone - skip lead (phone is mandatory)
                log("debug", "Lead dropped at insert (no phone)", url=source_url)
                continue
            vals = [
                r.get("name",""),
                r.get("rolle",""),
                r.get("email",""),
                r.get("telefon",""),
                r.get("quelle",""),
                r.get("score",0),
                r.get("tags",""),
                r.get("region",""),
                r.get("role_guess",""),
                r.get("lead_type",""),
                r.get("salary_hint",""),
                r.get("commission_hint",""),
                r.get("opening_line",""),
                r.get("ssl_insecure","no"),
                r.get("company_name",""),
                r.get("company_size","unbekannt"),
                r.get("hiring_volume","niedrig"),
                r.get("industry","unbekannt"),
                r.get("recency_indicator","unbekannt"),
                r.get("location_specific",""),
                r.get("confidence_score",0),
                r.get("last_updated",""),
                r.get("data_quality",0),
                r.get("phone_type",""),
                r.get("whatsapp_link",""),
                r.get("private_address",""),
                r.get("social_profile_url",""),
                r.get("ai_category",""),
                r.get("ai_summary",""),
                # New candidate fields
                r.get("candidate_status",""),
                r.get("experience_years",0),
                r.get("availability",""),
                r.get("mobility",""),
                r.get("skills",""),
                r.get("industries_experience",""),
                r.get("source_type",""),
                r.get("profile_url",""),
                r.get("cv_url",""),
                r.get("contact_preference",""),
                r.get("last_activity",""),
                r.get("name_validated",0)
            ]
            cur.execute(sql, vals)
            if cur.rowcount > 0:
                new_rows.append(r)
                # Learn from successful lead (with mobile number)
                if learning_engine:
                    try:
                        # Get the query context from metadata if available
                        query_context = r.get("_query_context", "")
                        learning_engine.learn_from_success(r, query=query_context)
                    except Exception as e:
                        # Don't fail lead insertion if learning fails
                        log("debug", "Learning failed", error=str(e))
        con.commit()
    except sqlite3.OperationalError as e:
        # Fallback: sehr alte DB migrieren (harte UNIQUE/fehlende Spalten)
        con.rollback(); con.close()
        migrate_db_unique_indexes()
        con = db(); cur = con.cursor()
        for r in leads:
            vals = [
                r.get("name",""),
                r.get("rolle",""),
                r.get("email",""),
                r.get("telefon",""),
                r.get("quelle",""),
                r.get("score",0),
                r.get("tags",""),
                r.get("region",""),
                r.get("role_guess",""),
                r.get("lead_type",""),
                r.get("salary_hint",""),
                r.get("commission_hint",""),
                r.get("opening_line",""),
                r.get("ssl_insecure","no"),
                r.get("company_name",""),
                r.get("company_size","unbekannt"),
                r.get("hiring_volume","niedrig"),
                r.get("industry","unbekannt"),
                r.get("recency_indicator","unbekannt"),
                r.get("location_specific",""),
                r.get("confidence_score",0),
                r.get("last_updated",""),
                r.get("data_quality",0),
                r.get("phone_type",""),
                r.get("whatsapp_link",""),
                r.get("private_address",""),
                r.get("social_profile_url",""),
                r.get("ai_category",""),
                r.get("ai_summary",""),
                # New candidate fields
                r.get("candidate_status",""),
                r.get("experience_years",0),
                r.get("availability",""),
                r.get("mobility",""),
                r.get("skills",""),
                r.get("industries_experience",""),
                r.get("source_type",""),
                r.get("profile_url",""),
                r.get("cv_url",""),
                r.get("contact_preference",""),
                r.get("last_activity",""),
                r.get("name_validated",0)
            ]
            cur.execute(sql, vals)
            if cur.rowcount > 0:
                new_rows.append(r)
        con.commit()
    finally:
        con.close()

    return new_rows

def start_run() -> int:
    con = db(); cur = con.cursor()
    cur.execute(
        "INSERT INTO runs(started_at,status,links_checked,leads_new) VALUES(datetime('now'),'running',0,0)"
    )
    run_id = cur.lastrowid
    con.commit(); con.close()
    return run_id

def finish_run(run_id: int, links_checked: Optional[int] = None, leads_new: Optional[int] = None, status: str = "ok", metrics: Optional[Dict[str, int]] = None):
    con = db(); cur = con.cursor()
    cur.execute(
        "UPDATE runs SET finished_at=datetime('now'), status=?, links_checked=?, leads_new=? WHERE id=?",
        (status, links_checked or 0, leads_new or 0, run_id)
    )
    con.commit(); con.close()
    if metrics:
        log("info", "Run metrics", **metrics)

def reset_history():
    con = db(); cur = con.cursor()
    a = cur.execute("SELECT COUNT(*) FROM queries_done").fetchone()[0]
    b = cur.execute("SELECT COUNT(*) FROM urls_seen").fetchone()[0]
    cur.execute("DELETE FROM queries_done")
    cur.execute("DELETE FROM urls_seen")
    con.commit(); con.close()
    return a, b


# =========================
# HTTP/SSL/robots — ASYNC
# =========================

# Content-Guards (konfigurierbar)
MAX_CONTENT_LENGTH = int(os.getenv("MAX_CONTENT_LENGTH", str(MAX_FETCH_SIZE)))  # Use MAX_FETCH_SIZE (2MB default)
BINARY_CT_PREFIXES = ("image/", "video/", "audio/")
DENY_CT_EXACT = {
    "application/octet-stream",
    "application/x-msdownload",
}
PDF_CT = "application/pdf"

# Circuit-Breaker pro Host
_HOST_STATE: Dict[str, Dict[str, Any]] = {}  # {host: {"penalty_until": float, "failures": int}}
# URLs, die wegen 429/403 in die zweite Welle sollen
CB_BASE_PENALTY = int(os.getenv("CB_BASE_PENALTY", "90"))  # Sekunden
CB_MAX_PENALTY  = int(os.getenv("CB_MAX_PENALTY", "900"))

RETRY_INCLUDE_403 = (os.getenv("RETRY_INCLUDE_403", "0") == "1")
RETRY_MAX_PER_URL = int(os.getenv("RETRY_MAX_PER_URL", "2"))
RETRY_BACKOFF_BASE = float(os.getenv("RETRY_BACKOFF_BASE", "6.0"))
_RETRY_URLS: Dict[str, Dict[str, Any]] = {}  # url -> {"retries": int, "status": int}
_SITEMAP_FAILED_HOSTS: set[str] = set()
RUN_METRICS = {
    "removed_by_dropper": 0,
    "portal_dropped": 0,
    "impressum_dropped": 0,
    "pdf_dropped": 0,
    "retry_count": 0,
    "status_429": 0,
    "status_403": 0,
    "status_5xx": 0,
}

def _reset_metrics():
    global RUN_METRICS
    RUN_METRICS = {k: 0 for k in RUN_METRICS}

def _record_drop(reason: str):
    RUN_METRICS["removed_by_dropper"] += 1
    if reason in ("portal_host", "portal_domain"):
        RUN_METRICS["portal_dropped"] += 1
    if reason == "impressum_no_contact":
        RUN_METRICS["impressum_dropped"] += 1
    if reason == "pdf_without_cv_hint":
        RUN_METRICS["pdf_dropped"] += 1

def _record_retry(status: int):
    RUN_METRICS["retry_count"] += 1
    if status == 429:
        RUN_METRICS["status_429"] += 1
    elif status == 403:
        RUN_METRICS["status_403"] += 1
    if 500 <= status < 600:
        RUN_METRICS["status_5xx"] += 1

def _host_from(url: str) -> str:
    try:
        return urllib.parse.urlparse(url).netloc.lower()
    except Exception:
        return ""

def _penalize_host(host: str):
    if not host:
        return
    if host in {"www.googleapis.com", "googleapis.com"}:
        # Google API: never hard-penalize; at most a short cool-down
        st = _HOST_STATE.setdefault(host, {"penalty_until": 0.0, "failures": 0})
        penalty = random.uniform(10, 30)
        st["penalty_until"] = time.time() + penalty
        st["failures"] = 0
        log("info", "Google API backoff (soft)", host=host, penalty_s=penalty)
        return
    st = _HOST_STATE.setdefault(host, {"penalty_until": 0.0, "failures": 0})
    st["failures"] = min(st["failures"] + 1, 10)
    penalty = min(CB_BASE_PENALTY * (2 ** (st["failures"] - 1)), CB_MAX_PENALTY)
    st["penalty_until"] = time.time() + penalty
    log("warn", "Circuit-Breaker: Host penalized", host=host, failures=st["failures"], penalty_s=penalty)

def _host_allowed(host: str) -> bool:
    if host in {"www.googleapis.com", "googleapis.com"}:
        return True
    st = _HOST_STATE.get(host)
    if not st:
        return True
    if time.time() >= st.get("penalty_until", 0.0):
        st["failures"] = 0
        st["penalty_until"] = 0.0
        return True
    return False

def _should_retry_status(status: int) -> bool:
    if status in (429, 503, 504):
        return True
    if status == 403 and RETRY_INCLUDE_403:
        return True
    return False

def _schedule_retry(url: str, status: int):
    _record_retry(status)
    if not url:
        return
    if url in _RETRY_URLS:
        return
    _RETRY_URLS[url] = {"retries": 0, "status": status, "last_ts": time.time()}

# Globale Async-Client-Fabrik + Rate-Limiter
_CLIENT_SECURE: Optional[AsyncSession] = None
_CLIENT_INSECURE: Optional[AsyncSession] = None

async def get_client(secure: bool = True) -> AsyncSession:
    global _CLIENT_SECURE, _CLIENT_INSECURE
    proxy_cfg = {"http://": "socks5://127.0.0.1:9050", "https://": "socks5://127.0.0.1:9050"} if USE_TOR else None
    if secure:
        if _CLIENT_SECURE is None:
            _CLIENT_SECURE = AsyncSession(
                impersonate="chrome120",
                headers={"User-Agent": USER_AGENT, "Accept-Language": "de-DE,de;q=0.9,en;q=0.8"},
                verify=True,
                timeout=HTTP_TIMEOUT,
                proxies=proxy_cfg,
            )
        return _CLIENT_SECURE
    else:
        if _CLIENT_INSECURE is None:
            _CLIENT_INSECURE = AsyncSession(
                impersonate="chrome120",
                headers={"User-Agent": USER_AGENT, "Accept-Language": "de-DE,de;q=0.9,en;q=0.8"},
                verify=False,
                timeout=HTTP_TIMEOUT,
                proxies=proxy_cfg,
            )
        return _CLIENT_INSECURE
    
def _make_client(secure: bool, ua: str, proxy_url: Optional[str], force_http1: bool, timeout_s: int):
    # === TASK 2: Harden proxy handling for normal HTTP requests ===
    if USE_TOR:
        proxy_url = "socks5://127.0.0.1:9050"
    headers = {
        "User-Agent": ua or USER_AGENT,
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    }
    
    # When secure=True and NOT using Tor, explicitly disable proxies
    # to prevent environment variable lookup that causes ConnectTimeout
    if secure and not USE_TOR and not proxy_url:
        proxies = None  # Explicitly None to bypass all proxy lookups
    else:
        proxies = {"http://": proxy_url, "https://": proxy_url} if proxy_url else None
    
    return AsyncSession(
        impersonate="chrome120",
        headers=headers,
        verify=True if secure else False,
        timeout=timeout_s,
        proxies=proxies,
    )



# robots.txt Cache mit TTL
_ROBOTS_CACHE_TTL = int(os.getenv("ROBOTS_CACHE_TTL", "21600"))  # 6h
_ROBOTS_CACHE: Dict[str, Tuple[RobotFileParser, float]] = {}     # {base: (rp, ts)}

def _acceptable_by_headers(hdrs: Dict[str, str]) -> Tuple[bool, str]:
    ct = (hdrs.get("Content-Type", "") or "").lower().split(";")[0].strip()
    if any(ct.startswith(p) for p in BINARY_CT_PREFIXES) or ct in DENY_CT_EXACT:
        return False, f"content-type={ct}"
    if (PDF_CT in ct) and (not CFG.allow_pdf):
        return False, "pdf-not-allowed"
    try:
        cl = int(hdrs.get("Content-Length", "0"))
        if cl > 0 and cl > MAX_CONTENT_LENGTH:
            return False, f"too-large:{cl}"
    except Exception:
        pass
    return True, ct or "unknown"

# Letzten HTTP-Status pro URL merken
_LAST_STATUS: Dict[str, int] = {}

async def http_get_async(url, headers=None, params=None, timeout=HTTP_TIMEOUT):
    """
    GET mit optionalem HEAD-Preflight, Proxys/UA-Rotation, HTTP/2→1.1 Fallback
    Verbesserungen:
      - HEAD-Preflight: 405/501 nicht „bestrafen“ (kein erneutes HEAD), aber bei 405 Host sanft penalizen.
      - Host-Penalty zusätzlich bei 503/504 (Rate drosseln, spätere Retries wahrscheinlicher).
    """
    # --- Rotation wählen ---
    ua = random.choice(UA_POOL) if UA_POOL else USER_AGENT
    proxy = random.choice(PROXY_POOL) if PROXY_POOL else None

    # Eingehende Header mergen (UA aus Rotation hat Vorrang)
    headers = {**(headers or {})}
    headers["User-Agent"] = ua

    host = _host_from(url)
    if not _host_allowed(host):
        log("warn", "Circuit-Breaker: host muted (skip)", url=url, host=host)
        return None

    base_to = max(5, min(timeout, 45))
    eff_timeout = base_to + random.uniform(0.0, 1.25)

    # 1) HEAD-Preflight (secure, HTTP/2/1 je nach Param) – optional
    r_head = None
    try:
        async with _make_client(True, ua, proxy, force_http1=False, timeout_s=eff_timeout) as client_head:
            r_head = await client_head.head(url, headers=headers, params=params, allow_redirects=True, timeout=eff_timeout)
            if r_head is not None:
                # Wenn HEAD 405/501 → kein erneutes HEAD versuchen (spart Roundtrips).
                # Zusätzlich: bei 405 sanfte Host-Penalty (viele Sites blocken HEAD hart).
                if r_head.status_code == 405:
                    _penalize_host(host)
                    log("info", "HEAD 405: host penalized, continue with GET", url=url)
                if r_head.status_code in (405, 501):
                    pass  # einfach mit GET fortfahren

                elif r_head.status_code in (200, 204):
                    ok, reason = _acceptable_by_headers(r_head.headers or {})
                    if not ok:
                        log("info", "Head-preflight: skipped by headers", url=url, reason=reason)
                        return None
    except Exception:
        # Preflight ist optional – still & silent
        r_head = None

    async def _do_get(secure: bool, force_http1: bool) -> Optional[Any]:
        async with _make_client(secure, ua, proxy, force_http1, eff_timeout) as cl:
            return await cl.get(url, headers=headers, params=params, timeout=eff_timeout, allow_redirects=True)

    # 2) Primär GET (secure, HTTP/2 erlaubt)
    try:
        r = await _do_get(secure=True, force_http1=False)
        if r.status_code == 200:
            ok, reason = _acceptable_by_headers(r.headers or {})
            if not ok:
                log("info", "GET: skipped by headers", url=url, reason=reason)
                return None
            setattr(r, "insecure_ssl", False)
            return r
        if _should_retry_status(r.status_code):
            _penalize_host(host)
            _schedule_retry(url, r.status_code)
            log("warn", f"{r.status_code} received", url=url)
            return r
    except Exception:
        # 2a) Retry als HTTP/1.1
        try:
            r = await _do_get(secure=True, force_http1=True)
            if r.status_code == 200:
                ok, reason = _acceptable_by_headers(r.headers or {})
                if not ok:
                    return None
                setattr(r, "insecure_ssl", False)
                return r
            if _should_retry_status(r.status_code):
                _penalize_host(host)
                _schedule_retry(url, r.status_code)
                log("warn", f"{r.status_code} received (HTTP/1.1 retry)", url=url)
                return r
        except Exception:
            pass

    # 3) SSL-Fallback (unsicher), erst HTTP/2, dann HTTP/1.1
    if CFG.allow_insecure_ssl:
        try:
            r2 = await _do_get(secure=False, force_http1=False)
            if r2.status_code == 200:
                ok, reason = _acceptable_by_headers(r2.headers or {})
                if not ok:
                    return None
                setattr(r2, "insecure_ssl", True)
                log("warn", "SSL Fallback ohne Verify genutzt", url=url)
                return r2
            if _should_retry_status(r2.status_code):
                _penalize_host(host)
                _schedule_retry(url, r2.status_code)
                log("warn", f"{r2.status_code} received (insecure TLS)", url=url)
                return r2
        except Exception:
            try:
                r2 = await _do_get(secure=False, force_http1=True)
                if r2.status_code == 200:
                    ok, reason = _acceptable_by_headers(r2.headers or {})
                    if not ok:
                        return None
                    setattr(r2, "insecure_ssl", True)
                    log("warn", "SSL Fallback (HTTP/1.1) genutzt", url=url)
                    return r2
                if _should_retry_status(r2.status_code):
                    _penalize_host(host)
                    _schedule_retry(url, r2.status_code)
                    log("warn", f"{r2.status_code} received (insecure TLS, HTTP/1.1)", url=url)
                    return r2
            except Exception:
                pass

    if "sitemap" in (url or "").lower():
        log("debug", "Sitemap nicht verfügbar", url=url)
    else:
        log("error", "HTTP GET endgültig gescheitert", url=url)
    return None


async def fetch_response_async(url: str, headers=None, params=None, timeout=HTTP_TIMEOUT):
    r = await http_get_async(url, headers=headers, params=params, timeout=timeout)
    if r is None:
        _LAST_STATUS[url] = -1
        return None
    status = getattr(r, "status_code", 0)
    if status != 200:
        _LAST_STATUS[url] = status
        log("warn", "Nicht-200 beim Abruf – skip", url=url, status=status)
        return None
    _LAST_STATUS[url] = 200
    return r

def check_robots_txt(url: str, rp: Optional[RobotFileParser] = None) -> bool:
    return True

async def robots_allowed_async(url: str) -> bool:
    return True

# =========================
# Suche (modular)
# =========================

def _normalize_cx(s: str) -> str:
    if not s: return ""
    try:
        p = urllib.parse.urlparse(s)
        if p.query:
            q = urllib.parse.parse_qs(p.query)
            val = q.get("cx", [""])[0].strip()
            if val: return val
    except Exception:
        pass
    return s.strip()

def _jitter(a=0.2,b=0.8): return a + random.random()*(b-a)

GCS_CX = _normalize_cx(GCS_CX_RAW)
# Multi-Key/CX Rotation + Limits
GCS_KEYS = [k.strip() for k in os.getenv("GCS_KEYS","").split(",") if k.strip()] or ([GCS_API_KEY] if GCS_API_KEY else [])
GCS_CXS  = [_normalize_cx(x) for x in os.getenv("GCS_CXS","").split(",") if _normalize_cx(x)] or ([GCS_CX] if GCS_CX else [])
MAX_GOOGLE_PAGES = int(os.getenv("MAX_GOOGLE_PAGES","2"))  # Reduced to 1-2 for cost control

# ======= SUCHE: Branchen & Query-Baukasten (modular) =======
REGION = '(NRW OR "Nordrhein-Westfalen" OR Düsseldorf OR Köln OR Essen OR Dortmund OR Bochum OR Duisburg OR Mönchengladbach)'
CONTACT = '(kontakt OR impressum OR ansprechpartner OR "e-mail" OR email OR telefon OR whatsapp)'
SALES   = '(vertrieb OR d2d OR "call center" OR telesales OR outbound OR verkauf OR sales)'

# Kurze, treffsichere Query-Sets je Branche
INDUSTRY_QUERIES: dict[str, list[str]] = {
    "candidates": [
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 1: KLEINANZEIGEN STELLENGESUCHE (Primäre Quelle!)
        # ══════════════════════════════════════════════════════════════
        
        # Kleinanzeigen.de - Stellengesuche Vertrieb
        'site:kleinanzeigen.de/s-stellengesuche "vertrieb" "NRW"',
        'site:kleinanzeigen.de/s-stellengesuche "sales" "nordrhein-westfalen"',
        'site:kleinanzeigen.de/s-stellengesuche "außendienst" "erfahrung"',
        'site:kleinanzeigen.de/s-stellengesuche "key account" "suche"',
        'site:kleinanzeigen.de/s-stellengesuche "handelsvertreter" "freiberuflich"',
        'site:kleinanzeigen.de/s-stellengesuche "verkäufer" "mobil"',
        'site:kleinanzeigen.de/s-stellengesuche "kundenberater" "NRW"',
        'site:kleinanzeigen.de/s-stellengesuche "akquise" "erfahrung"',
        'site:kleinanzeigen.de/s-stellengesuche "telesales" "homeoffice"',
        'site:kleinanzeigen.de/s-stellengesuche "call center" "agent"',
        'site:kleinanzeigen.de "ich suche arbeit" "vertrieb"',
        'site:kleinanzeigen.de "suche stelle" "verkauf" "NRW"',
        'site:kleinanzeigen.de "biete meine dienste" "vertrieb"',
        
        # Markt.de - Stellengesuche
        'site:markt.de/stellengesuche "vertrieb"',
        'site:markt.de/stellengesuche "sales manager"',
        'site:markt.de/stellengesuche "außendienst" "PKW"',
        'site:markt.de/stellengesuche "handelsvertreter"',
        'site:markt.de "suche arbeit" "vertrieb" "NRW"',
        
        # Quoka - Stellengesuche
        'site:quoka.de/stellengesuche "vertrieb"',
        'site:quoka.de/stellengesuche "verkauf" "erfahrung"',
        'site:quoka.de "suche job" "sales"',
        
        # Kalaydo (Regional NRW)
        'site:kalaydo.de/stellengesuche "vertrieb"',
        'site:kalaydo.de "suche stelle" "außendienst"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 2: BUSINESS NETZWERKE (Xing, LinkedIn)
        # ══════════════════════════════════════════════════════════════
        
        # Xing - Offene Kandidaten
        'site:xing.com/profile "offen für angebote" "vertrieb" "NRW"',
        'site:xing.com/profile "auf jobsuche" "sales"',
        'site:xing.com/profile "suche neue herausforderung" "key account"',
        'site:xing.com/profile "verfügbar ab" "vertriebsleiter"',
        'site:xing.com/profile "in ungekündigter stellung" "außendienst"',
        'site:xing.com/profile "wechselwillig" "sales manager"',
        'site:xing.com "freiberuflicher handelsvertreter" "sucht"',
        'site:xing.com "sales" "open to work" "düsseldorf"',
        'site:xing.com "vertrieb" "neue chancen" "köln"',
        'site:xing.com "b2b vertrieb" "suche" "NRW"',
        
        # LinkedIn - Open to Work
        'site:linkedin.com/in "open to work" "sales" "germany"',
        'site:linkedin.com/in "offen für" "vertrieb" "NRW"',
        'site:linkedin.com/in "looking for" "sales" "düsseldorf"',
        'site:linkedin.com/in "seeking" "business development" "köln"',
        'site:linkedin.com/in "available" "account manager" "essen"',
        'site:linkedin.com/in "#opentowork" "vertrieb"',
        'site:linkedin.com "actively looking" "sales representative" "germany"',
        'site:linkedin.com "job seeker" "key account" "NRW"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 3: SOCIAL MEDIA (Facebook, Instagram, TikTok)
        # ══════════════════════════════════════════════════════════════
        
        # Facebook - Job Suche Posts
        'site:facebook.com "suche neue herausforderung" "vertrieb"',
        'site:facebook.com "suche arbeit" "sales" "NRW"',
        'site:facebook.com "wer kennt wen" "vertriebsjob"',
        'site:facebook.com "bin auf jobsuche" "verkauf"',
        'site:facebook.com "suche stelle" "außendienst" "erfahrung"',
        'site:facebook.com/groups "vertriebler" "jobsuche"',
        'site:facebook.com/groups "sales jobs" "deutschland"',
        'site:facebook.com "freelancer" "vertrieb" "verfügbar"',
        'site:facebook.com "gekündigt" "vertrieb" "suche"',
        
        # Instagram - Job Seeker Bios
        'site:instagram.com "open for work" "sales"',
        'site:instagram.com "job gesucht" "vertrieb"',
        'site:instagram.com "DM for work" "sales"',
        'site:instagram.com "suche job" "verkauf"',
        'site:instagram.com "available for hire" "business"',
        'site:instagram.com "freelancer" "vertrieb" "kontakt"',
        
        # TikTok - Karriere Content
        'site:tiktok.com "jobsuche" "vertrieb"',
        'site:tiktok.com "suche arbeit" "sales"',
        'site:tiktok.com "arbeitslos" "vertriebler"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 4: MESSENGER GRUPPEN (Telegram, WhatsApp, Discord)
        # ══════════════════════════════════════════════════════════════
        
        # Telegram - Öffentliche Gruppen
        'site:t.me "vertrieb" "jobs" "gruppe"',
        'site:t.me "sales jobs" "germany"',
        'site:t.me "jobsuche" "NRW"',
        'site:t.me "vertriebler" "netzwerk"',
        'site:t.me "stellengesuche" "vertrieb"',
        'site:t.me/joinchat "vertrieb"',
        'site:t.me/joinchat "sales" "jobs"',
        
        # WhatsApp - Öffentliche Einladungslinks
        'site:chat.whatsapp.com "vertrieb" "jobs"',
        'site:chat.whatsapp.com "sales" "netzwerk"',
        'site:chat.whatsapp.com "vertriebler" "gruppe"',
        'site:chat.whatsapp.com "jobsuche" "NRW"',
        
        # Discord - Karriere Server
        'site:discord.gg "vertrieb" "jobs"',
        'site:discord.gg "sales" "karriere"',
        'site:discord.com/invite "jobs" "deutschland"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 5: FOREN & COMMUNITIES
        # ══════════════════════════════════════════════════════════════
        
        # Reddit - Deutsche Karriere-Subreddits
        'site:reddit.com/r/arbeitsleben "vertrieb" "neuer job"',
        'site:reddit.com/r/arbeitsleben "sales" "kündigung"',
        'site:reddit.com/r/arbeitsleben "außendienst" "wechsel"',
        'site:reddit.com/r/de_EDV "vertrieb" "jobsuche"',
        'site:reddit.com/r/FragReddit "vertriebsjob" "erfahrung"',
        'site:reddit.com/r/Finanzen "vertrieb" "gehalt" "wechsel"',
        'site:reddit.com "gekündigt" "sales" "was tun"',
        'site:reddit.com "arbeitslos" "vertrieb" "bewerbung"',
        
        # Gutefrage.net
        'site:gutefrage.net "vertrieb" "stelle suchen"',
        'site:gutefrage.net "sales job" "erfahrung"',
        'site:gutefrage.net "außendienst" "bewerbung"',
        'site:gutefrage.net "handelsvertreter werden"',
        
        # Wer-weiss-was.de
        'site:wer-weiss-was.de "vertriebsjob" "suche"',
        'site:wer-weiss-was.de "sales karriere"',
        
        # Motor-Talk (Autoverkäufer)
        'site:motor-talk.de "autoverkäufer" "suche stelle"',
        'site:motor-talk.de "verkaufsberater" "wechsel"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 6: FREELANCER PORTALE
        # ══════════════════════════════════════════════════════════════
        
        # Freelancer Profile
        'site:freelancermap.de "vertrieb" "verfügbar"',
        'site:freelancermap.de "sales" "freiberufler"',
        'site:freelance.de "vertriebsexperte" "profil"',
        'site:freelance.de "handelsvertreter" "selbstständig"',
        'site:gulp.de "sales" "verfügbar"',
        'site:gulp.de "vertrieb" "freiberuflich"',
        
        # Fiverr/Upwork (Vertriebs-Freelancer)
        'site:fiverr.com "sales" "germany" "b2b"',
        'site:upwork.com "sales representative" "german"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 7: LEBENSLAUF-DATENBANKEN & JOBPORTALE
        # ══════════════════════════════════════════════════════════════
        
        # Lebenslauf-Portale (wo Kandidaten sich eintragen)
        'site:lebenslaufmuster.de "vertrieb" "berufserfahrung"',
        '"mein lebenslauf" "vertrieb" "NRW" "mobil" filetype:pdf',
        '"curriculum vitae" "sales" "germany" "phone" filetype:pdf',
        '"bewerbung" "vertriebserfahrung" "kontakt" filetype:pdf',
        
        # StepStone Kandidatenprofile
        'site:stepstone.de/kandidat "vertrieb"',
        'site:stepstone.de "profil" "sales manager"',
        
        # Indeed Kandidatensuche
        'site:indeed.com/r "vertrieb" "NRW"',
        'site:indeed.com "lebenslauf" "sales"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 8: BRANCHEN-SPEZIFISCHE KANDIDATEN
        # ══════════════════════════════════════════════════════════════
        
        # D2D / Door-to-Door Vertriebler
        'site:kleinanzeigen.de "d2d" "suche" "erfahrung"',
        '"door to door" "vertriebler" "sucht" "mobil"',
        '"haustürgeschäft" "erfahrung" "suche arbeit"',
        
        # Call Center / Telesales
        '"call center agent" "suche" "homeoffice" "NRW"',
        '"telesales" "erfahrung" "suche stelle"',
        '"telefonvertrieb" "freiberuflich" "verfügbar"',
        'site:kleinanzeigen.de "telefonverkäufer" "suche"',
        
        # Energie / Solar Vertriebler
        '"solarvertrieb" "suche" "erfahrung" "NRW"',
        '"energieberater" "freiberuflich" "sucht"',
        '"photovoltaik" "vertrieb" "suche stelle"',
        
        # Versicherung / Finanz
        '"versicherungsvertreter" "suche" "neue"',
        '"finanzberater" "wechselwillig" "kontakt"',
        '"makler" "sucht" "neue herausforderung"',
        
        # Telekommunikation
        '"telekom vertrieb" "suche" "erfahrung"',
        '"mobilfunk" "sales" "suche stelle"',
        '"provider" "vertrieb" "wechsel"',
        
        # Medizin / Pharma
        '"pharmareferent" "sucht" "neue"',
        '"medizinprodukteberater" "verfügbar"',
        '"healthcare sales" "germany" "open"',
        
        # IT / Software
        '"software sales" "suche" "germany"',
        '"it vertrieb" "suche stelle" "NRW"',
        '"saas" "account executive" "open to"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 9: REGIONALE NRW-SUCHE
        # ══════════════════════════════════════════════════════════════
        
        # Düsseldorf
        '"vertrieb" "suche" "düsseldorf" "mobil"',
        '"sales" "jobsuche" "düsseldorf" "kontakt"',
        'site:xing.com "vertrieb" "düsseldorf" "offen für"',
        
        # Köln
        '"vertrieb" "suche" "köln" "erfahrung"',
        '"außendienst" "köln" "suche stelle"',
        'site:linkedin.com "sales" "köln" "open"',
        
        # Essen/Ruhrgebiet
        '"vertrieb" "essen" "suche" "mobil"',
        '"sales" "ruhrgebiet" "jobsuche"',
        '"dortmund" "vertrieb" "suche stelle"',
        
        # Weitere NRW Städte
        '"vertrieb" "wuppertal" "suche"',
        '"sales" "bonn" "suche stelle"',
        '"außendienst" "münster" "suche"',
        '"vertrieb" "bielefeld" "jobsuche"',
        '"verkauf" "aachen" "suche stelle"',
        
        # ══════════════════════════════════════════════════════════════
        # KATEGORIE 10: KARRIERE-EVENTS & NETZWERKE
        # ══════════════════════════════════════════════════════════════
        
        # Karrieremessen Teilnehmer
        '"jobmesse" "NRW" "vertrieb" "besucher"',
        '"karrieretag" "düsseldorf" "sales"',
        '"connecticum" "vertrieb" "teilnehmer"',
        
        # Vertriebler-Netzwerke
        'site:vertriebsmanager.de "mitglied" "profil"',
        'site:salesjob.de "kandidat" "profil"',
        '"bdvt" "mitglied" "vertrieb" "kontakt"',
        
        # Alumni-Netzwerke
        '"sales" "alumni" "NRW" "kontakt"',
        '"vertrieb" "absolvent" "suche stelle"',
    ],
}

# NEU: Recruiter-spezifische Queries für Vertriebler-Rekrutierung
RECRUITER_QUERIES = {
    'recruiter': [
        'site:kleinanzeigen.de/s-stellengesuche/ "außendienst" NRW',
        'site:kleinanzeigen.de/s-stellengesuche/ "handelsvertreter" NRW',
        'site:kleinanzeigen.de/s-stellengesuche/ "verkauf" NRW',
        'site:kleinanzeigen.de/s-stellengesuche/ "kundenberater" NRW',
        'site:kleinanzeigen.de/s-stellengesuche/ "suche job" "vertrieb" NRW',
        'site:kleinanzeigen.de/s-stellengesuche/ "arbeit" "verkauf" NRW',
        'site:kleinanzeigen.de/s-stellengesuche/ "quereinsteiger" "vertrieb" NRW',
        # Facebook Gruppen & Profile
        'site:facebook.com "suche arbeit" "vertrieb" ("017" OR "016" OR "015" OR "+49")',
        'site:facebook.com/groups/ "stellengesuche" ("017" OR "016")',
        'site:facebook.com "jobsuche" "verkauf" ("017" OR "016" OR "015")',
    ]
}


# Fallback für "alle" Branchen – Reihenfolge
INDUSTRY_ORDER = ["nrw","social","solar","telekom","versicherung","bau","ecom","household"]

def build_queries(
    selected_industry: Optional[str] = None,
    per_industry_limit: int = 20000
) -> List[str]:
    """
    Build a compact set of high-precision Handelsvertreter dorks.
    If candidates/recruiter mode is selected, use INDUSTRY_QUERIES["candidates"] instead.
    """
    # CRITICAL FIX: Use candidates queries when in candidates/recruiter mode
    if selected_industry and selected_industry.lower() in ("candidates", "recruiter"):
        base = INDUSTRY_QUERIES.get("candidates", [])
        queries = list(dict.fromkeys(base))
        random.shuffle(queries)
        cap = min(max(1, per_industry_limit), len(queries))
        return queries[:cap]
    
    # Standard mode: use DEFAULT_QUERIES
    queries: List[str] = list(dict.fromkeys(DEFAULT_QUERIES))
    random.shuffle(queries)

    cap = min(max(1, per_industry_limit), len(queries))
    return queries[:cap]

def is_denied(url: str) -> bool:
    p = urllib.parse.urlparse(url)
    host = (p.netloc or "").lower()
    # Normalisieren: www./m. abschneiden
    if host.startswith("www."):
        host = host[4:]
    if host.startswith("m."):
        host = host[2:]

    if host in SOCIAL_HOSTS:
        return False

    # Harte Domain-Blockliste (bestehend)
    for d in DENY_DOMAINS:
        d = d.lower()
        if host == d or host.endswith("." + d):
            return True

    # Bestehende Heuristiken
    if host.startswith(("uni-", "cdu-", "stadtwerke-")):
        return True
    if host in {"ruhr-uni-bochum.de"}:
        return True

    # NEU: NRW-Rauschen — IHK/HWK/Bildung
    if host.endswith(".ihk.de") or host.startswith(("ihk-", "hwk-")):
        return True
    if any(k in host for k in (
        "schule", "berufskolleg", "weiterbildung", "bildungszentrum",
        "akademie", "bbw-", "bfw-", "leb-"
    )):
        return True

    # NEU: Jobportale/Aggregatoren (ziehen Budget, liefern selten direkte tel/mail/wa)
    PORTAL_HOST_HINTS = (
        "stepstone", "indeed", "monster", "jobware", "stellenanzeigen",
        "jobvector", "yourfirm", "metajob", "stellenangebotevertrieb",
        "jobanzeiger", "jobboerse.arbeitsagentur", "meinestadt"
    )
    if any(h in host for h in PORTAL_HOST_HINTS):
        return True

    return False


def path_ok(url: str) -> bool:
    p = urllib.parse.urlparse(url)
    path = (p.path or "").lower()
    q    = (p.query or "").lower()
    frag = (p.fragment or "").lower()
    host = (p.netloc or "").lower()
    if host.startswith("www."):
        host = host[4:]
    if host in SOCIAL_HOSTS:
        return True
    if host.endswith("kleinanzeigen.de") or host.endswith("ebay-kleinanzeigen.de"):
        return True
    if any(bad in path for bad in NEG_PATH_HINTS):
        return False
    positive = any(h in path for h in ALLOW_PATH_HINTS) \
               or any(h in q for h in ALLOW_PATH_HINTS) \
               or any(h in frag for h in ALLOW_PATH_HINTS)
    is_rootish = (path in ("", "/"))
    return positive or is_rootish

def _normalize_for_dedupe(u: str) -> str:
    try:
        pu = urllib.parse.urlparse(u)

        # Query normalisieren – Tracking & Paginierung raus
        q = urllib.parse.parse_qsl(pu.query, keep_blank_values=False)
        q = [
            (k, v) for (k, v) in q
            if not (
                k.lower().startswith(("utm_",))                       # utm_*
                or k.lower() in {"gclid", "fbclid", "mc_eid", "page"} # + page-Param raus
            )
        ]
        new_q = urllib.parse.urlencode(q, doseq=True)

        # Pfad normalisieren: Fragment weg, trailing slash schlucken
        path = pu.path or "/"
        if path != "/" and path.endswith("/"):
            path = path[:-1]

        pu = pu._replace(query=new_q, fragment="", path=path)
        return urllib.parse.urlunparse(pu)
    except Exception:
        return u


UrlLike = Union[str, Dict[str, Any]]


def _extract_url(item: UrlLike) -> str:
    if isinstance(item, str):
        return item
    if isinstance(item, dict):
        return item.get("url") or item.get("link", "")
    return ""


from typing import List
import re, urllib.parse

def prioritize_urls(urls: List[str]) -> List[str]:
    """
    Priorisiert typische Kontaktseiten deutlich höher als Karriere/Jobs/Datenschutz.
    - Additives Scoring (nicht nur erstes Pattern)
    - Starke Upvotes: /kontakt, /impressum
    - Downvotes: /karriere, /jobs, /stellen, /datenschutz, /privacy, /agb
    - Leichte Bevorzugung kurzer/oberflächlicher Pfade; Abwertung bei Paginierung/Fragmenten
    """
    # --- POSITIVE & NEGATIVE Schlüsselwörter ---
    # Hinweis: additive Bewertung; ein Pfad kann mehrere Treffer bekommen.
    PRIORITY_PATTERNS = [
        (r'/kontakt(?:/|$)',                    +40),
        (r'/kontaktformular(?:/|$)',            +32),
        (r'/impressum(?:/|$)',                  +35),
        (r'/team(?:/|$)',                       +12),
        (r'/ueber-uns|/über-uns|/unternehmen',  +10),
        # NEU: Jobseeker-Signale boosten
        (r'/karriere(?:/|$)',                   +18),
        (r'/jobs?(?:/|$)',                      +18),
        (r'/stellen(?:/|$)',                    +16),
        (r'/bewerb|/lebenslauf|/cv|/profil',    +24),
    ]

    PENALTY_PATTERNS = [
        (r'/datenschutz|/privacy|/policy',      -40),
        (r'/agb|/terms|/bedingungen',           -16),
        (r'/login|/signin|/signup|/account',    -70),
        (r'/cart|/warenkorb|/checkout|/search', -70),
        (r'/blog/|/news/',                      -30),
    ]


    def _score(u: str) -> int:
        s = 50
        path = urllib.parse.urlparse(u).path or "/"
        low = u.lower()

        # Positive Muster: additiv werten
        for pat, pts in PRIORITY_PATTERNS:
            if re.search(pat, low, re.I):
                s += pts

        # Negative Muster: additiv abwerten
        for pat, pts in PENALTY_PATTERNS:
            if re.search(pat, low, re.I):
                s += pts

        # Query-Hints (selten, aber wenn vorhanden → leicht positiv)
        if re.search(r'(\?|#).*(kontakt|impressum)', low, re.I):
            s += 12

        # Pfad-Länge / -Tiefe: kurze, flache Pfade bevorzugen
        depth = max(0, path.count('/') - 1)
        s += max(0, 20 - len(path))           # kürzerer Pfad = besser
        s += max(0, 10 - 5*depth)             # weniger Unterverzeichnisse = besser

        # Rauschfilter / Paginierung / Fragmente
        if len(u) > 220:                      # sehr lange URLs wirken oft „Rauschen“
            s -= 40
        if re.search(r'(\?|&)page=\d{1,3}\b', low):
            s -= 25
        if re.search(r'/page/\d{1,3}\b', low):
            s -= 25
        if '#' in u:                          # Anker/Kommentare/Blogfragmente
            s -= 10

        return s

    # Dedupe (auf Basis deiner Normalisierung) + Scoring + Sortierung
    normed, seen = [], set()
    for u in urls:
        nu = _normalize_for_dedupe(u)
        if nu in seen:
            continue
        seen.add(nu)
        normed.append(nu)

    scored = [(u, _score(u)) for u in normed]
    scored.sort(key=lambda x: (-x[1], x[0]))
    return [u for u, _ in scored]


async def search_perplexity_async(query: str) -> List[Dict[str, str]]:
    """
    Perplexity (sonar) search returning citation URLs.
    """
    pplx_key = os.getenv("PERPLEXITY_API_KEY", "")
    if not pplx_key:
        return []

    url = "https://api.perplexity.ai/chat/completions"
    headers = {
        "Authorization": f"Bearer {pplx_key}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "sonar",
        "messages": [
            {"role": "system", "content": "You are a lead generation engine. Search for the user's query and return relevant business URLs. ensure citations are included."},
            {"role": "user", "content": query}
        ]
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(url, headers=headers, json=payload, timeout=HTTP_TIMEOUT) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    citations = data.get("citations", []) or []
                    log("info", "Perplexity found citations", count=len(citations))
                    return [{'link': u, 'title': 'Perplexity Source', 'snippet': 'AI Verified'} for u in citations if u]
                log("error", "Perplexity API Error", status=resp.status)
                return []
    except Exception as e:
        log("error", "Perplexity Exception", error=str(e))
        return []

async def google_cse_search_async(q: str, max_results: int = 60, date_restrict: Optional[str] = None) -> Tuple[List[Dict[str, str]], bool]:
    if os.getenv("DISABLE_GOOGLE") == "1" or not (GCS_KEYS and GCS_CXS):
        log("info","Google CSE disabled; skipping"); return [], False

    def _preview(txt: Optional[str]) -> str:
        if not txt: return ""
        # HTML-Tags & übermäßige Whitespaces entfernen, dann hart bei 200 cutten
        txt = re.sub(r"<[^>]+>", " ", txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt[:200]

    url = "https://www.googleapis.com/customsearch/v1"
    results: List[Dict[str, str]] = []
    page_no, key_i, cx_i = 0, 0, 0
    had_429 = False
    page_cap = int(os.getenv("MAX_GOOGLE_PAGES","2"))  # Reduced to 1-2 for cost control
    while len(results) < max_results and page_no < page_cap:
        params = {
            "key": GCS_KEYS[key_i], "cx": GCS_CXS[cx_i], "q": q,
            "num": min(10, max_results - len(results)),
            "start": 1 + page_no*10, "lr":"lang_de", "safe":"off",
        }
        if date_restrict:
            params["dateRestrict"] = date_restrict

        try:
            r = await http_get_async(url, headers=None, params=params, timeout=HTTP_TIMEOUT)
        except (ReadError, ClientConnectorError, ServerDisconnectedError) as e:
            log("warn", "Google CSE Netzfehler, retry", error=str(e))
            await asyncio.sleep(3)
            continue
        except Exception as e:
            log("warn", "Google CSE Fehler", error=str(e))
            await asyncio.sleep(3)
            continue
        if not r:
            key_i = (key_i + 1) % max(1,len(GCS_KEYS))
            cx_i  = (cx_i  + 1) % max(1,len(GCS_CXS))
            sleep_s = 4 + int(4*_jitter())
            await asyncio.sleep(sleep_s)
            continue

        if r.status_code == 429:
            had_429 = True
            log("warn","Google 429 – skip this query without retry")
            return [], had_429

        if r.status_code != 200:
            log("error","Google CSE Status != 200", status=r.status_code, body=_preview(r.text))
            break

        try:
            data = r.json()
        except Exception:
            log("error","Google CSE JSON-Parsing fehlgeschlagen", text=_preview(r.text))
            break

        items = data.get("items", []) or []
        batch = [
            {
                "url": it.get("link"),
                "title": it.get("title", "") or "",
                "snippet": it.get("snippet", "") or "",
            }
            for it in items
            if it.get("link")
        ]
        results.extend(batch)
        log("info","Google CSE Batch", q=q, batch=len(batch), total=len(results), page_no=page_no)

        if not batch: break
        page_no += 1
        await asyncio.sleep(0.5 + _jitter(0,0.6))

    uniq_items: List[Dict[str, str]] = []
    seen = set()
    for entry in results:
        raw_url = entry.get("url", "")
        if not raw_url:
            continue
        nu = _normalize_for_dedupe(raw_url)
        if nu in seen: continue
        seen.add(nu)
        if is_denied(nu): continue
        if not path_ok(nu): continue
        uniq_items.append({
            "url": nu,
            "title": entry.get("title", "") or "",
            "snippet": entry.get("snippet", "") or "",
        })
    if uniq_items:
        ordered = prioritize_urls([item["url"] for item in uniq_items])
        order_map = {u: i for i, u in enumerate(ordered)}
        uniq_items.sort(key=lambda item: order_map.get(item["url"], len(order_map)))
    return uniq_items, had_429


async def duckduckgo_search_async(query: str, max_results: int = 10, date_restrict: Optional[str] = None) -> List[Dict[str, str]]:
    """
    DuckDuckGo-Suche mit strenger Proxy-Steuerung, um ConnectTimeouts zu vermeiden.
    
    === TASK 3: Hardened proxy control ===
    Explicitly manages environment variables to force direct connection (USE_TOR=False)
    or TOR routing (USE_TOR=True) before DDGS initialization.
    """
    if not HAVE_DDG:
        log("warn", "DuckDuckGo-Modul fehlt.")
        return []

    results: List[Dict[str, str]] = []

    for attempt in range(1, 3):  # Max 1 retry (2 attempts total)
        try:
            # === CRITICAL: Set environment variables *inside* the try block ===
            # This ensures clean state on each retry attempt
            if USE_TOR:
                # Force TOR routing via SOCKS5 proxy
                os.environ["HTTP_PROXY"] = "socks5://127.0.0.1:9050"
                os.environ["HTTPS_PROXY"] = "socks5://127.0.0.1:9050"
                os.environ["ALL_PROXY"] = "socks5://127.0.0.1:9050"
                os.environ["http_proxy"] = "socks5://127.0.0.1:9050"
                os.environ["https_proxy"] = "socks5://127.0.0.1:9050"
                # Remove no_proxy to ensure proxy is used
                os.environ.pop("no_proxy", None)
                os.environ.pop("NO_PROXY", None)
                log("debug", "DuckDuckGo: TOR proxy configured", proxy="socks5://127.0.0.1:9050")
            else:
                # === Force direct connection by nuclear cleanup ===
                # Clear ALL proxy variables to prevent ConnectTimeout (WinError 10060)
                for key in PROXY_ENV_VARS:
                    os.environ.pop(key, None)
                
                # Explicitly set no_proxy (both case variants) to bypass any system-level proxy settings
                # This is the "nuclear option" to ensure ddgs/httpx/curl_cffi don't use any proxy
                os.environ["no_proxy"] = "*"
                os.environ["NO_PROXY"] = "*"
                log("debug", "DuckDuckGo: Direct connection configured (no_proxy='*', all proxies cleared)")
            
            # Initialize DDGS (Kurzes Timeout für "Fail Fast")
            with DDGS(timeout=10) as ddgs:
                gen = ddgs.text(
                    query,
                    region="de-de",
                    safesearch="off",
                    timelimit="y",
                    max_results=max_results
                )
                count = 0
                for r in gen:
                    if count >= max_results:
                        break
                    link = r.get("href")
                    title = r.get("title", "")
                    snippet = r.get("body", "")
                    if link:
                        results.append({
                            "link": link,
                            "title": title,
                            "snippet": snippet
                        })
                        count += 1

                if results:
                    log("info", "DuckDuckGo Treffer", q=query, count=len(results))
                else:
                    log("info", "DuckDuckGo: Keine Treffer (Seite leer)", q=query)
                    # Log query as weak after "No results" on first attempt
                    # Metrics system tracks this for adaptive dork selection
                    if attempt == 1:
                        log("debug", "DuckDuckGo: Query weak (no results)", q=query)
                return results

        except Exception as e:
            err_msg = str(e)
            if "ConnectTimeout" in err_msg or "WinError 10060" in err_msg:
                log("warn", "DuckDuckGo: Netzwerkproblem, überspringe", q=query)
                return []
            if attempt < 2:  # Only 1 retry allowed
                log("warn", f"DuckDuckGo Retry {attempt}/2 wegen Fehler", error=err_msg, q=query)
                await asyncio.sleep(3)
            else:
                log("warn", "DuckDuckGo endgültig gescheitert nach 1 Retry", error=err_msg, q=query)
                # Log query as weak after failure - metrics system tracks this
                log("debug", "DuckDuckGo: Query weak (failed)", q=query)

    return []

def _ka_keywords_from_query(q: str) -> str:
    if not q:
        return ""
    cleaned = re.sub(r"site:[^\s]+", " ", q, flags=re.I)
    cleaned = re.sub(r"[()\"']", " ", cleaned)
    cleaned = re.sub(r"\b(OR|AND)\b", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned[:120]


async def kleinanzeigen_search_async(q: str, max_results: int = KLEINANZEIGEN_MAX_RESULTS) -> List[Dict[str, str]]:
    if (not ENABLE_KLEINANZEIGEN) or max_results <= 0:
        return []

    keywords = _ka_keywords_from_query(q)
    if not keywords:
        return []

    url = "https://www.kleinanzeigen.de/s-stellengesuche/k0"
    try:
        r = await http_get_async(
            url,
            params={
                "keywords": keywords,
                "locationStr": "NRW",
            },
            timeout=HTTP_TIMEOUT
        )
    except Exception as e:
        log("warn", "Kleinanzeigen-Suche fehlgeschlagen", q=keywords, err=str(e))
        return []
    if not r:
        return []
    if r.status_code != 200:
        log("warn", "Kleinanzeigen Status != 200", status=r.status_code, q=keywords)
        return []

    html = r.text or ""
    soup = BeautifulSoup(html, "html.parser")
    items: List[Dict[str, str]] = []
    for art in soup.select("li.ad-listitem article.aditem"):
        href = art.get("data-href") or ""
        if not href:
            a_tag = art.find("a", href=True)
            if a_tag:
                href = a_tag.get("href", "")
        if not href:
            continue
        full = urllib.parse.urljoin("https://www.kleinanzeigen.de", href)
        title_el = art.select_one("h2 a")
        desc_el = art.select_one(".aditem-main--middle--description")
        title = title_el.get_text(" ", strip=True) if title_el else ""
        snippet = desc_el.get_text(" ", strip=True) if desc_el else ""
        items.append({"url": full, "title": title, "snippet": snippet})
        if len(items) >= max_results:
            break

    uniq: List[Dict[str, str]] = []
    seen = set()
    for entry in items:
        u = _extract_url(entry)
        if not u:
            continue
        nu = _normalize_for_dedupe(u)
        if nu in seen:
            continue
        seen.add(nu)
        if is_denied(nu):
            continue
        uniq.append({**entry, "url": nu})

    if uniq:
        log("info", "Kleinanzeigen Treffer", q=keywords, count=len(uniq))
    return uniq


async def crawl_kleinanzeigen_listings_async(listing_url: str, max_pages: int = 5) -> List[str]:
    """
    Crawl Kleinanzeigen listing pages directly (not via Google) and extract all ad links.
    Supports pagination (page 1, 2, 3...).
    
    Args:
        listing_url: Base URL for the listing (e.g., https://www.kleinanzeigen.de/s-stellengesuche/...)
        max_pages: Maximum number of pages to crawl (default: 5)
        
    Returns:
        List of ad detail URLs
    """
    if not ENABLE_KLEINANZEIGEN:
        return []
    
    ad_links: List[str] = []
    seen_urls = set()
    page_num = 1  # Initialize before loop
    
    for page_num in range(1, max_pages + 1):
        # Build URL with page parameter, properly handling existing query params
        if page_num == 1:
            url = listing_url
        else:
            # Parse URL and add page parameter
            parsed = urllib.parse.urlparse(listing_url)
            params = urllib.parse.parse_qs(parsed.query)
            params['page'] = [str(page_num)]
            new_query = urllib.parse.urlencode(params, doseq=True)
            url = urllib.parse.urlunparse(parsed._replace(query=new_query))
        
        try:
            # Rate limiting: 2-3 seconds between requests
            if page_num > 1:
                await asyncio.sleep(2.0 + _jitter(0.5, 1.0))
            
            log("info", "Crawling Kleinanzeigen listing", url=url, page=page_num)
            
            r = await http_get_async(url, timeout=HTTP_TIMEOUT)
            if not r or r.status_code != 200:
                log("warn", "Failed to fetch listing page", url=url, status=r.status_code if r else "None")
                break
            
            html = r.text or ""
            soup = BeautifulSoup(html, "html.parser")
            
            # Extract ad links from listing
            page_links = 0
            for art in soup.select("li.ad-listitem article.aditem"):
                # Try data-href first
                href = art.get("data-href") or ""
                if not href:
                    # Fallback to anchor tag
                    a_tag = art.find("a", href=True)
                    if a_tag:
                        href = a_tag.get("href", "")
                
                if not href or "/s-anzeige/" not in href:
                    continue
                
                # Build full URL using the base URL from the listing
                parsed_listing = urllib.parse.urlparse(listing_url)
                base_url = f"{parsed_listing.scheme}://{parsed_listing.netloc}"
                full_url = urllib.parse.urljoin(base_url, href)
                norm_url = _normalize_for_dedupe(full_url)
                
                if norm_url in seen_urls:
                    continue
                
                seen_urls.add(norm_url)
                ad_links.append(full_url)
                page_links += 1
            
            log("info", "Extracted ad links from page", page=page_num, count=page_links)
            
            # If no links found, we've reached the end
            if page_links == 0:
                log("info", "No more ads found, stopping pagination", page=page_num)
                break
                
        except Exception as e:
            log("error", "Error crawling listing page", url=url, error=str(e))
            break
    
    log("info", "Completed Kleinanzeigen listing crawl", total_ads=len(ad_links), pages=page_num)
    return ad_links


async def extract_kleinanzeigen_detail_async(url: str) -> Optional[Dict[str, Any]]:
    """
    Crawl individual Kleinanzeigen ad detail page and extract contact information.
    
    Args:
        url: URL of the ad detail page
        
    Returns:
        Dict with lead data or None if extraction failed
    """
    try:
        r = await http_get_async(url, timeout=HTTP_TIMEOUT)
        if not r or r.status_code != 200:
            log("debug", "Failed to fetch detail page", url=url, status=r.status_code if r else "None")
            return None
        
        html = r.text or ""
        soup = BeautifulSoup(html, "html.parser")
        
        # Extract title
        title_elem = soup.select_one("h1#viewad-title, h1.boxedarticle--title")
        title = title_elem.get_text(" ", strip=True) if title_elem else ""
        
        # Extract description
        desc_elem = soup.select_one("#viewad-description-text, .boxedarticle--description")
        description = desc_elem.get_text(" ", strip=True) if desc_elem else ""
        
        # Combine text for extraction
        full_text = f"{title} {description}"
        
        # Extract mobile phone numbers (015x, 016x, 017x patterns)
        phones = []
        phone_matches = MOBILE_RE.findall(full_text)
        for phone_match in phone_matches:
            normalized = normalize_phone(phone_match)
            if normalized:
                is_valid, phone_type = validate_phone(normalized)
                if is_valid and is_mobile_number(normalized):
                    phones.append(normalized)
        
        # Extract email
        email = ""
        email_matches = EMAIL_RE.findall(full_text)
        if email_matches:
            email = email_matches[0]
        
        # Extract WhatsApp link
        wa_link = soup.select_one('a[href*="wa.me"], a[href*="api.whatsapp.com"]')
        whatsapp = ""
        if wa_link:
            wa_href = wa_link.get("href", "")
            # Extract phone from WhatsApp link
            wa_phone = re.sub(r'\D', '', wa_href)
            if wa_phone:
                wa_normalized = "+" + wa_phone
                is_valid, phone_type = validate_phone(wa_normalized)
                if is_valid and is_mobile_number(wa_normalized):
                    whatsapp = wa_normalized
                    if wa_normalized not in phones:
                        phones.append(wa_normalized)
        
        # Extract location/region
        location_elem = soup.select_one("#viewad-locality, .boxedarticle--details--locality")
        location = location_elem.get_text(" ", strip=True) if location_elem else ""
        
        # Extract name (from title or text)
        # Use the enhanced name extractor which handles various patterns
        name = extract_name_enhanced(full_text)
        
        # Only create lead if we found at least one mobile number
        if not phones:
            log("debug", "No mobile numbers found in ad", url=url)
            return None
        
        # Use first mobile number found
        main_phone = phones[0]
        
        # Build lead data
        lead = {
            "name": name or "",
            "rolle": "Vertrieb",  # Default role
            "email": email,
            "telefon": main_phone,
            "quelle": url,
            "score": 85,  # High score for direct Kleinanzeigen finds
            "tags": "kleinanzeigen,candidate,mobile,direct_crawl",
            "lead_type": "candidate",
            "phone_type": "mobile",
            "opening_line": title[:200] if title else "",
            "firma": "",
            "firma_groesse": "",
            "branche": "",
            "region": location if location else "",
            "frische": "neu",
            "confidence": 0.85,
            "data_quality": 0.80,
        }
        
        log("info", "Extracted lead from Kleinanzeigen ad", url=url, has_phone=bool(main_phone), has_email=bool(email))
        return lead
        
    except Exception as e:
        log("error", "Error extracting Kleinanzeigen detail", url=url, error=str(e))
        return None


def _mark_url_seen(url: str, source: str = ""):
    """
    Helper function to mark a URL as seen in the database.
    
    Args:
        url: The URL to mark as seen
        source: Optional source name for logging (e.g., "Markt.de", "Quoka")
    """
    try:
        con = db()
        cur = con.cursor()
        cur.execute("INSERT OR IGNORE INTO urls_seen (url) VALUES (?)", (url,))
        con.commit()
        con.close()
        _seen_urls_cache.add(_normalize_for_dedupe(url))
    except Exception as e:
        log_prefix = f"{source}: " if source else ""
        log("warn", f"{log_prefix}Konnte URL nicht als gesehen markieren", url=url, error=str(e))


async def crawl_markt_de_listings_async() -> List[Dict]:
    """
    Crawlt markt.de Stellengesuche-Seiten.
    
    HTML-Struktur (typisch):
    - Listing: <div class="ad-list-item"> oder <article class="result-item">
    - Link: <a href="/anzeige/...">
    - Titel: <h2> oder <h3>
    
    Pagination: ?page=2, ?page=3, etc.
    
    Returns:
        Liste von Lead-Dicts
    """
    if not DIRECT_CRAWL_SOURCES.get("markt_de", True):
        return []
    
    leads = []
    max_pages = 3  # Limit pages per URL to avoid overload
    
    for base_url in MARKT_DE_URLS:
        for page in range(1, max_pages + 1):
            if page == 1:
                url = base_url
            else:
                # Add page parameter
                separator = "&" if "?" in base_url else "?"
                url = f"{base_url}{separator}page={page}"
            
            try:
                # Rate limiting
                await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                
                log("info", "Markt.de: Listing-Seite", url=url, page=page)
                
                r = await http_get_async(url, timeout=HTTP_TIMEOUT)
                if not r or r.status_code != 200:
                    log("warn", "Markt.de: Failed to fetch", url=url, status=r.status_code if r else "None")
                    break
                
                html = r.text or ""
                soup = BeautifulSoup(html, "html.parser")
                
                # Extract ad links - try multiple selectors
                ad_links = []
                
                # Try common selectors for markt.de
                for selector in [
                    'a[href*="/anzeige/"]',
                    'a[href*="/stellengesuche/"]',
                    '.ad-list-item a',
                    'article a[href*="/anzeige/"]'
                ]:
                    links = soup.select(selector)
                    for link in links:
                        href = link.get("href", "")
                        if href and ("/anzeige/" in href or "/stellengesuche/" in href):
                            full_url = urllib.parse.urljoin("https://www.markt.de", href)
                            if full_url not in ad_links:
                                ad_links.append(full_url)
                
                log("info", "Markt.de: Anzeigen gefunden", url=url, count=len(ad_links))
                
                if not ad_links:
                    break  # No more ads, stop pagination
                
                # Extract details from each ad
                for ad_url in ad_links:
                    if url_seen(ad_url):
                        continue
                    
                    # Rate limiting
                    await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                    
                    lead = await extract_generic_detail_async(ad_url, source_tag="markt_de")
                    if lead and lead.get("telefon"):
                        leads.append(lead)
                        log("info", "Markt.de: Lead extrahiert", url=ad_url, has_phone=True)
                        
                        # Mark as seen
                        _mark_url_seen(ad_url, source="Markt.de")
                    else:
                        log("debug", "Markt.de: Keine Handynummer", url=ad_url)
                
            except Exception as e:
                log("error", "Markt.de: Fehler beim Crawlen", url=url, error=str(e))
                break
    
    log("info", "Markt.de: Crawling abgeschlossen", total_leads=len(leads))
    return leads


async def crawl_quoka_listings_async() -> List[Dict]:
    """
    Crawlt quoka.de Stellengesuche-Seiten.
    
    HTML-Struktur:
    - Listing: <li class="q-ad"> oder <div class="result-list-item">
    - Link: <a href="/stellengesuche/...">
    
    Returns:
        Liste von Lead-Dicts
    """
    if not DIRECT_CRAWL_SOURCES.get("quoka", True):
        return []
    
    leads = []
    max_pages = 3
    
    for base_url in QUOKA_DE_URLS:
        for page in range(1, max_pages + 1):
            if page == 1:
                url = base_url
            else:
                separator = "&" if "?" in base_url else "?"
                url = f"{base_url}{separator}page={page}"
            
            try:
                await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                
                log("info", "Quoka: Listing-Seite", url=url, page=page)
                
                r = await http_get_async(url, timeout=HTTP_TIMEOUT)
                if not r or r.status_code != 200:
                    log("warn", "Quoka: Failed to fetch", url=url, status=r.status_code if r else "None")
                    break
                
                html = r.text or ""
                soup = BeautifulSoup(html, "html.parser")
                
                # Extract ad links
                ad_links = []
                
                for selector in [
                    'a.q-ad-link',
                    'li.q-ad a',
                    'a[href*="/stellengesuche/"]',
                    '.result-list-item a'
                ]:
                    links = soup.select(selector)
                    for link in links:
                        href = link.get("href", "")
                        if href and "/stellengesuche/" in href:
                            full_url = urllib.parse.urljoin("https://www.quoka.de", href)
                            if full_url not in ad_links:
                                ad_links.append(full_url)
                
                log("info", "Quoka: Anzeigen gefunden", url=url, count=len(ad_links))
                
                if not ad_links:
                    break
                
                for ad_url in ad_links:
                    if url_seen(ad_url):
                        continue
                    
                    await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                    
                    lead = await extract_generic_detail_async(ad_url, source_tag="quoka")
                    if lead and lead.get("telefon"):
                        leads.append(lead)
                        log("info", "Quoka: Lead extrahiert", url=ad_url, has_phone=True)
                        
                        _mark_url_seen(ad_url, source="Quoka")
                    else:
                        log("debug", "Quoka: Keine Handynummer", url=ad_url)
                
            except Exception as e:
                log("error", "Quoka: Fehler beim Crawlen", url=url, error=str(e))
                break
    
    log("info", "Quoka: Crawling abgeschlossen", total_leads=len(leads))
    return leads


async def crawl_kalaydo_listings_async() -> List[Dict]:
    """
    Crawlt kalaydo.de Stellengesuche-Seiten.
    Kalaydo ist besonders stark im Rheinland/NRW!
    
    Returns:
        Liste von Lead-Dicts
    """
    if not DIRECT_CRAWL_SOURCES.get("kalaydo", True):
        return []
    
    leads = []
    max_pages = 3
    
    for base_url in KALAYDO_DE_URLS:
        for page in range(1, max_pages + 1):
            if page == 1:
                url = base_url
            else:
                separator = "&" if "?" in base_url else "?"
                url = f"{base_url}{separator}page={page}"
            
            try:
                await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                
                log("info", "Kalaydo: Listing-Seite", url=url, page=page)
                
                r = await http_get_async(url, timeout=HTTP_TIMEOUT)
                if not r or r.status_code != 200:
                    log("warn", "Kalaydo: Failed to fetch", url=url, status=r.status_code if r else "None")
                    break
                
                html = r.text or ""
                soup = BeautifulSoup(html, "html.parser")
                
                # Extract ad links
                ad_links = []
                
                for selector in [
                    'article.classified-ad a',
                    'a[href*="/anzeige/"]',
                    'a[href*="/stellengesuche/"]',
                    '.ad-item a'
                ]:
                    links = soup.select(selector)
                    for link in links:
                        href = link.get("href", "")
                        if href and ("/anzeige/" in href or "/stellengesuche/" in href):
                            full_url = urllib.parse.urljoin("https://www.kalaydo.de", href)
                            if full_url not in ad_links:
                                ad_links.append(full_url)
                
                log("info", "Kalaydo: Anzeigen gefunden", url=url, count=len(ad_links))
                
                if not ad_links:
                    break
                
                for ad_url in ad_links:
                    if url_seen(ad_url):
                        continue
                    
                    await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                    
                    lead = await extract_generic_detail_async(ad_url, source_tag="kalaydo")
                    if lead and lead.get("telefon"):
                        leads.append(lead)
                        log("info", "Kalaydo: Lead extrahiert", url=ad_url, has_phone=True)
                        
                        _mark_url_seen(ad_url, source="Kalaydo")
                    else:
                        log("debug", "Kalaydo: Keine Handynummer", url=ad_url)
                
            except Exception as e:
                log("error", "Kalaydo: Fehler beim Crawlen", url=url, error=str(e))
                break
    
    log("info", "Kalaydo: Crawling abgeschlossen", total_leads=len(leads))
    return leads


async def crawl_meinestadt_listings_async() -> List[Dict]:
    """
    Crawlt meinestadt.de Stellengesuche-Seiten.
    Städte-basiert, gut für lokale Kandidaten.
    
    Returns:
        Liste von Lead-Dicts
    """
    if not DIRECT_CRAWL_SOURCES.get("meinestadt", True):
        return []
    
    leads = []
    max_pages = 3
    
    for base_url in MEINESTADT_DE_URLS:
        for page in range(1, max_pages + 1):
            if page == 1:
                url = base_url
            else:
                separator = "&" if "?" in base_url else "?"
                url = f"{base_url}{separator}page={page}"
            
            try:
                await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                
                log("info", "Meinestadt: Listing-Seite", url=url, page=page)
                
                r = await http_get_async(url, timeout=HTTP_TIMEOUT)
                if not r or r.status_code != 200:
                    log("warn", "Meinestadt: Failed to fetch", url=url, status=r.status_code if r else "None")
                    break
                
                html = r.text or ""
                soup = BeautifulSoup(html, "html.parser")
                
                # Extract ad links
                ad_links = []
                
                for selector in [
                    'a[href*="/stellengesuche/anzeige/"]',
                    'a[href*="/anzeige/"]',
                    '.job-listing a',
                    'article a'
                ]:
                    links = soup.select(selector)
                    for link in links:
                        href = link.get("href", "")
                        if href and ("/stellengesuche/" in href or "/anzeige/" in href):
                            full_url = urllib.parse.urljoin("https://www.meinestadt.de", href)
                            if full_url not in ad_links:
                                ad_links.append(full_url)
                
                log("info", "Meinestadt: Anzeigen gefunden", url=url, count=len(ad_links))
                
                if not ad_links:
                    break
                
                for ad_url in ad_links:
                    if url_seen(ad_url):
                        continue
                    
                    await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                    
                    lead = await extract_generic_detail_async(ad_url, source_tag="meinestadt")
                    if lead and lead.get("telefon"):
                        leads.append(lead)
                        log("info", "Meinestadt: Lead extrahiert", url=ad_url, has_phone=True)
                        
                        _mark_url_seen(ad_url, source="Meinestadt")
                    else:
                        log("debug", "Meinestadt: Keine Handynummer", url=ad_url)
                
            except Exception as e:
                log("error", "Meinestadt: Fehler beim Crawlen", url=url, error=str(e))
                break
    
    log("info", "Meinestadt: Crawling abgeschlossen", total_leads=len(leads))
    return leads


async def extract_generic_detail_async(url: str, source_tag: str = "direct_crawl") -> Optional[Dict[str, Any]]:
    """
    Generic function to extract contact information from any ad detail page.
    Similar to extract_kleinanzeigen_detail_async but works for multiple sites.
    
    Args:
        url: URL of the ad detail page
        source_tag: Tag to identify the source (e.g., "markt_de", "quoka")
        
    Returns:
        Dict with lead data or None if extraction failed
    """
    try:
        r = await http_get_async(url, timeout=HTTP_TIMEOUT)
        if not r or r.status_code != 200:
            log("debug", f"{source_tag}: Failed to fetch detail", url=url, status=r.status_code if r else "None")
            return None
        
        html = r.text or ""
        soup = BeautifulSoup(html, "html.parser")
        
        # Extract title - try multiple selectors
        title = ""
        for selector in ["h1", "h1.title", ".ad-title", ".listing-title"]:
            title_elem = soup.select_one(selector)
            if title_elem:
                title = title_elem.get_text(" ", strip=True)
                break
        
        # Extract description - get all text from body
        description = soup.get_text(" ", strip=True)
        
        # Combine text for extraction
        full_text = f"{title} {description}"
        
        # Extract mobile phone numbers
        phones = []
        phone_matches = MOBILE_RE.findall(full_text)
        for phone_match in phone_matches:
            normalized = normalize_phone(phone_match)
            if normalized:
                is_valid, phone_type = validate_phone(normalized)
                if is_valid and is_mobile_number(normalized):
                    phones.append(normalized)
        
        # Extract email
        email = ""
        email_matches = EMAIL_RE.findall(full_text)
        if email_matches:
            email = email_matches[0]
        
        # Extract WhatsApp link
        wa_link = soup.select_one('a[href*="wa.me"], a[href*="api.whatsapp.com"]')
        if wa_link:
            wa_href = wa_link.get("href", "")
            wa_phone = re.sub(r'\D', '', wa_href)
            if wa_phone:
                wa_normalized = "+" + wa_phone
                is_valid, phone_type = validate_phone(wa_normalized)
                if is_valid and is_mobile_number(wa_normalized):
                    if wa_normalized not in phones:
                        phones.append(wa_normalized)
        
        # Extract name
        name = extract_name_enhanced(full_text)
        
        # Only create lead if we found at least one mobile number
        if not phones:
            return None
        
        # Use first mobile number found
        main_phone = phones[0]
        
        # Build lead data
        lead = {
            "name": name or "",
            "rolle": "Vertrieb",
            "email": email,
            "telefon": main_phone,
            "quelle": url,
            "score": 85,
            "tags": f"{source_tag},candidate,mobile,direct_crawl",
            "lead_type": "candidate",
            "phone_type": "mobile",
            "opening_line": title[:200] if title else "",
            "firma": "",
            "firma_groesse": "",
            "branche": "",
            "region": "",
            "frische": "neu",
            "confidence": 0.85,
            "data_quality": 0.80,
        }
        
        return lead
        
    except Exception as e:
        log("error", f"{source_tag}: Error extracting detail", url=url, error=str(e))
        return None


# =========================
# Regex/Scoring/Enrichment
# =========================

EMAIL_RE   = re.compile(r'\b(?!noreply|no-reply|donotreply)[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,24}\b', re.I)
PHONE_RE   = re.compile(r'(?:\+49|0049|0)\s?(?:\(?\d{2,5}\)?[\s\/\-]?)?\d(?:[\s\/\-]?\d){5,10}')
MOBILE_RE  = re.compile(r'(?:\+49|0049|0)\s*1[5-7]\d(?:[\s\/\-]?\d){6,10}')
SALES_RE   = re.compile(r'\b(vertrieb|vertriebs|sales|account\s*manager|key\s*account|business\s*development|aussendienst|aussendienst|handelsvertreter|telesales|call\s*center|outbound|haust?r|d2d)\b', re.I)
PROVISION_HINT = re.compile(r'\b(Provisionsbasis|nur\s*Provision|hohe\s*Provision(en)?|Leistungsprovision)\b', re.I)
D2D_HINT       = re.compile(r'\b(Door[-\s]?to[-\s]?door|Haustür|Kaltakquise|D2D)\b', re.I)
CALLCENTER_HINT= re.compile(r'\b(Call\s*Center|Telesales|Outbound|Inhouse-?Sales)\b', re.I)
B2C_HINT       = re.compile(
    r'\b(?:B2C|Privatkunden|Haushalte|Endkunden|Privatperson(?:en)?|'
    r'verschuld\w*|schuldenhilfe|schuldnerberatung|schuldner|'
    r'inkasso(?:[-\s]?f(?:ä|ae)lle?)?)\b',
    re.I,
)
JOBSEEKER_RE  = re.compile(r'\b(jobsuche|stellensuche|arbeitslos|lebenslauf|bewerb(ung)?|cv|portfolio|offen\s*f(?:ür|uer)\s*neues)\b', re.I)
CANDIDATE_TEXT_RE = re.compile(r'(?is)\b(ich\s+suche|suche\s+job|biete\s+mich|arbeitslos|stellengesuch|open\s+to\s+work)\b')
EMPLOYER_TEXT_RE  = re.compile(r'(?is)\b(wir\s+suchen|wir\s+stellen\s+ein|deine\s+aufgaben|unser\s+angebot|jetzt\s+bewerben)\b')
RECRUITER_RE  = re.compile(r'\b(recruit(er|ing)?|hr|human\s*resources|personalvermittlung|headhunter|wir\s*suchen|join\s*our\s*team)\b', re.I)

# Spezifische Handelsvertreter-Fingerprint-Phrasen
AGENT_FINGERPRINTS = (
    "selbstständiger handelsvertreter", "handelsvertretung cdh", "cdh-mitglied",
    "gemäß § 84 hgb", "gemäß §84 hgb", "industrievertretung",
    "vertriebsbüro inhaber", "auf provisionsbasis", "vertretung für plz",
    "freie handelsvertretung", "vertriebsagentur", "gebiete: plz",
    "mitglied im handelsvertreterverband", "iucab", "handelsregister a",
    "einzelkaufmann", "e.k."
)


WHATSAPP_RE    = re.compile(r'(?i)\b(WhatsApp|Whats\s*App)[:\s]*\+?\d[\d \-()]{6,}\b')
WA_LINK_RE     = re.compile(r'(?:https?://)?(?:wa\.me/\d+|api\.whatsapp\.com/send\?phone=\d+|chat\.whatsapp\.com/[A-Za-z0-9]+)', re.I)
WHATS_RE       = re.compile(r'(?:\+?\d{2,3}\s?)?(?:\(?0\)?\s?)?\d{2,4}[\s\-]?\d{3,}.*?(?:whatsapp|wa\.me|api\.whatsapp)', re.I)
WHATSAPP_PHRASE_RE = re.compile(
    r'(?i)(meldet\s+euch\s+per\s+whatsapp|schreib(?:t)?\s+mir\s+(?:per|bei)\s+whatsapp|per\s+whatsapp\s+melden)'
)
TELEGRAM_LINK_RE = re.compile(r'(?:https?://)?(?:t\.me|telegram\.me)/[A-Za-z0-9_/-]+', re.I)

CITY_RE        = re.compile(r'\b(NRW|Nordrhein[-\s]?Westfalen|Düsseldorf|Köln|Essen|Dortmund|Mönchengladbach|Bochum|Wuppertal|Bonn)\b', re.I)
NAME_RE        = re.compile(r'\b([A-ZÄÖÜ][a-zäöüß]+(?:\s+[A-ZÄÖÜ][a-zäöüß\-]+){0,2})\b')
# --- Kontext-Fenster für Heuristiken (Sales/Jobseeker) ---
# Sales-Fenster: erkennt Vertriebs-/Akquise-Kontext im Umfeld von Text
SALES_WINDOW = re.compile(
    r'(?is).{0,400}(vertrieb|verkauf|sales|akquise|außendienst|aussendienst|'
    r'call\s*center|telefonverkauf|door\s*to\s*door|d2d|provision).{0,400}'
)

# Jobseeker-Fenster: erkennt Lebenslauf/Bewerbung/Jobsuche-Kontext
JOBSEEKER_WINDOW = re.compile(
    r'(?is).{0,400}(jobsuche|stellensuche|arbeitslos|bewerb(?:ung)?|lebenslauf|'
    r'cv|portfolio|offen\s*f(?:ür|uer)\s*neues|profil).{0,400}'
)

CANDIDATE_KEYWORDS = [
    "suche job",
    "biete mich an",
    "open to work",
    "neue herausforderung",
    "gesuch",
    "lebenslauf",
]

IGNORE_KEYWORDS = [
    "wir suchen",
    "stellenanzeige",
    "gmbh",
    "hr manager",
    "karriere bei uns",
]

JOB_AD_MARKERS = [
    "(m/w/d)", "(w/m/d)", "(d/m/w)", "(m/f/d)", "(f/m/d)",
    "wir suchen", "we are hiring", "stellenangebot", "jobangebot",
    "ausbildungsplatz", "azubi", "ausbildung zum",
    "teamleiter", "sales manager", "mitarbeiter",
    "gesucht", "join us", "karriere bei", "vacancies",
    "wir bieten", "deine aufgaben", "ihr profil", "your profile", "benefits", "corporate benefits",
    "obstkorb", "flache hierarchien", "bewerben sie sich", "apply now",
]

STRICT_JOB_AD_MARKERS = [
    "(m/w/d)", "(w/m/d)", "(d/m/w)", "(m/f/d)", "(f/m/d)", "(gn)",
    "wir suchen", "we are hiring", "wir stellen ein", "join our team",
    "ausbildungsplatz", "azubi gesucht", "ab sofort gesucht",
    "ihre aufgaben", "wir bieten", "dein profil", "your profile", "benefits:",
    "stellenanzeige", "jobangebot", "vacancy", "karriere bei",
    "teamleiter gesucht", "sales manager gesucht", "benefits", "corporate benefits",
]

# CANDIDATE POSITIVE SIGNALS - Menschen die Jobs SUCHEN (dürfen NICHT als job_ad geblockt werden!)
CANDIDATE_POSITIVE_SIGNALS = [
    "suche job",
    "suche arbeit",
    "suche stelle",
    "suche neuen job",
    "suche neue stelle",
    "ich suche",
    "stellengesuch",
    "auf jobsuche",
    "offen für angebote",
    "offen für neue",
    "suche neue herausforderung",
    "suche neuen wirkungskreis",
    "verfügbar ab",
    "freigestellt",
    "open to work",
    "#opentowork",
    "looking for opportunities",
    "seeking new",
    "jobsuchend",
    "arbeitslos",
    "gekündigt",
    "wechselwillig",
    "bin auf der suche",
    "suche eine neue",
]

# JOB OFFER SIGNALS - Firmen die Mitarbeiter SUCHEN (SOLLEN geblockt werden)
JOB_OFFER_SIGNALS = [
    "(m/w/d)",
    "(w/m/d)",
    "wir suchen",
    "gesucht:",
    "ab sofort zu besetzen",
    "stellenangebot",
    "job zu vergeben",
    "mitarbeiter gesucht",
    "verstärkung gesucht",
    "team sucht",
    "für unser team",
]

# Minimum number of job offer signals required to override a candidate signal
MIN_JOB_OFFER_SIGNALS_TO_OVERRIDE = 2

HIRING_INDICATORS = (
    "wir suchen", "suchen wir", "wir stellen ein", "join our team",
    "deine aufgaben", "ihr profil", "your profile", "wir bieten", "benefits",
    "bewerben sie sich", "jetzt bewerben", "apply now",
    "stellenanzeige", "jobangebot", "vacancy", "karriere bei",
    "(m/w/d)", "(m/f/d)", "(gn)", "einstiegsgehalt", "festanstellung",
)

SOLO_BIZ_INDICATORS = (
    # Original
    "handelsvertretung", "handelsvertreter", "selbstständiger vertriebspartner",
    "industrievertretung", "agentur für", "vertriebsbüro",
    # NEW from Research
    "§ 84 hgb", "§84 hgb", "cdh-mitglied", "cdh mitglied",
    "freie handelsvertretung", "vertriebsunternehmer", "eigenständige vertriebsunternehmung",
    "handelsagentur", "vertriebsrepräsentanz", "gebietsvertretung",
    "provisionsbasis", "courtage-vereinbarung", "übernehme vertretungen",
    "inhaltlich verantwortlicher", "einzelunternehmen", "einzelkaufmann",
    "registriert bei handelsvertreter.de", "iucab"
)

RETAIL_ROLES = [
    "kuechen", "kueche", "moebel", "einzelhandel", "verkaeufer", "verkaeuferin",
    "kassierer", "servicekraft", "filialleiter", "shop manager", "baecker",
    "metzger", "floor manager", "call center", "telefonist", "promoter",
]

def is_likely_human_name(text: str) -> bool:
    """Heuristic: 2-3 words, no digits, no company/ad markers."""
    if not text:
        return False
    s = re.sub(r"\s+", " ", text).strip(" ,;|")
    if not s:
        return False
    lower = s.lower()
    corporate_tokens = (
        "gmbh", "ag", "kg", "ug", "co. kg", "inc", "ltd", "gbr",
        "unternehmen", "firma", "company", "holding",
    )
    ad_tokens = (
        "team", "karriere", "job", "stellenanzeige", "gesucht", "wir suchen",
        "abteilung",
    )
    if any(tok in lower for tok in corporate_tokens + ad_tokens):
        return False
    if re.search(r"\d", s):
        return False
    words = [w for w in re.split(r"[\s\-\|_]+", s) if w]
    if not (2 <= len(words) <= 3):
        return False
    if any(len(w) == 1 for w in words):
        return False
    return True

def looks_like_company(text: str) -> bool:
    if not text:
        return False
    tl = text.lower()
    company_tokens = [
        "gmbh", "ag", "kg", "ug", "ltd", "inc", "team", "firma", "unternehmen",
        "holding", "group", "gbr", "gesucht"
    ]
    return any(tok in tl for tok in company_tokens)

EMPLOYER_EMAIL_PREFIXES = {
    "jobs", "job", "karriere", "career", "recruiting", "recruit",
    "bewerbung", "application", "hr", "humanresources", "personal",
    "info", "kontakt", "contact", "service", "support", "office",
    "gf", "geschaeftsfuehrung", "management", "hello", "mail",
    "team", "admin", "datenschutz",
}

ROLE_DEFINITIONS = {
    "high_value": ["account manager", "key account", "vertriebsleiter", "head of sales", "aussendienst", "handelsvertreter", "sales manager"],
    "retail_craft": ["verkaeufer", "verkaeuferin", "kuechen", "kuechenfachberater", "einzelhandel", "baeckerei", "kassierer", "servicekraft", "promoter", "call center", "telefonist", "moebel"],
    "junior": ["trainee", "praktikant", "werkstudent", "azubi", "junior", "einstieg"],
}

INTENT_TITLE_KEYWORDS = (
    "gesuch",
    "suche",
    "biete",
    "bewerbung",
    "lebenslauf",
    "profil",
    "cv",
    "curriculum vitae",
    "kenntnisse",
    "qualifikation",
    "verfuegbar",
    "freelancer",
    "stellengesuch",
    "bewerber",
    "open to work",
    "mich vor",
    "ueber mich",
    "kurzprofil",
)

# --- COMPLETING THE MISSING CONSTANTS ---

# Positive signals in URL path (if present, we scan it even if it looks weird)
ALLOW_PATH_HINTS = (
    "/profil", "/profile", "/cv", "/lebenslauf", "/resume",
    "/user", "/nutzer", "/candidate", "/bewerber",
    "/in/", "/pub/", "/p/"  # Social media profile markers
)

# Hints that suggest a low-paying or irrelevant job (used for scoring)
LOW_PAY_HINT = (
    "mindestlohn", "minijob", "aushilfe", "student", "praktikum",
    "ferienjob", "saison", "helfer", "azubi", "ausbildung",
    "promoter", "kommission"
)

# Hints that suggest high commission or sales focus (used for scoring)
COMMISSION_HINT = (
    "provision", "bonus", "uncapped", "hohe verdienstmoeglichkeiten",
    "abschlussstark", "hunter", "akquise", "vertriebsstark"
)

# Initialize missing filter lists if they don't exist yet (Safety fallback)
if 'BAD_DOMAINS' not in globals():
    BAD_DOMAINS = {"example.com", "localhost"}
if 'IGNORE_EXTENSIONS' not in globals():
    IGNORE_EXTENSIONS = (".jpg", ".png", ".css", ".js")

NEG_PATH_HINTS = (
    "/jobs/", "/job/", "/stellenangebot/", "/company/", "/companies/",
    "/unternehmen/", "/business/", "/school/", "/university/",
    "/pulse/", "/article/", "/learning/", "/groups/", "/feed/",
    "/events/", "/search/", "/salary/", "/gehalt/",
)

IGNORE_EXTENSIONS = (
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".webp",
    ".css", ".js", ".json", ".xml", ".ico", ".mp4", ".mp3",
    ".woff", ".woff2", ".ttf", ".eot", ".zip", ".rar", ".7z",
    ".tar", ".gz", ".exe", ".msi", ".apk", ".dmg", ".iso",
)

BAD_TITLE_KEYWORDS = (
    "jobs", "stellenangebote", "karriere", "career", "vacancies",
    "wir suchen", "gesucht", "hiring", "jobbörse", "company",
    "unternehmen", "über uns", "about us", "impressum", "kontakt",
    "datenschutz", "login", "register", "anmelden", "sign up",
    "ag", "gmbh", "k.g.", "limited", "inc.",
)

DENY_DOMAINS = {
    "stepstone.de", "indeed.com", "monster.de", "arbeitsagentur.de",
    "xing.com/jobs", "linkedin.com/jobs", "meinestadt.de", "kimeta.de",
    "jobware.de", "stellenanzeigen.de", "absolventa.de", "glassdoor.de",
    "kununu.com", "azubiyo.de", "ausbildung.de", "gehalt.de", "lehrstellen-radar.de",
    "wikipedia.org", "youtube.com", "amazon.de", "ebay.de",
    "heyjobs.de", "heyjobs.co", "softgarden.io", "jobijoba.de", "jobijoba.com",
    "ok.ru", "tiktok.com", "patents.google.com",
}

DENY_DOMAINS.update({
    "stellenonline.de", "stellenmarkt.de", "stepstone.de", "indeed.com",
    "meinestadt.de", "kimeta.de", "jobware.de", "monster.de",
    "arbeitsagentur.de", "nadann.de", "freshplaza.de", "kikxxl.de",
    "xing.com/jobs", "linkedin.com/jobs", "gehalt.de", "kununu.com",
    "ausbildung.de", "azubiyo.de", "lehrstellen-radar.de",
    "softgarden.io", "jobijoba.de", "jobijoba.com", "heyjobs.de", "heyjobs.co",
})


NEGATIVE_HINT  = re.compile(r'\b(Behörde|Amt|Universität|Karriereportal|Blog|Ratgeber|Software|SaaS|Bank|Versicherung)\b', re.I)
WHATSAPP_INLINE= re.compile(r'\+?\d[\d ()\-]{6,}\s*(?:WhatsApp|WA)', re.I)
PERSON_PREFIX  = re.compile(r'\b(Herr|Frau|Hr\.|Fr\.)\s+[A-ZÄÖÜ][a-zäöüß\-]+(?:\s+[A-ZÄÖÜ][a-zäöüß\-]+)?')
JOBSEEKER_WINDOW = re.compile(
    r'(?is).{0,400}(jobsuche|stellensuche|arbeitslos|bewerb(ung)?|lebenslauf|cv|portfolio|offen\s*f(?:ür|uer)\s*neues).{0,400}'
)
INDUSTRY_PATTERNS = {
    "energie":      r'\b(Energie|Strom|Gas|Ökostrom|Versorger|Photovoltaik|PV|Solar|Wärmepumpe)\b',
    "telekom":      r'\b(Telekommunikation|Telefonie|Internet|DSL|Mobilfunk|Glasfaser|Telekom)\b',
    "versicherung": r'\b(Versicherung(en)?|Versicherungsmakler|Bausparen|Finanzberatung|Finanzen)\b',
    "bau":          r'\b(Bau|Handwerk|Sanierung|Fenster|Türen|Dämm|Energieberatung)\b',
    "ecommerce":    r'\b(E-?Commerce|Onlineshop|Bestellhotline|Kundengewinnung)\b',
    "household":    r'\b(Vorwerk|Kobold|Staubsauger|Haushaltswaren)\b'
}

CONTACT_HINTS = [
    "kontakt", "ansprechpartner", "e-mail", "email", "mail",
    "telefon", "tel", "tel.", "whatsapp", "telegram", "anruf", "hotline"
]

SOCIAL_HOSTS = {
    "linkedin.com", "www.linkedin.com", "de.linkedin.com",
    "xing.com", "www.xing.com",
    "facebook.com", "www.facebook.com",
    "instagram.com", "www.instagram.com",
    "twitter.com", "x.com",
    "tiktok.com",
}

INDUSTRY_HINTS = [
    "solar", "photovoltaik", "pv", "energie", "energieberatung", "strom", "gas",
    "glasfaser", "telekom", "telekommunikation", "dsl", "mobilfunk", "internet",
    "vorwerk", "kobold", "staubsauger", "haushaltswaren",
    "fenster", "türen", "tueren", "dämm", "daemm", "wärmepumpe", "waermepumpe",
    "versicherung", "versicherungen", "bausparen", "immobilien", "makler",
    "onlineshop", "e-commerce", "shop", "kundengewinnung"
]

BAD_MAILBOXES = {
    "noreply", "no-reply", "donotreply", "do-not-reply",
    "info", "kontakt", "contact", "office", "service", "support", "news", "presse",
    "bewerbung", "recruiting", "karriere", "jobs", "hr", "humanresources",
    "talent", "people", "personal", "datenschutz", "privacy"
}

PORTAL_DOMAINS = {
    "adecco.de",
    "arbeitsagentur.de",
    "glassdoor.de",
    "hays.de",
    "heyjobs.co",
    "heyjobs.de",
    "indeed.com",
    "jobware.de",
    "join.com",
    "kimeta.de",
    "kununu.com",
    "linkedin.com",
    "meinestadt.de",
    "monster.de",
    "randstad.de",
    "softgarden.io",
    "stellenanzeigen.de",
    "stepstone.de",
    "talents.studysmarter.de",
    "xing.com",
    "jobijoba.de",
    "jobijoba.com",
    "ok.ru",
    "patents.google.com",
    "tiktok.com",
}
DROP_MAILBOX_PREFIXES = {
    "info",
    "kontakt",
    "contact",
    "support",
    "service",
    "privacy",
    "datenschutz",
    "noreply",
    "no-reply",
    "donotreply",
    "do-not-reply",
    "jobs",
    "karriere",
}
DROP_PORTAL_DOMAINS = {
    "stepstone.de",
    "indeed.com",
    "heyjobs.co",
    "heyjobs.de",
    "softgarden.io",
    "jobijoba.de",
    "jobijoba.com",
    "jobware.de",
    "monster.de",
    "kununu.com",
    "ok.ru",
    "tiktok.com",
    "patents.google.com",
    "linkedin.com",
    "xing.com",
    "arbeitsagentur.de",
    "meinestadt.de",
    "kimeta.de",
    "stellenanzeigen.de",
    "bewerbung.net",
    "freelancermap.de",
    "reddit.com",
    "jobboard-deutschland.de",
    "kleinanzeigen.de",
    "praca.egospodarka.pl",
    "tabellarischer-lebenslauf.net",
    "lexware.de",
    "tribeworks.de",
    "junico.de",
    "qonto.com",
    "accountable.de",
    "sevdesk.de",
    "mlp.de",
    "netspor-tv.com",
    "trendyol.com",
}
DROP_PORTAL_PATH_FRAGMENTS = ("linkedin.com/jobs", "xing.com/jobs")
IMPRINT_PATH_RE = re.compile(r"/(impressum|datenschutz|privacy|agb)(?:/|\\?|#|$)", re.I)
CV_HINT_RE = re.compile(r"\b(lebenslauf|curriculum vitae|cv)\b", re.I)
ALLOW_PDF_NON_CV = (os.getenv("ALLOW_PDF_NON_CV", "0") == "1")

# Blacklist path/title patterns for pre-fetch filtering (case-insensitive)
BLACKLIST_PATH_PATTERNS = {
    "lebenslauf", "vorlage", "muster", "sitemap", "seminar", 
    "academy", "weiterbildung", "job", "stellenangebot", 
    "news", "blog", "ratgeber", "portal"
}

# Whitelist for promising patterns - always allow these through
ALWAYS_ALLOW_PATTERNS = [
    r'industrievertretung',
    r'handelsvertret',
    r'vertriebspartner',
    r'/ansprechpartner/',
    r'/team/',
    r'/ueber-uns/',
    r'/kontakt/',
    r'/impressum',
]

def _matches_hostlist(host: str, blocked: set[str]) -> bool:
    h = (host or "").lower()
    if h.startswith("www."):
        h = h[4:]
    return any(h == d or h.endswith("." + d) for d in (b.lower() for b in blocked))


def _is_candidates_mode() -> bool:
    """Check if we're in candidates/recruiter mode based on INDUSTRY env var."""
    industry = str(os.getenv("INDUSTRY", "")).lower()
    return "recruiter" in industry or "candidates" in industry


def is_candidate_url(url: Optional[str]) -> Optional[bool]:
    """
    Prüft ob URL ein Kandidaten-Profil sein könnte.
    Returns: True (candidate), False (definitely not), None (uncertain, needs further analysis)
    """
    if not url:
        return None
    
    url_lower = url.lower()
    
    # CRITICAL FIX: Check for candidate keywords first (highest priority)
    # Don't block /jobs/ or /stellenangebote/ if it contains candidate keywords
    if 'stellengesuch' in url_lower or 'jobgesuch' in url_lower:
        return True
    
    # POSITIV - Kandidaten-URLs (ALWAYS allow these!)
    positive_patterns = [
        '/s-stellengesuche/',       # Kleinanzeigen Stellengesuche
        '/stellengesuche/',          # Andere Portale Stellengesuche  
        'linkedin.com/in/',          # LinkedIn Profile (nicht /jobs/ oder /company/)
        'xing.com/profile/',         # Xing Profile
        '/freelancer/',              # Freelancer Profile
        'facebook.com/groups/',      # Facebook Gruppen
        't.me/',                     # Telegram Gruppen
        'chat.whatsapp.com/',        # WhatsApp Gruppen
        'reddit.com/r/',             # Reddit Threads
        'gutefrage.net',             # Fragen zu Jobsuche
        'freelancermap.de',          # Freelancer Portale
        'freelance.de',              # Freelancer Portale
        'gulp.de',                   # Freelancer Portale
    ]
    
    for pos in positive_patterns:
        if pos in url_lower:
            return True
    
    # NEGATIV - Only block if definitely NOT a candidate
    # NOTE: Candidate keywords (stellengesuch/jobgesuch) are checked FIRST above,
    # so /jobs/ URLs with those keywords will already have returned True by this point.
    # This means we can safely block generic /jobs/ URLs here without affecting candidate URLs.
    negative_patterns = [
        '/company/',                # Firmen-Profile
        '/impressum',               # Firmen-Impressum
        '/kontakt',                 # Firmen-Kontakt
        '/about',                   # Über uns Seiten
        '/karriere/stellenangebote', # Firmen-Karriereseiten mit Stellenangeboten
        '/karriere/jobs',           # Firmen-Karriereseiten mit Jobs
        '/stellenangebote/',        # Firmen-Stellenangebote (without "stellengesuch")
        '/jobs/',                   # Generic job listings (without "stellengesuch")
        'stepstone.de',             # Job boards
        'indeed.com',               # Job boards
        'monster.de',               # Job boards
        'linkedin.com/jobs/',       # Job listings
        'xing.com/jobs/',           # Job listings
        'jobboerse',                # Jobbörsen
    ]
    
    for neg in negative_patterns:
        if neg in url_lower:
            return False
    
    # Default: Nicht sicher, weitere Prüfung nötig
    return None


def should_skip_url_prefetch(url: str, title: str = "", snippet: str = "", is_snippet_jackpot: bool = False) -> Tuple[bool, str]:
    """
    Pre-fetch URL filtering: check blacklist hosts and path patterns.
    Snippet jackpots with contact data are always allowed through (unless job postings).
    Returns (should_skip, reason).
    """
    if not url:
        return False, ""
    
    try:
        parsed = urllib.parse.urlparse(url)
        host = (parsed.netloc or "").lower()
        path = (parsed.path or "").lower()
        url_lower = url.lower()
        title_lower = (title or "").lower()
        snippet_lower = (snippet or "").lower()
        
        # Check for job posting - always block
        if is_job_posting(url=url, title=title, snippet=snippet):
            return True, "job_posting"
        
        # In candidates mode, use candidate URL filtering
        if _is_candidates_mode():
            candidate_check = is_candidate_url(url)
            if candidate_check is False:
                # Definitely not a candidate URL
                return True, "not_candidate_url"
            elif candidate_check is True:
                # Definitely a candidate URL, allow through
                return False, ""
            # If None, continue with normal filtering
        
        # Check whitelist first - always allow promising patterns
        combined_text = f"{url_lower} {path} {title_lower}"
        for pattern in ALWAYS_ALLOW_PATTERNS:
            if re.search(pattern, combined_text, re.I):
                return False, ""  # Allow through
        
        # Snippet jackpots with contact data always pass (unless job posting - already checked)
        if is_snippet_jackpot:
            return False, ""
        
        # Check host blacklist
        if _matches_hostlist(host, DROP_PORTAL_DOMAINS):
            return True, "blacklist_host"
        
        # Check path/title patterns (case-insensitive)
        for pattern in BLACKLIST_PATH_PATTERNS:
            if pattern in combined_text:
                return True, f"blacklist_pattern_{pattern}"
        
        return False, ""
    except Exception:
        return False, ""


def should_drop_lead(lead: Dict[str, Any], page_url: str, text: str = "", title: str = "") -> Tuple[bool, str]:
    email = (lead.get("email") or "").strip().lower()
    url_lower = (page_url or "").lower()
    text_lower = (text or "").lower()
    title_lower = (title or "").lower()
    host = (urllib.parse.urlparse(page_url or "").netloc or "").lower()
    person_blob = " ".join([lead.get("name", ""), lead.get("rolle", "")]).strip()

    def _drop(reason: str) -> Tuple[bool, str]:
        log("debug", "lead dropped", reason=reason, url=page_url)
        return True, reason
    
    # CRITICAL: Check for job postings first - NEVER save as lead
    if is_job_posting(url=page_url, title=title, snippet=lead.get("opening_line", ""), content=text):
        return _drop("job_posting")

    # Telefonnummer Pflicht - strict validation
    phone = (lead.get("telefon") or "").strip()
    if not phone:
        return _drop("no_phone")
    
    is_valid, phone_type = validate_phone(phone)
    if not is_valid:
        return _drop("no_phone")
    
    # STRICT: Only mobile numbers allowed
    # Normalize phone first before checking if it's mobile
    normalized_phone = normalize_phone(phone)
    if not is_mobile_number(normalized_phone):
        return _drop("not_mobile_number")

    if email:
        local, _, domain = email.partition("@")
        if local in DROP_MAILBOX_PREFIXES:
            return _drop("generic_mailbox")
        if domain and _matches_hostlist(domain, DROP_PORTAL_DOMAINS):
            return _drop("portal_domain")

    host_is_portal = _matches_hostlist(host, DROP_PORTAL_DOMAINS)
    if host_is_portal and host in {"linkedin.com", "www.linkedin.com", "xing.com", "www.xing.com"}:
        if "/jobs" not in url_lower:
            host_is_portal = False
    if host_is_portal or any(frag in url_lower for frag in DROP_PORTAL_PATH_FRAGMENTS):
        return _drop("portal_host")

    if IMPRINT_PATH_RE.search(url_lower):
        has_phone = bool(lead.get("telefon") or PHONE_RE.search(text_lower))
        has_messenger = bool(
            WA_LINK_RE.search(text_lower) or WHATS_RE.search(text_lower) or
            WHATSAPP_RE.search(text_lower) or TELEGRAM_LINK_RE.search(text_lower)
        )
        has_email = bool(lead.get("email") or EMAIL_RE.search(text or ""))
        has_person = (
            is_likely_human_name(person_blob) or
            len(person_blob.split()) >= 2 or
            bool(NAME_RE.search(text or ""))
        )
        if not (has_person or has_phone or has_messenger):
            if not has_email:
                return _drop("impressum_no_contact")

    if url_lower.endswith(".pdf"):
        hint_blob = " ".join([url_lower, person_blob, title_lower])
        has_cv_hint = bool(
            CV_HINT_RE.search(text_lower) or
            CV_HINT_RE.search(hint_blob) or
            CV_HINT_RE.search(title_lower)
        )
        if not has_cv_hint and not ALLOW_PDF_NON_CV:
            return _drop("pdf_without_cv_hint")

    return False, ""

GENERIC_BOXES = {"sales", "vertrieb", "verkauf", "marketing", "kundenservice", "hotline"}

def etld1(host: str) -> str:
    if not host:
        return ""
    ex = tldextract.extract(host)
    dom = getattr(ex, "top_domain_under_public_suffix", None) or ex.registered_domain
    return dom.lower() if dom else host.lower()


def _dedup_run(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Entdoppelt pro Run anhand (eTLD+1 der Quelle, normalisierte E-Mail, normalisierte Tel).
    Spart sinnlose DB-Inserts/Exports.
    """
    seen: set[Tuple[str, str, str]] = set()
    out: List[Dict[str, Any]] = []
    for r in rows:
        try:
            dom = etld1(urllib.parse.urlparse(r.get("quelle","")).netloc)
        except Exception:
            dom = ""
        e = (r.get("email") or "").strip().lower()
        t = re.sub(r'\D','', r.get("telefon") or "")
        key = (dom, e, t)
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def same_org_domain(page_url: str, email_domain: str) -> bool:
    try:
        host = urllib.parse.urlparse(page_url).netloc
        return etld1(host) == etld1(email_domain)
    except Exception:
        return False

_OBFUSCATION_PATTERNS = [
    (r'\s*(\[\s*at\s*\]|\(\s*at\s*\)|\{\s*at\s*\}|\s+at\s+|\s+ät\s+)\s*', '@'),
    (r'\s*(\[\s*dot\s*\]|\(\s*dot\s*\)|\{\s*dot\s*\}|\s+dot\s+|\s+punkt\s*|\s*\.\s*)\s*', '.'),
    (r'\s*(ät|@t)\s*', '@'),   # zusätzliche Varianten
]

def deobfuscate_text_for_emails(text: str) -> str:
    s = text
    for pat, rep in _OBFUSCATION_PATTERNS:
        s = re.sub(pat, rep, s, flags=re.I)
    return s

def email_quality(email: str, page_url: str) -> str:
    if not email:
        return "reject"
    low = email.strip().lower()
    local, _, domain = low.partition("@")
    if local in EMPLOYER_EMAIL_PREFIXES:
        return "reject"
    if any(domain.endswith(p) for p in PORTAL_DOMAINS):
        return "reject"
    if local in BAD_MAILBOXES:
        return "reject"
    if page_url and not same_org_domain(page_url, domain):
        if domain in {"gmail.com", "outlook.com", "hotmail.com", "gmx.de", "web.de"}:
            return "weak"
        return "reject"
    if re.match(r'^[a-z]\.?[a-z]+(\.[a-z]+)?$', local) or re.match(r'^[a-z]+[._-][a-z]+$', local):
        return "personal"
    if local in GENERIC_BOXES:
        return "team"
    return "generic"

def is_employer_email(email: str) -> bool:
    """Detect recruiter/company mailboxes by local-part prefixes."""
    if not email or "@" not in email:
        return False
    local = email.split("@", 1)[0].lower()
    normalized_local = re.sub(r"[._\\-]+", "", local)
    return any(normalized_local.startswith(pref) for pref in EMPLOYER_EMAIL_PREFIXES)

def is_candidate_seeking_job(text: str = "", title: str = "", url: str = "") -> bool:
    """Prüft ob es sich um einen jobsuchenden Kandidaten handelt (darf NICHT geblockt werden!)."""
    text_lower = (text + " " + title).lower()
    
    # ERST prüfen: Ist es ein KANDIDAT der einen Job SUCHT?
    for signal in CANDIDATE_POSITIVE_SIGNALS:
        if signal in text_lower:
            return True  # Das ist ein Kandidat! Nicht blocken!
    
    return False


# =========================
# NEW CANDIDATE-FOCUSED VALIDATION FUNCTIONS
# =========================

def _validate_name_heuristic(name: str) -> Tuple[bool, int, str]:
    """Fallback-Heuristik für Namensvalidierung ohne KI."""
    if not name or len(name.strip()) < 3:
        return False, 0, "Name zu kurz"
    
    name_lower = name.lower().strip()
    
    # Blacklist
    blacklist = [
        "gmbh", "ag", "kg", "ug", "ltd", "inc",
        "team", "vertrieb", "sales", "info", "kontakt",
        "ansprechpartner", "unknown", "n/a", "k.a.",
        "firma", "unternehmen", "company", "abteilung",
    ]
    if any(b in name_lower for b in blacklist):
        return False, 90, f"Blacklist-Wort gefunden"
    
    # Muss mindestens 2 Wörter haben
    words = name.split()
    if len(words) < 2:
        return False, 70, "Nur ein Wort"
    
    # Keine Zahlen
    if re.search(r'\d', name):
        return False, 85, "Enthält Zahlen"
    
    return True, 75, "Heuristik: wahrscheinlich echter Name"


def _looks_like_company_name(name: str) -> bool:
    """
    Prüft ob Name wie eine Firma aussieht.
    
    Returns True für:
        - "GmbH", "AG", "KG", "UG" im Namen
        - "Team", "Vertrieb", "Sales" im Namen
        - Nur ein Wort
        - Enthält Zahlen
    """
    if not name:
        return True
    
    company_indicators = [
        "gmbh", "ag", "kg", "ug", "ltd", "inc", 
        "team", "vertrieb", "sales", "group", "holding",
        "firma", "unternehmen", "company"
    ]
    name_lower = name.lower()
    
    # Check for company indicators
    if any(ind in name_lower for ind in company_indicators):
        return True
    
    # Check for single word (not a full name)
    if len(name.split()) < 2:
        return True
    
    # Check for numbers
    if re.search(r'\d', name):
        return True
    
    return False


async def validate_real_name_with_ai(name: str, context: str = "") -> Tuple[bool, int, str]:
    """
    KI-Prüfung ob der Name ein echter menschlicher Name ist.
    
    Returns: (is_real_name, confidence, reason)
    """
    if not OPENAI_API_KEY:
        # Fallback: Einfache Heuristik
        return _validate_name_heuristic(name)
    
    prompt = f"""
    Prüfe ob dieser Name ein ECHTER menschlicher Name ist:
    
    Name: "{name}"
    Kontext: "{context}"
    
    ECHTE Namen:
    ✅ "Max Mustermann" - Vor- und Nachname
    ✅ "Anna-Maria Schmidt" - Doppelname
    ✅ "Dr. Hans Meier" - Mit Titel
    ✅ "Mehmet Yılmaz" - Internationale Namen
    
    KEINE echten Namen:
    ❌ "Vertrieb NRW" - Firmenname
    ❌ "Team Sales" - Abteilung
    ❌ "info" - Generisch
    ❌ "123456" - Zahlen
    ❌ "Ansprechpartner" - Rolle
    ❌ "GmbH" - Firma
    ❌ "Unknown Candidate" - Platzhalter
    ❌ "N/A", "k.A.", "-" - Leer
    
    Return JSON: {{"is_real_name": true/false, "confidence": 0-100, "reason": "..."}}
    """
    
    try:
        import openai
        client = openai.AsyncOpenAI(api_key=OPENAI_API_KEY)
        response = await client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.1,
            max_tokens=150
        )
        
        result = json.loads(response.choices[0].message.content)
        is_real = result.get("is_real_name", False)
        confidence = result.get("confidence", 0)
        reason = result.get("reason", "")
        
        return (is_real, confidence, reason)
    except Exception as e:
        log("warn", "AI name validation failed, using heuristic", error=str(e))
        return _validate_name_heuristic(name)


def analyze_wir_suchen_context(text: str, url: str = "") -> Tuple[str, str]:
    """
    Analysiert "wir suchen" im Kontext.
    
    Returns: (classification, reason)
    - "job_ad" = Stellenanzeige (BLOCK)
    - "candidate" = Kandidat sucht (ALLOW)
    - "business_inquiry" = Geschäftsanfrage (ALLOW mit Prüfung)
    - "unclear" = Unklar (weitere Prüfung)
    """
    text_lower = text.lower()
    
    wir_suchen_matches = list(re.finditer(r'wir\s+suchen', text_lower))
    
    if not wir_suchen_matches:
        return "unclear", "Kein 'wir suchen' gefunden"
    
    for match in wir_suchen_matches:
        start = max(0, match.start() - 200)
        end = min(len(text_lower), match.end() + 300)
        context = text_lower[start:end]
        
        # === STELLENANZEIGE (BLOCK) ===
        job_ad_signals = [
            "(m/w/d)", "(w/m/d)", "(d/m/w)",
            "zum nächstmöglichen", "zur verstärkung",
            "für unser team", "ihre aufgaben", "ihr profil",
            "wir bieten", "benefits", "festanstellung",
            "vollzeit", "teilzeit", "homeoffice möglich",
            "bewerbung an", "bewerben sie sich",
        ]
        job_ad_count = sum(1 for signal in job_ad_signals if signal in context)
        
        if job_ad_count >= 2:
            return "job_ad", f"Stellenanzeige ({job_ad_count} Signale)"
        
        # === KANDIDAT SUCHT (ALLOW) ===
        candidate_signals = [
            "wir suchen ein unternehmen", "wir suchen eine firma",
            "wir suchen aufträge", "wir suchen neue kunden",
            "wir suchen vertretungen", "wir suchen partner",
            "wir suchen projekte", "ich und mein team suchen",
            "wir als handelsvertretung suchen",
            "wir als freelancer suchen",
        ]
        for signal in candidate_signals:
            if signal in context:
                return "candidate", f"Kandidat/Team sucht: '{signal}'"
        
        # === EINZELPERSON (Pluralis Majestatis) ===
        solo_signals = [
            "biete meine dienste", "meine erfahrung",
            "mein profil", "ich bin", "verfügbar ab",
            "kontaktieren sie mich",
        ]
        for signal in solo_signals:
            if signal in context:
                return "candidate", f"Einzelperson ('{signal}' im Kontext)"
        
        # === BUSINESS INQUIRY ===
        business_signals = [
            "suchen lieferanten", "suchen hersteller",
            "suchen dienstleister", "suchen zusammenarbeit",
        ]
        for signal in business_signals:
            if signal in context:
                return "business_inquiry", f"Geschäftsanfrage: '{signal}'"
    
    return "unclear", "Kontext nicht eindeutig"


def detect_hidden_gem(text: str, url: str = "") -> Tuple[Optional[str], int, str]:
    """
    Erkennt Quereinsteiger mit Vertriebspotenzial.
    
    Returns: (gem_type, confidence, reason) oder (None, 0, "")
    """
    text_lower = text.lower()
    
    for gem_type, pattern in HIDDEN_GEMS_PATTERNS.items():
        keyword_hits = sum(1 for k in pattern["keywords"] if k in text_lower)
        positive_hits = sum(1 for p in pattern["positive"] if p in text_lower)
        
        if keyword_hits >= 1 and positive_hits >= 1:
            confidence = min(95, 50 + keyword_hits * 15 + positive_hits * 10)
            return gem_type, confidence, pattern["reason"]
    
    return None, 0, ""

def has_nrw_signal(text: str) -> bool:
    """Prüft ob Text NRW-Bezug hat."""
    text_lower = text.lower()
    return any(region in text_lower for region in NRW_REGIONS)

def is_job_advertisement(text: str = "", title: str = "", snippet: str = "") -> bool:
    """Return True if content/title/snippet contains strict job-ad markers (company hiring)."""
    combined = " ".join([(text or ""), (title or ""), (snippet or "")]).lower()
    
    # FIRST: Check if this is a CANDIDATE seeking a job - NEVER block candidates!
    if is_candidate_seeking_job(text, title):
        # If candidate signal is found, only mark as job ad if there are MULTIPLE strong job offer signals
        job_offer_count = sum(1 for offer in JOB_OFFER_SIGNALS if offer in combined)
        if job_offer_count < MIN_JOB_OFFER_SIGNALS_TO_OVERRIDE:
            return False  # It's a candidate, not a job ad!
    
    # THEN: Check for job offer signals
    return any(m in combined for m in STRICT_JOB_AD_MARKERS)

def classify_role(text: str = "", title: str = "") -> str:
    """Classify lead into role category for tagging."""
    combined = " ".join([(title or ""), (text or "")]).lower()
    def _match(keywords):
        return any(k in combined for k in keywords)
    if _match(ROLE_DEFINITIONS.get("high_value", [])):
        return "pro_sales"
    if _match(ROLE_DEFINITIONS.get("retail_craft", [])):
        return "retail_sales"
    if _match(ROLE_DEFINITIONS.get("junior", [])):
        return "junior_sales"
    return "general_sales"

def analyze_page_intent(text_lower: str, title_lower: str) -> Tuple[bool, str]:
    """Distinguish hiring/job ads from self-presentations of reps/agents."""
    hiring_score = sum(1 for phrase in HIRING_INDICATORS if phrase in text_lower)
    offering_score = sum(1 for phrase in SOLO_BIZ_INDICATORS if phrase in text_lower)

    if hiring_score >= 2:
        return False, "Detected Job Advertisement"

    if "handelsvertreter gesucht" in text_lower or "suche handelsvertreter" in text_lower:
        return False, "Detected 'Rep Wanted' Ad"

    if offering_score >= 1:
        return True, "Identified as Sales Rep/Solo Biz"

    if ("profil" in title_lower) or ("profile" in title_lower) or ("/in/" in text_lower):
        return True, "Likely a Profile"

    return False, "Content too generic / No clear intent"

def classify_lead(lead: Dict[str, Any], title: str = "", text: str = "") -> str:
    """Classify lead/persona type for lead_type column."""
    name = (lead.get("name") or "").strip()
    combined = " ".join([(title or ""), (text or ""), name]).lower()
    name_lower = name.lower()
    title_lower = (title or "").lower()
    text_lower = (text or "").lower()

    job_triggers = set(STRICT_JOB_AD_MARKERS + [
        "your profile",
        "wir bieten",
        "gesucht",
    ])

    # Hard job-ad signals
    if any(m in combined for m in job_triggers):
        return "job_ad"

    company_markers = (
        "gmbh", "ag", "kg", "ug", "co. kg", "inc", "ltd", "holding", "group",
        "unternehmen", "firma", "company",
    )
    if any(tok in name_lower for tok in company_markers) or any(tok in title_lower for tok in company_markers):
        return "company"
    if looks_like_company(name) or looks_like_company(title):
        return "company"

    sales_tokens = [
        "vertrieb", "verkauf", "sales", "handelsvertreter", "aussendienst",
        "account manager", "key account", "call center", "telefonverkauf",
        "sales representative", "commercial agent", "account executive",
    ]
    is_human = is_likely_human_name(name)
    has_sales = any(tok in combined for tok in sales_tokens)
    if is_human and has_sales:
        return "individual"
    if is_human:
        return "candidate"
    if has_sales:
        return "candidate"
    return "candidate"

def clean_email(email: str) -> str:
    if not email:
        return ""
    return email.replace("remove-this.", "").replace(".nospam", "")

def normalize_email(e: str) -> str:
    if not e:
        return ""
    e = clean_email(e.strip().lower())
    local, _, domain = e.partition("@")
    if domain == "gmail.com":
        local = local.split("+", 1)[0].replace(".", "")
    return f"{local}@{domain}"

def extract_company_name(title_text:str)->str:
    if not title_text: return ""
    m = re.split(r'[-–|•·:]', title_text)
    base = m[0].strip() if m else title_text.strip()
    if re.search(r'\b(Job|Karriere|Kontakt|Impressum)\b', base, re.I): return ""
    return base[:120]

def detect_company_size(text:str)->str:
    patterns = {
        "klein":  r'\b(1-10|klein|Inhaber|Familienunternehmen)\b',
        "mittel": r'\b(11-50|50-250|Mittelstand|[1-9]\d?\s*Mitarbeiter)\b',
        "groß":   r'\b(250\+|Konzern|Tochtergesellschaft|international|[2-9]\d{2,}\s*Mitarbeiter)\b'
    }
    for size, pat in patterns.items():
        if re.search(pat, text, re.I): return size
    return "unbekannt"

def detect_industry(text:str)->str:
    for k, pat in INDUSTRY_PATTERNS.items():
        if re.search(pat, text, re.I): return k
    return "unbekannt"

def detect_recency(html:str)->str:
    if re.search(r'\b(2025|2024)\-(0[1-9]|1[0-2])\-(0[1-9]|[12]\d|3[01])\b', html): return "aktuell"
    if re.search(r'\b(0?[1-9]|[12]\d|3[01])\.(0?[1-9]|1[0-2])\.(2024|2025)\b', html): return "aktuell"
    if re.search(r'\b(ab\s*sofort|sofort|zum\s*nächsten?\s*möglichen\s*termin)\b', html, re.I): return "sofort"
    return "unbekannt"

def estimate_hiring_volume(text:str)->str:
    if re.search(r'\b(mehrere|Teams|Team-?Erweiterung|Verstärkung|wir wachsen)\b', text, re.I): return "hoch"
    if len(re.findall(r'\b(Stelle|Stellen|Job)\b', text, re.I)) > 1: return "mittel"
    return "niedrig"

ADDRESS_RE = re.compile(
    r'\b([A-ZÄÖÜ][\wÄÖÜäöüß\-\.\s]{2,40}?(?:straße|str\.|weg|platz|allee|gasse|ring|ufer|chaussee|damm|steig|pfad|stieg|promenade)\s*\d+[a-zA-Z]?(?:\s*,\s*)?(?:\d{5}\s+[A-ZÄÖÜ][\wÄÖÜäöüß\-\.\s]{2,40})?)',
    re.I
)
ADDRESS_RE_ALT = re.compile(
    r'\b(\d{5}\s+[A-ZÄÖÜ][\wÄÖÜäöüß\-\.\s]{2,40}\s*,?\s*[A-ZÄÖÜ][\wÄÖÜäöüß\-\.\s]{2,40}?(?:straße|str\.|weg|platz|allee|gasse|ring|ufer|chaussee|damm|steig|pfad|stieg|promenade)\s*\d+[a-zA-Z]?)',
    re.I
)

SOCIAL_URL_RE = re.compile(
    r'(https?://(?:www\.)?(?:linkedin\.com/(?:in|company)/[^\s"\'<>]+|xing\.com/profile/[^\s"\'<>]+|facebook\.com/[^\s"\'<>]+|instagram\.com/[^\s"\'<>]+|x\.com/[^\s"\'<>]+|twitter\.com/[^\s"\'<>]+|tiktok\.com/@[^\s"\'<>]+))',
    re.I
)
SOCIAL_DOMAINS = ("linkedin.com", "xing.com", "x.com", "twitter.com", "facebook.com", "instagram.com", "tiktok.com")

def extract_private_address(text: str) -> str:
    """
    Greift die erste wahrscheinliche Anschrift aus dem Text heraus (Straße + Hausnummer, optional PLZ/Ort).
    """
    if not text:
        return ""
    snippet = text[:5000]
    for pat in (ADDRESS_RE, ADDRESS_RE_ALT):
        m = pat.search(snippet)
        if m:
            addr = re.sub(r"\s+", " ", m.group(1)).strip(" ,;")
            return addr
    return ""

def extract_social_profile_url(soup: Optional[BeautifulSoup], text: str) -> str:
    """
    Sucht nach Social-Links (LinkedIn/Xing/etc.) und liefert den ersten priorisierten Treffer.
    """
    candidates: list[str] = []

    def _add(url: str):
        if not url:
            return
        url = url.strip()
        if url.startswith("//"):
            url = "https:" + url
        low = url.lower()
        if not any(d in low for d in SOCIAL_DOMAINS):
            return
        clean = url.split("#", 1)[0].split("?", 1)[0]
        if clean not in candidates:
            candidates.append(clean)

    if soup:
        for a in soup.find_all("a", href=True):
            _add(a.get("href", ""))

    if not candidates and text:
        for m in SOCIAL_URL_RE.finditer(text):
            _add(m.group(1))

    if not candidates:
        return ""

    priority = ["linkedin.com", "xing.com", "x.com", "twitter.com", "facebook.com", "instagram.com", "tiktok.com"]
    candidates.sort(key=lambda u: min((priority.index(p) if p in u.lower() else len(priority) for p in priority)))
    return candidates[0]

def extract_locations(text:str)->str:
    cities = re.findall(r'\b(Düsseldorf|Köln|Essen|Dortmund|Mönchengladbach|Bochum|Wuppertal|Bonn|NRW|Nordrhein[-\s]?Westfalen)\b', text, re.I)
    out, seen=[], set()
    for c in cities:
        k=c.lower()
        if k in seen: continue
        seen.add(k); out.append(c)
    return ", ".join(out[:6]) if out else ""

def tags_from(text:str)->str:
    t=[]
    if JOBSEEKER_RE.search(text): t.append("jobseeker")
    elif RECRUITER_RE.search(text): t.append("recruiter")

    if CALLCENTER_HINT.search(text): t.append("callcenter")
    if D2D_HINT.search(text): t.append("d2d")
    if PROVISION_HINT.search(text): t.append("provision")
    if B2C_HINT.search(text): t.append("b2c")
    if CITY_RE.search(text): t.append("nrw")
    if WHATSAPP_RE.search(text) or WHATSAPP_INLINE.search(text) or WHATS_RE.search(text) \
       or WHATSAPP_PHRASE_RE.search(text) or WA_LINK_RE.search(text):
        t.append("whatsapp")
    if TELEGRAM_LINK_RE.search(text) or re.search(r'\btelegram\b', text, re.I):
        t.append("telegram")
    for k, pat in INDUSTRY_PATTERNS.items():
        if re.search(pat, text, re.I): t.append(k)
    return ",".join(dict.fromkeys(t))


def opening_line(lead:dict)->str:
    t = (lead.get("tags","") or "")
    if "d2d" in t: return "D2D in NRW: tägliche Touren + überdurchschnittliche Provision, Auszahlung wöchentlich."
    if "callcenter" in t: return "Outbound-Leads mit hoher Abschlussquote – Provision top, Auszahlung wöchentlich."
    return "Warme Leads, starker Provisionsplan, schnelle Auszahlung."

def normalize_phone(p: str) -> str:
    """
    DE-Telefon-Normalisierung (E.164-ähnlich) mit Edge-Cases wie '(0)'.
    Beispiele:
      '0211 123456'            -> '+49211123456'
      '+49 (0) 211 123456'     -> '+49211123456'
      '0049 (0) 211-123456'    -> '+49211123456'
      '+49-(0)-176 123 45 67'  -> '+491761234567'
    """
    if not p:
        return ""

    s = str(p).strip()

    # (0)-Edgecases vor der eigentlichen Bereinigung behandeln
    # Erfasst (), [], {} und beliebige Whitespaces: '(0)', '( 0 )', '[0]' etc.
    s = re.sub(r'[\(\[\{]\s*0\s*[\)\]\}]', '0', s)

    # Häufige Extension-/Zusatzangaben am Ende entfernen (ext, Durchwahl, DW, Tel.)
    s = re.sub(r'(?:durchwahl|dw|ext\.?|extension)\s*[:\-]?\s*\d+\s*$', '', s, flags=re.I)

    # Alle Zeichen außer Ziffern und Plus entfernen
    s = re.sub(r'[^\d+]', '', s)

    # Internationale Präfixe vereinheitlichen
    s = re.sub(r'^00', '+', s)        # 0049 -> +49, 0033 -> +33, etc.
    s = re.sub(r'^\+049', '+49', s)   # Tippfehler-Variante absichern
    s = re.sub(r'^0049', '+49', s)    # (redundant, aber explizit)

    # Optionales (0) hinter +49 entfernen (nach oben bereits zu '0' gemacht)
    # Beispiele: '+490211...' -> '+49211...'
    s = re.sub(r'^\+490', '+49', s)

    # Nationale führende 0 → +49
    if s.startswith('0') and not s.startswith('+'):
        s = '+49' + s[1:]

    # Mehrfaches '+' absichern
    if s.count('+') > 1:
        s = '+' + re.sub(r'\D', '', s)

    # Falls noch kein '+' vorhanden ist, hinzufügen (E.164-ähnlich)
    if not s.startswith('+'):
        s = '+' + re.sub(r'\D', '', s)

    # Plausibilitätscheck (Gesamtziffernzahl)
    digits = re.sub(r'\D', '', s)
    if len(digits) < 8 or len(digits) > 16:
        return s  # unbearbeitet zurückgeben, wenn außerhalb Range

    return s


# German mobile prefixes (015x, 016x, 017x)
MOBILE_PREFIXES_DE = {'150', '151', '152', '155', '156', '157', '159',
                      '160', '162', '163', '170', '171', '172', '173',
                      '174', '175', '176', '177', '178', '179'}


def validate_phone(phone: str) -> Tuple[bool, str]:
    """
    Robust phone validation with strict requirements.
    Returns (is_valid, phone_type) where phone_type is 'mobile', 'landline', or 'invalid'.
    
    Requirements:
    - Must be in E.164 format or DE format
    - Length: 10-15 digits after normalization
    - Support DE/intl formats: +49/0049/0...
    - Detect mobile prefixes: 015/016/017
    """
    if not phone or not isinstance(phone, str):
        return False, "invalid"
    
    normalized = normalize_phone(phone)
    if not normalized:
        return False, "invalid"
    
    # Extract digits only
    digits = re.sub(r'\D', '', normalized)
    
    # Length check: 10-15 digits
    if len(digits) < 10 or len(digits) > 15:
        return False, "invalid"
    
    # Check if it's a valid German number
    if normalized.startswith('+49'):
        # Extract area/mobile code (first 3 digits after country code)
        if len(digits) >= 5:
            prefix = digits[2:5]  # Skip '49' country code
            if prefix in MOBILE_PREFIXES_DE:
                return True, "mobile"
            # Landline numbers typically start with area codes
            elif prefix[0] in '23456789':  # Valid landline area code starts
                return True, "landline"
    
    # International numbers (non-DE)
    elif normalized.startswith('+') and len(digits) >= 10:
        return True, "international"
    
    return False, "invalid"


def guess_name_around(pos:int, text:str, window=120):
    seg = text[max(0,pos-window):pos+window]
    m = PERSON_PREFIX.search(seg) or NAME_RE.search(seg)
    if not m: return ""
    return m.group(0).replace("Hr.","Herr").replace("Fr.","Frau")

def validate_contact(contact: dict, page_url: str = "", page_text: str = "") -> bool:
    email = (contact.get("email") or "").strip()
    phone = (contact.get("telefon") or "").strip()
    text_lower = (page_text or "").lower()
    info_like_localparts = {"info", "kontakt", "service", "office", "news", "presse"}
    candidate_context = any(tok in text_lower for tok in CANDIDATE_POS_MARKERS)
    if email:
        low = email.lower()
        if not EMAIL_RE.search(email):
            return False
        local, _, domain = low.partition("@")
        if local in EMPLOYER_EMAIL_PREFIXES:
            return False
        if local in info_like_localparts and not candidate_context:
            return False
        if local in BAD_MAILBOXES and local not in info_like_localparts:
            return False
        if any(domain.endswith(p) for p in PORTAL_DOMAINS):
            return False
        if page_url and not same_org_domain(page_url, domain):
            if domain not in {"gmail.com", "outlook.com", "hotmail.com", "gmx.de", "web.de"}:
                return False
        if re.fullmatch(r'\d{8,}', (local or "")):
            return False
    if phone:
        phone = normalize_phone(phone)
        digits = re.sub(r'\D', '', phone)
        if len(digits) < 8:
            return False
        if re.search(r'(\d)\1{5,}', digits):
            return False
    if not email and not phone:
        return False
    return True

def regex_extract_contacts(text: str, src_url: str):
    # Text absichern
    text = text or ""

    # 2× De-Obfuscation (schluckt verschachtelte Varianten wie [at](punkt), ät, {dot}, etc.)
    for _ in range(2):
        text = deobfuscate_text_for_emails(text)

    # Gate: Sales/Jobseeker ODER offensichtliche Kontakt-/Messenger-URL zulassen
    is_contact_like = any(x in (src_url or "").lower() for x in ("/kontakt","/kontaktformular","/impressum","/team","/ansprechpartner"))
    messenger_hit = bool(
        WHATSAPP_RE.search(text) or WHATS_RE.search(text) or WA_LINK_RE.search(text) or
        WHATSAPP_PHRASE_RE.search(text) or TELEGRAM_LINK_RE.search(text) or re.search(r'\btelegram\b', text, re.I)
    )
    if not (SALES_WINDOW.search(text) or JOBSEEKER_WINDOW.search(text) or is_contact_like or messenger_hit):
        log("info", "Regex-Fallback: kein Sales/Jobseeker/Messenger-Kontext", url=src_url)
        return []


    # Näheprüfung mit erweitertem Fenster (±400 Zeichen)
    def _sales_near(a: int, b: int) -> bool:
        span = text[max(0, a - 400): min(len(text), b + 400)]
        return bool(SALES_WINDOW.search(span))

    # Treffer sammeln
    email_hits = [(m.group(0), m.start(), m.end()) for m in EMAIL_RE.finditer(text)]
    mobile_hits = [(m.group(0), m.start(), m.end(), True) for m in MOBILE_RE.finditer(text)]
    phone_hits_generic = [(m.group(0), m.start(), m.end(), False) for m in PHONE_RE.finditer(text)]
    phone_hits = []
    seen_spans = set()
    for hit in mobile_hits + phone_hits_generic:
        key = (hit[1], hit[2])
        if key in seen_spans:
            continue
        seen_spans.add(key)
        phone_hits.append(hit)
    wa_hits    = [(m.group(0), m.start(), m.end()) for m in WHATSAPP_RE.finditer(text)]
    wa_hits2   = [(m.group(0), m.start(), m.end()) for m in WHATS_RE.finditer(text)]

    rows = []
    if not email_hits and not phone_hits and not wa_hits and not wa_hits2:
        log("info", "Regex-Fallback: keine Treffer", url=src_url)
        return rows

    # E-Mails mit nächster Telefonnummer verbinden (falls vorhanden) + Name raten
    for e, es, ee in email_hits:
        if not _sales_near(es, ee):
            continue
        best_p, best_ppos, best_mobile = "", None, False
        best_dist = 10**9
        for p, ps, pe, is_mobile in phone_hits:
            if not _sales_near(ps, pe):
                continue
            if not _phone_context_ok(text, ps, pe):
                continue
            d = min(abs(ps - es), abs(pe - ee))
            if d < best_dist or (d == best_dist and is_mobile and not best_mobile):
                best_dist, best_ppos, best_p, best_mobile = d, ps, p, is_mobile
        name = guess_name_around(es, text) or (guess_name_around(best_ppos, text) if best_ppos is not None else "")
        rows.append({
            "name": name,
            "rolle": "",
            "email": normalize_email(e),
            "telefon": normalize_phone(best_p) if best_p else "",
            "quelle": src_url
        })

    # Telefonnummern ergänzen, die noch nicht genutzt wurden
    used_tel = set(r["telefon"] for r in rows if r.get("telefon"))
    for p, ps, pe, _is_mobile in phone_hits:
        if not _sales_near(ps, pe):
            continue
        if not _phone_context_ok(text, ps, pe):
            continue
        np = normalize_phone(p)
        if np and np in used_tel:
            continue
        rows.append({
            "name": guess_name_around(ps, text),
            "rolle": "",
            "email": "",
            "telefon": np,
            "quelle": src_url
        })

    # WhatsApp-Telefonnummern aus Text
    for (w, ws, we) in (wa_hits + wa_hits2):
        if not _sales_near(ws, we):
            continue
        if not _phone_context_ok(text, ws, we):
            continue
        tel_candidates = re.findall(r'\+?\d[\d \-()]{6,}', w)
        tel = normalize_phone(tel_candidates[0]) if tel_candidates else ""
        if tel:
            rows.append({
                "name": guess_name_around(ws, text),
                "rolle": "",
                "email": "",
                "telefon": tel,
                "quelle": src_url
            })

    # WhatsApp-Links (wa.me / api.whatsapp.com)
    for m in WA_LINK_RE.finditer(text):
        if not _phone_context_ok(text, m.start(), m.end()):
            continue
        tel = re.sub(r'\D', '', m.group(0))
        if tel:
            tel_fmt = "+" + tel if not tel.startswith("+") else tel
            if tel_fmt not in {r.get("telefon") for r in rows}:
                rows.append({
                    "name": "",
                    "rolle": "",
                    "email": "",
                    "telefon": tel_fmt,
                    "quelle": src_url
                })

    log("info", "Regex-Fallback genutzt", url=src_url,
        emails=len(email_hits), phones=len(phone_hits),
        whatsapp=len(wa_hits) + len(wa_hits2), rows=len(rows))
    return rows


# Kandidaten-Heuristiken
CANDIDATE_NEG_MARKERS = (
    "wir suchen", "team sucht", "bewirb dich", "bewirb-dich", "gmbh", " ag ", "aktuell suchen",
)
CANDIDATE_POS_MARKERS = (
    "suche job", "suche arbeit", "suche stelle", "suche neue herausforderung",
    "biete mich an", "zu sofort", "arbeitslos", "arbeitssuchend",
    "lebenslauf", "cv", "curriculum vitae", "profil", "qualifikation",
    "kenntnisse", "fuehrerschein", "fuehrerschein", "erfahrung im vertrieb",
    "stellengesuch", "bewerberprofil",
)
CANDIDATE_PHONE_CONTEXT = (
    "lebenslauf", "cv", "profil", "erfahrung", "qualifikation", "vita",
    "bewerbung", "stellengesuch", "ich suche", "ich bin", "open to work", "freelancer", "freiberuf",
)

# Social media profile URL patterns for candidates mode
SOCIAL_PROFILE_PATTERNS = (
    "linkedin.com/in/",
    "xing.com/profile/",
    "instagram.com/",
    "facebook.com/",
    "twitter.com/",
    "x.com/",
    "t.me/",
    "chat.whatsapp.com/",
)


def is_garbage_context(text: str, url: str = "", title: str = "", h1: str = "") -> Tuple[bool, str]:
    """Detect obvious non-candidate contexts (blogs, shops, company imprint, job ads)."""
    
    # NEU: Im Candidates-Modus - Social-Profile NIEMALS als Garbage markieren!
    if _is_candidates_mode():
        url_lower = (url or "").lower()
        if any(pattern in url_lower for pattern in SOCIAL_PROFILE_PATTERNS):
            return False, ""  # Social-Profile durchlassen!
    
    t = (text or "").lower()
    t_lower = t
    ttl = (title or "").lower()
    h1l = (h1 or "").lower()

    # FIRST: Check if this is a CANDIDATE seeking a job - NEVER mark candidates as garbage!
    if is_candidate_seeking_job(text, title, url):
        return False, ""  # Not garbage - this is a candidate!
    
    # NEW: Allow social media profiles in candidate mode
    url_lower = url.lower() if url else ""
    social_profile_urls = [
        "linkedin.com/in/", "xing.com/profile/", "facebook.com/profile/",
        "instagram.com/", "twitter.com/", "x.com/",
    ]
    if any(social_url in url_lower for social_url in social_profile_urls):
        return False, ""  # Social profiles are allowed in candidate mode

    news_tokens = (
        "news", "artikel", "bericht", "tipps", "trends", "black friday",
        "angebot", "webinar", "termine", "veranstaltung",
    )
    if any(tok in ttl for tok in news_tokens) or any(tok in h1l for tok in news_tokens):
        return True, "news_blog"

    shop_tokens = ("warenkorb", "kasse", "preis inkl", "preis inkl. mwst", "versandkosten", "lieferzeit", "bestellen")
    if any(tok in t for tok in shop_tokens):
        return True, "shop_product"

    company_tokens = (" gmbh", "gmbh", " ag", " kg")
    profile_tokens = ("profil", "lebenslauf", "gesuch")
    if any(tok in ttl for tok in company_tokens) and not any(pk in ttl for pk in profile_tokens):
        return True, "company_imprint"

    job_ad_tokens = ("wir suchen", "wir bieten", "deine aufgaben", "bewirb dich jetzt", "stellenanzeige", "jobangebot") + tuple(STRICT_JOB_AD_MARKERS)
    if any(tok in t for tok in job_ad_tokens):
        # Double-check: make sure it's not a candidate with these words in context
        if not is_candidate_seeking_job(text, title, url):
            return True, "job_ad"

    return False, ""

def _extract_name_from_text_block(txt: str) -> str:
    if not txt:
        return ""
    m = PERSON_PREFIX.search(txt) or NAME_RE.search(txt)
    if not m:
        return ""
    cand = (m.group(0) or "").strip()
    if cand.lower() in {"wir", "team"}:
        return ""
    return cand

def _ensure_candidate_name(record: Dict[str, Any], text: str, soup: Optional[BeautifulSoup], url: str, title_hint: str = "", h1_hint: str = "", *, linkedin_profile: bool = False) -> Dict[str, Any]:
    name = (record.get("name") or "").strip()
    email = (record.get("email") or "").strip()
    phone = (record.get("telefon") or "").strip()
    lower_url = (url or "").lower()
    is_linkedin_profile = linkedin_profile or ("linkedin.com/in/" in lower_url)

    def _clean_fragment(val: str) -> str:
        if not val:
            return ""
        cleaned = re.sub(r"\s*\|\s*(linkedin|xing).*?$", "", val, flags=re.I)
        cleaned = re.sub(r"\s*-\s*(linkedin|xing).*?$", "", cleaned, flags=re.I)
        return cleaned.strip()

    title_hint_clean = _clean_fragment(title_hint)
    h1_hint_clean = _clean_fragment(h1_hint)
    name = _clean_fragment(name)

    if name.lower() == "wir":
        name = ""

    # 0) LinkedIn-Slug -> Name
    slug_name = ""
    if is_linkedin_profile and "/in/" in lower_url:
        slug = lower_url.split("/in/", 1)[1]
        slug = slug.split("?", 1)[0].split("#", 1)[0].strip("/")
        parts = [p for p in re.split(r"[-_]+", slug) if p and not p.isdigit()]
        if len(parts) >= 2:
            slug_name = " ".join(p.capitalize() for p in parts[:3])
    if slug_name and (not name or not is_likely_human_name(name)):
        name = slug_name

    # 0b) Title/H1 cleanup-based guess
    if not name and title_hint_clean:
        maybe = _extract_name_from_text_block(title_hint_clean)
        if maybe:
            name = maybe
    if not name and h1_hint_clean:
        maybe = _extract_name_from_text_block(h1_hint_clean)
        if maybe:
            name = maybe

    # 1) E-Mail-Local-Part -> Name
    if (not name) and email and "@" in email:
        local = email.split("@", 1)[0]
        parts = re.split(r"[._\\-]+", local)
        if len(parts) >= 2:
            guess = " ".join(p.capitalize() for p in parts if p)
            if len(guess.split()) >= 2:
                name = guess

    # 2) Impressum-Pattern
    if (not name) and ("/impressum" in lower_url or "legal" in lower_url):
        m = re.search(r"angaben gemäß §\s*5\s*tmg[\s:]*\n?\r?\s*([A-ZÄÖÜ][^\n\r]{2,80})", text, flags=re.I)
        if m:
            name = _extract_name_from_text_block(m.group(1))
        if not name:
            m2 = re.search(r"Inhaltlich Verantwortlicher:?\s*([A-ZÄÖÜ][^\n\r]{2,80})", text, flags=re.I)
            if m2:
                name = _extract_name_from_text_block(m2.group(1))

    # 3) PDF/CV Header-Heuristik
    if not name:
        m = re.search(r"Lebenslauf von\s+([A-ZÄÖÜ][^\n\r]{2,80})", text, flags=re.I)
        if m:
            name = _extract_name_from_text_block(m.group(1))
    if not name:
        for line in text.splitlines():
            ln = line.strip()
            if not ln:
                continue
            maybe = _extract_name_from_text_block(ln)
            if maybe and len(maybe.split()) >= 2:
                name = maybe
                break

    # 4) Phone owner via Title/H1 (nur wenn Nummer vorhanden)
    if (not name) and phone:
        digits = re.sub(r"\\D", "", phone)
        for hint in (title_hint_clean, h1_hint_clean):
            if not hint:
                continue
            if digits and digits not in re.sub(r"\\D", "", hint):
                continue
            maybe = _extract_name_from_text_block(hint)
            if maybe:
                name = maybe
                break

    if (not name) and (phone or email):
        name = "Unknown Candidate"

    name = _clean_fragment(name)
    record["name"] = name
    return record

def is_candidate_profile_text(text: str, url: str = "") -> bool:
    """
    Check if text represents a candidate profile.
    In candidate mode, social media profiles are always accepted.
    """
    # NEW: Social media profiles are always accepted in candidate mode
    if url:
        url_lower = url.lower()
        social_profile_urls = [
            "linkedin.com/in/", "xing.com/profile/", "facebook.com/profile/",
            "instagram.com/", "twitter.com/", "x.com/",
        ]
        if any(social_url in url_lower for social_url in social_profile_urls):
            return True  # Social profiles are always accepted
def is_candidate_profile_text(text: str) -> bool:
    # Im Candidates-Modus: Alle Profile durchlassen (verhindert Text-basierte Job-Ad-Erkennung)
    if _is_candidates_mode():
        return True
    
    t = (text or "").lower()
    has_pos = any(tok in t for tok in CANDIDATE_POS_MARKERS)
    if not has_pos:
        return False
    if EMPLOYER_TEXT_RE.search(t) or any(tok in t for tok in CANDIDATE_NEG_MARKERS):
        return False
    return True

def _phone_context_ok(text: str, start: int, end: int) -> bool:
    window = text[max(0, start - 120): min(len(text), end + 120)].lower()
    return any(k in window for k in CANDIDATE_PHONE_CONTEXT)


def _anchor_contacts_fast(soup: BeautifulSoup, url: str) -> List[Dict[str, Any]]:
    """Liest nur <a href> auf Kontakt-/Impressum-/Team-Seiten: tel:, mailto:, wa.me / api.whatsapp.
    Liefert deduplizierte, validierte Kontakt-Records (ohne Volltext-Parsing)."""
    rows: List[Dict[str, Any]] = []
    if not soup:
        return rows

    for a in soup.find_all("a", href=True):
        h = a.get("href", "").strip()
        hl = h.lower()

        if hl.startswith("tel:"):
            tel = normalize_phone(h[4:])
            if tel:
                rows.append({"name": "", "rolle": "", "email": "", "telefon": tel, "quelle": url})

        elif hl.startswith("mailto:"):
            # alles nach mailto: bis evtl. '?' (Parameter) nehmen
            addr = h[7:].split("?", 1)[0].strip()
            if addr:
                rows.append({"name": "", "rolle": "", "email": addr, "telefon": "", "quelle": url})

        elif ("wa.me/" in hl) or ("api.whatsapp.com" in hl):
            digits = re.sub(r"\D", "", h)
            if digits:
                tel = "+" + digits if not digits.startswith("+") else digits
                tel = normalize_phone(tel)
                if tel:
                    rows.append({"name": "", "rolle": "", "email": "", "telefon": tel, "quelle": url})

    # Dedupe + Validierung (gleiche Logik wie sonst)
    out: List[Dict[str, Any]] = []
    seen: set[Tuple[str, str]] = set()
    for r in rows:
        key = ((r.get("email") or "").lower(), r.get("telefon") or "")
        if key in seen:
            continue
        if validate_contact(r, page_url=url):
            out.append(r)
            seen.add(key)
    return out

def _has_contact_anchor(soup: BeautifulSoup) -> bool:
    if not soup:
        return False
    for a in soup.find_all("a", href=True):
        h = a.get("href","").lower().strip()
        if h.startswith("mailto:") or h.startswith("tel:") or ("wa.me/" in h) or ("api.whatsapp.com" in h) \
           or ("chat.whatsapp.com" in h) or ("t.me/" in h) or ("telegram.me/" in h):
            return True
    return False


def extract_kleinanzeigen(html:str, url:str):
    lu = (url or "").lower()
    if ("kleinanzeigen.de" not in lu) and ("ebay-kleinanzeigen.de" not in lu):
        return []
    soup = BeautifulSoup(html, "html.parser")
    rows=[]
    wa = soup.select_one('a[href*="wa.me"], a[href*="api.whatsapp.com"]')
    tel = ""
    if wa:
        href = wa.get("href","")
        tel = re.sub(r'\D','', href)
        if tel: tel = "+"+tel
    if tel:
        rows.append({"name":"","rolle":"", "email":"", "telefon":tel, "quelle":url})
    return rows

def extract_kleinanzeigen_links(html: str, base_url: str = "") -> List[str]:
    """
    Extrahiert Anzeigen-Links aus einer Kleinanzeigen-Hub-Seite.
    """
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        return []
    links: List[str] = []
    blacklist = ("fahrer","kurier","lager","amazon","zusteller","lieferant","reinigung","pflege","stapler")
    for a in soup.find_all("a", href=True):
        href = (a.get("href") or "").strip()
        if "/s-anzeige/" not in href:
            continue
        text = (a.get_text(" ", strip=True) or "").lower()
        href_lower = href.lower()
        if any(word in href_lower or word in text for word in blacklist):
            continue
        full = urllib.parse.urljoin(base_url or "https://www.kleinanzeigen.de", href)
        links.append(full)
    return list(dict.fromkeys(links))


def is_commercial_agent(text: str) -> bool:
    """
    Detects legal/professional Handelsvertreter fingerprints in text.
    """
    if not text:
        return False
    low = text.lower()
    return any(fp in low for fp in AGENT_FINGERPRINTS)

# =========================
# Scoring
# =========================

def compute_score(text: str, url: str, html: str = "") -> int:
    # --- FIX START ---
    # Define t_lower immediately so it is available for all checks
    t = text or ""
    t_lower = t.lower()
    # Preserve previous lowercased behavior for downstream checks
    t = t_lower
    # --- FIX END ---
    u = (url or "").lower()
    title_text = ""
    if html:
        try:
            soup_title = BeautifulSoup(html, "html.parser")
            ttag = soup_title.find("title")
            if ttag:
                title_text = ttag.get_text(" ", strip=True)
        except Exception:
            title_text = ""
    title_lower = title_text.lower()
    # Zusatzflags für Off-Target-Quellen
    job_host_hints = (
        "ebay-kleinanzeigen.de",
        "kleinanzeigen.de",
        "/s-jobs/",
        "/stellenangebote",
        "/stellenangebot",
        "/job/",
        "/jobs/"
    )
    is_job_board = any(h in u for h in job_host_hints)

    public_hints = (
        "bundestag.de",
        "/rathaus/",
        "/verwaltung/",
        "/gleichstellungsstelle",
        "/grundsicherung",
        "/soziales-gesundheit-wohnen-und-recht",
        "/partei",
        "/landtag",
        "/ministerium"
    )
    is_public_context = any(h in u for h in public_hints)

    hr_role_hints = (
        "personalabteilung",
        "personalreferent",
        "personalreferentin",
        "sachbearbeiter personal",
        "sachbearbeiterin personal",
        "hr-manager",
        "hr manager",
        "human resources",
        "bewerbungen richten sie an",
        "bewerbung richten sie an",
        "pressesprecher",
        "pressesprecherin",
        "unternehmenskommunikation",
        "pressekontakt",
        "events-team",
        "veranstaltungen",
        "seminarprogramm"
    )
    is_hr_or_press = any(h in t for h in hr_role_hints)
    score = 0
    reasons: List[str] = []
    has_mobile = bool(MOBILE_RE.search(t))
    has_tel_number = bool(PHONE_RE.search(t))
    has_tel_word = ("tel:" in t) or ("telefon" in t) or ("tel." in t) or bool(re.search(r'\btelefon\b|\btel\.', t))
    has_tel = has_mobile or has_tel_number or has_tel_word
    has_wa_phrase = bool(WHATSAPP_PHRASE_RE.search(t))
    has_wa_word = ("whatsapp" in t) or has_wa_phrase
    has_wa_link = bool(WA_LINK_RE.search(html or "")) or bool(WA_LINK_RE.search(t))
    has_tg_link = bool(TELEGRAM_LINK_RE.search(html or "")) or bool(TELEGRAM_LINK_RE.search(t))
    has_telegram = has_tg_link or ("telegram" in t)
    has_whatsapp = has_wa_word or has_wa_link
    has_email = ("mailto:" in t) or ("e-mail" in t) or ("email" in t) or bool(re.search(r'\bmail\b', t))
    generic_mail_fragments = ("noreply@", "no-reply@", "donotreply@", "do-not-reply@", "info@", "kontakt@", "contact@", "office@", "support@", "service@")
    has_personal_email = has_email and not any(g in t for g in generic_mail_fragments)
    has_switch_now = any(k in t for k in [
        "quereinsteiger", "ab sofort", "sofort starten", "sofort start",
        "keine erfahrung noetig", "ohne erfahrung", "jetzt bewerben",
        "heute noch bewerben", "direkt bewerben"
    ])
    has_candidate_kw = any(k in t for k in CANDIDATE_KEYWORDS)
    has_ignore_kw = any(k in t for k in IGNORE_KEYWORDS)
    # Check for low pay / provision hints
    has_lowpay_or_prov = any(h in t_lower for h in LOW_PAY_HINT) or any(k in t_lower for k in [
        "nur provision", "provisionsbasis", "fixum + provision",
        "freelancer", "selbststaendig", "werkvertrag",
    ])
    agent_fingerprint = is_commercial_agent(t)
    if agent_fingerprint:
        score += 40
        reasons.append("agent_fingerprint")
    has_d2d = bool(D2D_HINT.search(t)) or any(k in t for k in ["door to door", "haustür", "haustuer", "kaltakquise"])
    has_callcenter = bool(CALLCENTER_HINT.search(t))
    has_b2c = bool(B2C_HINT.search(t))
    # Commission / high-value signals
    if any(h in t_lower for h in COMMISSION_HINT):
        score += 15
        reasons.append("commission_terms")
    industry_hits = sum(1 for k in INDUSTRY_HINTS if k in t)
    in_nrw = bool(CITY_RE.search(t)) or any(k in t for k in [" nrw ", " nordrhein-westfalen "])
    on_contact_like = any(h in u for h in ["kontakt", "impressum"])
    on_sales_path = any(h in u for h in ["callcenter", "telesales", "outbound", "vertrieb", "verkauf", "sales", "d2d", "door-to-door"])
    job_like = any(h in u for h in ["jobs.", "/jobs", "/karriere", "/stellen", "/bewerb"])
    portal_like = any(b in u for b in ["google.com", "indeed", "stepstone", "monster.", "xing.", "linkedin.", "glassdoor."])
    negative_pages = any(k in u for k in ["/datenschutz", "/privacy", "/agb", "/terms", "/bedingungen", "/newsletter", "/search", "/login", "/account", "/warenkorb", "/checkout", "/blog/", "/news/"])

    if "kleinanzeigen.de" in u:
        if "/s-stellengesuche/" in u:
            score += 30
        elif ("gesuch" in title_lower) or ("suche" in title_lower):
            score += 30
        elif ("gesuch" in t) or ("suche" in t):
            score += 30
        if "/s-stellengesuche/" in u and not has_ignore_kw:
            score = max(score, 50)

    if ("facebook.com" in u) and ("/groups" in u):
        score += 10  # leichte Bevorzugung für Facebook-Gruppen

    if has_whatsapp:
        score += 28
        if has_wa_link:
            score += 6
    if has_telegram:
        score += 8
        if has_tg_link:
            score += 4
    if has_mobile:
        score += 50
    elif has_tel:
        score += 14
    if has_personal_email:
        score += 12
    elif has_email:
        score += 6
    channel_count = int(has_whatsapp) + int(has_mobile or has_tel) + int(has_email) + int(has_telegram)
    if channel_count >= 2:
        score += 10   # angehoben
    if channel_count >= 3:
        score += 16   # deutlich angehoben
    if has_lowpay_or_prov:
        score -= 20
        reasons.append("low_pay_terms")
    if has_switch_now:
        score += 12
    if has_d2d:
        score += 9
    if has_callcenter:
        score += 7
    if has_b2c:
        score += 4
    if has_candidate_kw:
        score += 20
    if has_ignore_kw:
        score -= 100
    
    # NEW: Candidate-focused scoring boosts
    candidate_signals = [sig.lower() for sig in CANDIDATE_ALWAYS_ALLOW]
    candidate_hits = sum(1 for sig in candidate_signals if sig in t_lower)
    if candidate_hits > 0:
        score += min(candidate_hits * 15, 45)  # Up to +45 for strong candidate signals
        reasons.append(f"candidate_signals_{candidate_hits}")
    
    # Social media profile boost
    if any(social in u for social in ["linkedin.com/in/", "xing.com/profile/"]):
        score += 20
        reasons.append("social_profile")
    
    # Freelancer/Available signals
    if any(sig in t_lower for sig in ["freelancer", "verfügbar ab", "ab sofort verfügbar", "freiberuflich"]):
        score += 10
        reasons.append("availability_signal")
    
    if industry_hits:
        score += min(industry_hits * 4, 16)
    if in_nrw:
        score += 6
    if on_contact_like:
        score += 14   # angehoben
    if on_sales_path:
        score += 6
    if job_like:
        score -= 32
    if portal_like:
        score -= 24
    if negative_pages:
        score -= 10
    if any(k in t for k in [
        "per whatsapp bewerben", "bewerbung via whatsapp", "per telefon bewerben", "ruf uns an", "anrufen und starten",
        "meldet euch per whatsapp", "schreib mir per whatsapp", "meldet euch bei whatsapp"
    ]):
        score += 12
    if ("chat.whatsapp.com" in u) or ("t.me" in u):
        score += 100
    rec = detect_recency(html or "")
    if rec in ("aktuell", "sofort"):
        score += 8

    # Harte Abwertungen für Off-Target-Kontexte
    if is_job_board:
        score -= 40
    if is_public_context:
        score -= 40
    if is_hr_or_press:
        score -= 30

    return max(0, min(int(score), 100))

# =========================
# Content/Process (ASYNC)
# =========================

def validate_content(html: str, url: str) -> bool:
    if not html or not html.strip():
        return False
    raw = html.lstrip()
    if raw.startswith("<?xml") or "<urlset" in raw[:200].lower() or "<sitemapindex" in raw[:200].lower():
        log("debug", "XML/Sitemap erkannt (kein Lead-Content)", url=url)
        return False
    def _parse_html(doc: str, use_lxml_if_available: bool = True):
        soup_local = BeautifulSoup(doc, "html.parser")
        for el in soup_local.find_all(['script', 'style', 'noscript']):
            el.decompose()
        for c in soup_local.find_all(string=lambda t: isinstance(t, Comment)):
            c.extract()
        text_local = soup_local.get_text(" ", strip=True) if soup_local else ""
        text_local = text_local.replace("\xa0", " ").replace("\u200b", " ")
        text_local = re.sub(r"\s+", " ", text_local).strip()
        if use_lxml_if_available and len(text_local) < 120:
            try:
                soup_lx = BeautifulSoup(doc, "lxml")
                for el in soup_lx.find_all(['script', 'style', 'noscript']):
                    el.decompose()
                for c in soup_lx.find_all(string=lambda t: isinstance(t, Comment)):
                    c.extract()
                text_lx = soup_lx.get_text(" ", strip=True) if soup_lx else ""
                text_lx = text_lx.replace("\xa0", " ").replace("\u200b", " ")
                text_lx = re.sub(r"\s+", " ", text_lx).strip()
                if len(text_lx) > len(text_local) * 1.2:
                    return soup_lx, text_lx
            except Exception:
                pass
        return soup_local, text_local
    soup, text = _parse_html(raw)
    lower = text.lower()
    login_gate = (
        re.search(r'\b(anmelden|login|passwort)\b', lower, re.I) and
        re.search(r'\b(geschützt|nur\s*für\s*mitglieder|zugang\s*verweigert|restricted)\b', lower, re.I)
    )
    is_404 = any(phrase in lower for phrase in ["404", "seite nicht gefunden", "page not found", "not found"])
    german_core = re.search(r'\b(der|die|das|und|in|zu|für|mit|auf)\b', lower, re.I) is not None
    contact_tolerant = any(k in lower for k in ["kontakt", "impressum", "telefon", "e-mail", "email", "bewerben"])
    no_german = not (german_core or contact_tolerant)
    checks = {
        "too_short": len(text) < 200 and not contact_tolerant,
        "no_german": no_german,
        "is_404": is_404,
        "login_required": bool(login_gate),
    }
    if any(checks.values()):
        log("debug", "Content validation failed", url=url, checks=checks)
        return False
    return True


def is_high_quality_lead(row: Dict[str, Any], *, linkedin_profile: bool = False, agent_hit: bool = False) -> Tuple[bool, str]:
    """
    Dual-mode filter: handles candidates (CV/Stellengesuch) and agencies (Handelsvertretung/Industrievertretung).
    """
    name = (row.get("name") or "").lower()
    role_guess = (row.get("role_guess") or "").lower()
    phone_raw = (row.get("telefon") or "").strip()
    source = (row.get("quelle") or "").lower()

    legal_service_trash = [
        "rechtsanwalt", "kanzlei", "legal", "law", "anwalt",
        "umzug", "transporte", "pflegedienst", "autohaus", "kfz",
        "stiftung", "verein", "news", "presse",
        "messe frankfurt", "messe münchen"
    ]
    if any(w in name or w in role_guess for w in legal_service_trash):
        return False, "Blocked: Irrelevant Industry (Legal/B2C)"

    trash_words = ["stellenangebot", "jobs", "job", "ausbildung", "gmbh", "co. kg", "minijob", "büro", "fahrer", "lager", "suche arbeit", "team", "karriere"]
    if "stellengesuch" not in name and "jobgesuch" not in name:
        if any(w in name for w in trash_words):
            return False, "Trash keyword in name"

    is_candidate = any(k in name or k in role_guess for k in [
        "lebenslauf", "stellengesuch", "jobgesuch", "bewerbung", "suche neue", "ab sofort", "curriculum vitae"
    ])
    is_agency = agent_hit or any(k in name or k in role_guess for k in [
        "handelsvertretung", "industrievertretung", "vertriebsbüro", "vertriebsbuero", "agentur", "cdh", "hgb", "partner", "vertretung"
    ])

    text_for_niche = f"{row.get('name','')} {row.get('role_guess','')}".lower()
    if any(k in text_for_niche for k in ["shk", "sanitär", "heizung", "bauelemente", "werksvertretung"]):
        row["industry"] = "Construction/SHK"
    elif any(k in text_for_niche for k in ["medizinprodukte", "medizintechnik", "klinikreferent", "pharmareferent"]):
        row["industry"] = "Medical"
    elif any(k in text_for_niche for k in ["food", "lebensmittel", "getränke", "fmcg", "gastronomie"]):
        row["industry"] = "Food/FMCG"
    elif any(k in text_for_niche for k in ["freelancer", "freiberuflich", "selbstständig", "selbststaendig", "interim"]):
        row["industry"] = "Freelance Sales"
    elif any(k in text_for_niche for k in ["closer", "setter", "high ticket", "remote sales"]):
        row["industry"] = "Modern Sales/Closer"
    if ("aussteller" in text_for_niche) or ("messe" in text_for_niche):
        row["industry"] = "Trade Fair List"

    source_url = (row.get("url") or row.get("quelle") or "").lower()
    is_guerilla_source = any(d in source_url for d in ["reddit.com", "gutefrage.net", "facebook.com", "instagram.com", "kununu.com", "wiwi-treff.de"])
    has_sentiment_keywords = any(k in text_for_niche for k in ["gekündigt", "entlassen", "insolvenz", "suche arbeit", "job gesucht", "quereinstieg", "ohne ausbildung"])
    if is_guerilla_source and has_sentiment_keywords:
        row["industry"] = "Guerilla/Sentiment"
        return True, "Sentiment Signal found (Manual Review)"

    phone_normalized = normalize_phone(phone_raw) if phone_raw else ""
    has_mobile = phone_normalized.startswith(("+4915", "+4916", "+4917"))
    has_landline = phone_normalized.startswith("+49") and not phone_normalized.startswith("+490")

    if is_candidate:
        if has_mobile:
            return True, "Candidate with Mobile"
        return False, "Candidate without Mobile (Dropped)"

    if is_agency:
        if has_mobile or has_landline:
            return True, "Agency with Phone"
        return False, "Agency without Phone (Dropped)"

    if ("linkedin.com/in/" in source) or ("xing.com/profile/" in source):
        return True, "Social Profile (Review)"

    return False, "No valid category or phone number"

async def analyze_content_async(text: str, url: str) -> Dict[str, Any]:
    """
    LLM-basiertes Scoring des Seiteninhalts. Liefert Score/Category/Summary.
    """
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return {"score": 100, "category": "Unchecked", "summary": "No AI Key"}

    clean_text = (text or "")[:2000].replace("\n", " ")
    endpoint = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    system_prompt = (
        "You are an expert Headhunter AI. Your goal is to find CANDIDATES, not companies.\n"
        "Analyze the text and determine if it contains contact details of a human that can be recruited.\n\n"
        "1. CLASSIFY the content:\n"
        '   - \"POACHING\": A specific employee listed on a company team page (e.g., \"Sales Manager\" at Company X).\n'
        '   - \"FREELANCER\": A freelancer/contractor website or profile offering services.\n'
        '   - \"CV_DISCOVERY\": A CV, Resume, or \"Lebenslauf\" document (often PDF) or an \"Open to Work\" post.\n'
        '   - \"IRRELEVANT\": Job boards, general company homepages, news, shops, or agencies trying to sell recruiting services.\n\n'
        "2. DECIDE:\n"
        "   - Set \"is_relevant\": true ONLY if it is POACHING, FREELANCER, or CV_DISCOVERY.\n"
        "   - Set \"is_relevant\": false for IRRELEVANT.\n\n"
        'Return JSON: {"is_relevant": bool, "lead_type": string, "score": int, "reason": string}'
    )
    payload = {
        "model": os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"URL: {url}\nTEXT: {clean_text}"}
        ]
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(endpoint, headers=headers, json=payload, timeout=HTTP_TIMEOUT) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    choices = data.get("choices") or []
                    content = ""
                    if choices and isinstance(choices, list):
                        content = ((choices[0] or {}).get("message") or {}).get("content", "")  # type: ignore[index]
                    if content:
                        try:
                            parsed = json.loads(content)
                            score_raw = parsed.get("score", 100)
                            try:
                                score_val = int(score_raw)
                            except (TypeError, ValueError):
                                score_val = 100
                            score_val = max(0, min(100, score_val))
                            lead_type_val = (parsed.get("lead_type") or parsed.get("category") or "").strip() or "N/A"
                            reason_val = (parsed.get("reason") or parsed.get("summary") or "").strip() or "No reason"
                            is_rel = bool(parsed.get("is_relevant", True))
                            return {
                                "score": score_val,
                                "lead_type": lead_type_val,
                                "reason": reason_val,
                                "is_relevant": is_rel,
                            }
                        except Exception as e:
                            log("warn", "AI analysis JSON parse failed", url=url, error=str(e))
                    else:
                        log("warn", "AI analysis empty response", url=url, status=resp.status)
                else:
                    log("warn", "AI analysis HTTP error", url=url, status=resp.status)
    except Exception as e:
        log("warn", "AI Analysis failed", url=url, error=str(e))

    return {"score": 50, "category": "Error", "summary": "Analysis failed"}

async def extract_contacts_with_ai(text_content: str, url: str) -> List[Dict[str, Any]]:
    """
    Extrahiert Kontakte (Name/Rolle/Email/Telefon) per LLM. Gibt leere Liste bei Fehlern/Fallback.
    """
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return []

    clean_text = (text_content or "")[:3000].replace("\n", " ")
    endpoint = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    # In candidates mode, use enhanced prompt for likely candidate URLs
    if _is_candidates_mode() and is_candidate_url(url) is True:
        # Enhanced prompt for candidate profiles
        system_prompt = (
            'Analyze this profile of a person SEEKING a job (Stellengesuch - NOT a job offer). '
            'IMPORTANT: '
            '- This is a JOB SEEKER profile - the person is LOOKING FOR work '
            '- Extract contact data of the PERSON (not a company) '
            '- Mobile phone number is REQUIRED (015x, 016x, 017x, 01xx format) '
            '- If this is a COMPANY (not a person), set is_job_seeker to false '
            '- If no mobile phone found, still extract but note it '
            'Return JSON: {"is_job_seeker": true/false, "contacts": [{"name": "...", "role": "...", "email": "...", "phone": "...", "location": "...", "availability": "..."}]}. '
            "If no specific job seeker found, return empty contacts list."
        )
    else:
        # Standard prompt for agencies/companies
        system_prompt = (
            'Extract contact persons. Return JSON: {"contacts": [{"name": "...", "role": "...", "email": "...", "phone": "..."}]}. '
            "If no specific person found, return empty list."
        )
    payload = {
        "model": os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"URL: {url}\nTEXT: {clean_text}"}
        ]
    }

    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(endpoint, headers=headers, json=payload, timeout=HTTP_TIMEOUT) as resp:
                if resp.status != 200:
                    log("warn", "AI contact extraction HTTP error", url=url, status=resp.status)
                    return []
                data = await resp.json()
                choices = data.get("choices") or []
                content = ""
                if choices and isinstance(choices, list):
                    content = ((choices[0] or {}).get("message") or {}).get("content", "")  # type: ignore[index]
                if not content:
                    return []
                try:
                    parsed = json.loads(content)
                except Exception as e:
                    log("warn", "AI contact extraction parse failed", url=url, error=str(e))
                    return []
                # Check if this is a candidate response (only present in candidate mode)
                # Note: is_job_seeker is None when using standard prompt (not candidate mode)
                # We only skip if it's explicitly False (AI determined it's a company)
                is_job_seeker = parsed.get("is_job_seeker") if isinstance(parsed, dict) else None
                if is_job_seeker is False:
                    # AI explicitly determined this is NOT a job seeker (e.g., company page)
                    log("debug", "AI: Not a job seeker profile", url=url)
                    return []
                
                contacts_raw = parsed.get("contacts") if isinstance(parsed, dict) else None
                if not isinstance(contacts_raw, list):
                    return []
                cleaned: List[Dict[str, Any]] = []
                for c in contacts_raw:
                    if not isinstance(c, dict):
                        continue
                    name = (c.get("name") or "").strip()
                    role = (c.get("role") or "").strip()
                    email = (c.get("email") or "").strip()
                    phone = normalize_phone(c.get("phone") or "")
                    location = (c.get("location") or "").strip()
                    availability = (c.get("availability") or "").strip()
                    
                    if not (email or phone):
                        continue
                    
                    contact_record = {
                        "name": name,
                        "rolle": role,
                        "email": email,
                        "telefon": phone,
                        "quelle": url
                    }
                    
                    # Add candidate-specific fields if present
                    if location:
                        contact_record["location"] = location
                    if availability:
                        contact_record["availability"] = availability
                    
                    cleaned.append(contact_record)
                return cleaned
    except Exception as e:
        log("warn", "AI contact extraction failed", url=url, error=str(e))
        return []
    return []

async def process_link_async(url: UrlLike, run_id: int, *, force: bool = False) -> Tuple[int, List[Dict[str, Any]]]:
    is_pdf = False  # Track PDF status early
    meta = url if isinstance(url, dict) else {}
    snippet_hint = (meta.get("snippet", "") or "")
    search_title_hint = (meta.get("title", "") or "")
    url = _extract_url(url)
    if not url:
        return (0, [])
    extra_followups: List[str] = []
    extra_checked = 0
    parsed = urllib.parse.urlparse(url)
    host = parsed.netloc.lower()
    path_lower = (parsed.path or "").lower()
    is_linkedin = host.endswith("linkedin.com")
    linkedin_profile = is_linkedin and path_lower.startswith("/in/")
    linkedin_snippet_text = " ".join([search_title_hint, snippet_hint]).strip()

    def _random_desktop_ua() -> str:
        pool = [ua for ua in UA_POOL if "mobile" not in ua.lower() and "android" not in ua.lower()]
        return random.choice(pool or UA_POOL or [USER_AGENT])

    # History / Vorab-Checks
    if (not force) and url_seen(url):
        log("debug", "URL bereits gesehen (skip)", url=url)
        return (0, [])
    if (is_denied(url) or not path_ok(url)) and not linkedin_profile:
        log("debug", "Vorprüfung blockt URL", url=url)
        mark_url_seen(url, run_id)
        return (1, [])
    if not await robots_allowed_async(url):
        log("warn", "robots.txt verbietet Zugriff", url=url)
        mark_url_seen(url, run_id)
        return (1, [])

    # HTTP holen
    resp: Optional[Any] = None
    html = ""
    ct = ""
    using_linkedin_snippet = False
    content_bytes: bytes = b""
    li_status = None

    if linkedin_profile:
        li_status = -1
        login_wall = False
        ua_choice = _random_desktop_ua()
        try:
            async with _make_client(True, ua_choice, proxy_url=None, force_http1=False, timeout_s=HTTP_TIMEOUT) as client:
                resp = await client.get(url, allow_redirects=True, timeout=HTTP_TIMEOUT)
            li_status = getattr(resp, "status_code", 0) if resp else -1
            _LAST_STATUS[url] = li_status
            final_path = urllib.parse.urlparse(str(getattr(resp, "url", url))).path.lower() if resp else ""
            login_wall = (li_status == 999) or ("/login" in final_path)
            if login_wall and linkedin_snippet_text:
                log("info", "LinkedIn Login-Wall – nutze Snippet", url=url, status=li_status)
                using_linkedin_snippet = True
                html = linkedin_snippet_text
                resp = None
            elif resp and li_status == 200:
                setattr(resp, "insecure_ssl", False)
            else:
                resp = None
        except Exception as e:
            log("warn", "LinkedIn Fetch fehlgeschlagen", url=url, error=str(e))
            resp = None
            li_status = -1
            _LAST_STATUS[url] = li_status
    else:
        resp = await fetch_response_async(url)

    if (resp is None) and (not using_linkedin_snippet):
        st = _LAST_STATUS.get(url, li_status if li_status is not None else -1)
        if st in (429, 403, -1):
            log("warn", "Kein Content – Retry später (nicht markiert)", url=url, status=st)
            return (1, [])
        mark_url_seen(url, run_id)
        return (1, [])

    if resp is not None:
        ct = (resp.headers.get("Content-Type", "") or "").lower()
        if "application/pdf" in ct or url.lower().endswith(".pdf"):
            is_pdf = True  # PDF erkannt
        if is_pdf and not CFG.allow_pdf:
            log("info", "PDF übersprungen (ALLOW_PDF=0)", url=url)
            mark_url_seen(url, run_id)
            return (1, [])
        if is_pdf:
            try:
                content_bytes = getattr(resp, "content", b"") or b""
                if not content_bytes and hasattr(resp, "read"):
                    try:
                        content_bytes = await resp.read()
                    except Exception:
                        content_bytes = b""
                f = io.BytesIO(content_bytes)
                reader = PdfReader(f)
                text_content = ""
                for page in reader.pages[:5]:
                    try:
                        text_content += (page.extract_text() or "") + "\n"
                    except Exception:
                        continue
                html = text_content
            except Exception as e:
                log("warn", "PDF parsing failed", url=url, error=str(e))
                html = ""
        else:
            try:
                html = resp.text
            except Exception as e:
                log("error", "Response fehlerhaft", url=url, error=str(e))
                mark_url_seen(url, run_id)
                return (1, [])

    lu = url.lower()
    is_kleinanzeigen = ("kleinanzeigen.de" in lu) or ("ebay-kleinanzeigen.de" in lu)
    is_list_page = ("/k0" in lu or "/s-" in lu) and ("/s-anzeige/" not in lu)
    if is_kleinanzeigen and is_list_page:
        ka_links = extract_kleinanzeigen_links(html, url)
        collected: List[Dict[str, Any]] = []
        if ka_links:
            for link in ka_links:
                if link not in extra_followups:
                    extra_followups.append(link)
        if extra_followups:
            log("debug", "Kleinanzeigen-Hub erkannt, folge Anzeigenlinks", url=url, count=len(extra_followups))
            for fu in extra_followups:
                try:
                    inc2, items2 = await process_link_async(fu, run_id, force=force)
                    extra_checked += inc2
                    if items2:
                        collected.extend(items2)
                except Exception:
                    pass
        mark_url_seen(url, run_id)
        return (1 + extra_checked, collected if extra_followups else [])

    # Titel-basierter Guard (früher Exit, bevor teure Extraktion)
    title_text = ""
    try:
        soup_title = BeautifulSoup(html, "html.parser")
        ttag = soup_title.find("title")
        if ttag:
            title_text = ttag.get_text(" ", strip=True)
    except Exception:
        title_text = ""
    if (using_linkedin_snippet or is_pdf) and not title_text:
        title_text = search_title_hint or snippet_hint

    pdf_cv_hint = False
    if is_pdf:
        pdf_window = (html or "")[:1000].lower()
        hint_blob = " ".join([
            (url or "").lower(),
            (title_text or "").lower(),
            (search_title_hint or "").lower(),
        ])
        pdf_cv_hint = any(k in pdf_window for k in ("lebenslauf", "cv", "curriculum vitae")) or any(
            k in hint_blob for k in ("lebenslauf", "cv", "curriculum vitae")
        )
        if (url.lower().endswith(".pdf") or not CFG.allow_pdf) and not pdf_cv_hint:
            log("debug", "PDF ohne CV/Lebenslauf skip", url=url)
            mark_url_seen(url, run_id)
            return (1, [])

    title_src = (title_text or linkedin_snippet_text or url or "").lower()
    DIRECTORY_KEYWORDS = (
        "portal", "verzeichnis", "netzwerk", "verband", "marktplatz",
        "forum", "community", "treffpunkt", "liste", "firmen",
        "datenbank", "mitglieder", "aussteller", "katalog"
    )
    pos_keys = ("vertrieb","sales","verkauf","account","aussendienst","außendienst","kundenberater",
                "handelsvertreter","handelsvertretung","makler","akquise","agent","berater","beraterin","geschäftsführer",
                "repräsentant","b2b","b2c","verkäufer","verkaeufer","vertriebler",
                "vertriebspartner","aushilfe verkauf",
                "stellengesuch")
    neg_keys = ("reinigung","putz","hilfe","helfer","lager","fahrer","zusteller","kommissionierer",
                "melker","tischler","handwerker","bauhelfer","produktionshelfer","stapler",
                "pflege","medizin","arzt","kassierer","kasse","verräumer","regal",
                "aushilfe","minijob","winterdienst","promoter","promotion","fundraiser","spendensammler",
                "museum","theater","verein","crypto","bitcoin","nft","casino","dating","sex","flohmarkt",
                "impressum","gmbh","ag","kg","hrb","ust-id","datenschutzerklärung","agb")
    has_pos_key = any(k in title_src for k in pos_keys) or pdf_cv_hint
    intent_hit = any(k in title_src for k in INTENT_TITLE_KEYWORDS)
    suche_biete_hit = (("suche" in title_src) or ("biete" in title_src)) and (has_pos_key or intent_hit)
    job_ad_hit = any(k in title_src for k in STRICT_JOB_AD_MARKERS)
    directory_hit = any(k in title_src for k in DIRECTORY_KEYWORDS)
    use_positive_guard = not bool(OPENAI_API_KEY)
    
    # Check if this is a CANDIDATE seeking a job - NEVER skip candidates!
    is_candidate = is_candidate_seeking_job(title_text or "", "", url)
    
    # NEU: Im Candidates-Modus - Titel-Guard für Social-Profile überspringen
    url_lower = url.lower()
    is_social_profile = any(p in url_lower for p in SOCIAL_PROFILE_PATTERNS)
    
    if _is_candidates_mode():
        # Social-Profile überspringen den Titel-Guard
        if not is_social_profile:
            if any(k in title_src for k in neg_keys) and ("handelsvertretung" not in title_src and "handelsvertreter" not in title_src):
                if not is_candidate:
                    log("debug", "Titel-Guard: Negative erkannt, skip", url=url, title=title_text)
                    mark_url_seen(url, run_id)
                    return (1, [])
    else:
        # Standard-Logik für nicht-Candidates-Modus
        if any(k in title_src for k in neg_keys) and ("handelsvertretung" not in title_src and "handelsvertreter" not in title_src):
            if not is_candidate:
                log("debug", "Titel-Guard: Negative erkannt, skip", url=url, title=title_text)
                mark_url_seen(url, run_id)
                return (1, [])
    if job_ad_hit:
        # Don't skip if it's a candidate seeking a job (e.g., "ich suche job (m/w/d)" would have both)
        if not is_candidate:
            log("debug", "Titel-Guard: Job-Ad Marker erkannt, skip", url=url, title=title_text)
            mark_url_seen(url, run_id)
            return (1, [])
    if use_positive_guard and not (has_pos_key or intent_hit or suche_biete_hit or directory_hit):
        log("debug", "Titel-Guard: Keine positiven Keywords, skip", url=url, title=title_text)
        mark_url_seen(url, run_id)
        return (1, [])
    if use_positive_guard and directory_hit and not (has_pos_key or intent_hit or suche_biete_hit):
        log("debug", "Titel-Guard: Directory-Keyword gefunden (Pass)", url=url, title=title_text)

    ssl_insecure = getattr(resp, "insecure_ssl", False)
    invite_link = ("chat.whatsapp.com" in url.lower()) or ("t.me" in url.lower())

    # XML/Sitemap ausfiltern (Invite-Links dürfen durch)
    if (("xml" in ct) or html.lstrip().startswith("<?xml") or "<urlset" in html[:200].lower() or "<sitemapindex" in html[:200].lower()) and not invite_link:
        log("debug", "XML/Sitemap erkannt (kein Lead-Content)", url=url)
        mark_url_seen(url, run_id)
        return (1, [])

    # Content validieren (Invite-Links überspringen strenge Prüfung)
    if (not invite_link) and (not using_linkedin_snippet) and (not is_pdf) and (not validate_content(html, url)):
        mark_url_seen(url, run_id)
        return (1, [])

    if invite_link:
        soup_inv = BeautifulSoup(html, "html.parser")
        title_tag = soup_inv.find("title")
        og_title = soup_inv.find("meta", attrs={"property": "og:title"})
        group_title = ""
        if og_title and og_title.get("content"):
            group_title = og_title.get("content").strip()
        if (not group_title) and title_tag:
            group_title = title_tag.get_text(strip=True)

        record = {
            "name": group_title or "",
            "rolle": "Messenger Gruppe",
            "email": "",
            "telefon": "",
            "quelle": url,
            "score": 100,
            "tags": "group_invite",
            "region": "",
            "role_guess": "Messenger",
            "salary_hint": "",
            "commission_hint": "",
            "opening_line": "",
            "ssl_insecure": "yes" if ssl_insecure else "no",
            "company_name": "",
            "company_size": "",
            "hiring_volume": "",
            "industry": "",
            "recency_indicator": "",
            "location_specific": "",
            "confidence_score": 100,
            "last_updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "data_quality": 50,
            "phone_type": "",
            "whatsapp_link": "yes" if "chat.whatsapp.com" in url.lower() else "",
            "private_address": "",
            "social_profile_url": "",
            "lead_type": "group_invite",
            "ai_category": "Group Invite",
            "ai_summary": "Messenger group invite link",
        }
        ok, reason = is_high_quality_lead(record, linkedin_profile=linkedin_profile, agent_hit=agent_fingerprint_hit)
        if not ok:
            log("debug", "Dropped: Quality filter", reason=reason, url=url)
            mark_url_seen(url, run_id)
            return (1, [])
        mark_url_seen(url, run_id)
        return (1, [record])

    # Robust parsen
    def _parse_html(doc: str) -> Tuple[BeautifulSoup, str]:
        soup_local = BeautifulSoup(doc, "html.parser")
        for el in soup_local.find_all(['script', 'style', 'noscript']):
            el.decompose()
        for c in soup_local.find_all(string=lambda t: isinstance(t, Comment)):
            c.extract()
        text_local = soup_local.get_text(" ", strip=True) if soup_local else ""
        text_local = text_local.replace("\xa0", " ").replace("\u200b", " ")
        text_local = re.sub(r"\s+", " ", text_local).strip()
        if len(text_local) < 120:
            try:
                soup_lx = BeautifulSoup(doc, "lxml")
                for el in soup_lx.find_all(['script', 'style', 'noscript']):
                    el.decompose()
                for c in soup_lx.find_all(string=lambda t: isinstance(t, Comment)):
                    c.extract()
                text_lx = soup_lx.get_text(" ", strip=True) if soup_lx else ""
                text_lx = text_lx.replace("\xa0", " ").replace("\u200b", " ")
                text_lx = re.sub(r"\s+", " ", text_lx).strip()
                if len(text_lx) > len(text_local) * 1.2:
                    return soup_lx, text_lx
            except Exception:
                pass
        return soup_local, text_local

    soup, text = _parse_html(html)
    if soup is None:
        soup = BeautifulSoup(html, "html.parser")

    page_title = ""
    h1_hint = ""
    if soup:
        try:
            page_title = soup.title.get_text(" ", strip=True) if soup.title else ""
        except Exception:
            page_title = ""
        try:
            h1 = soup.find("h1")
            h1_hint = h1.get_text(" ", strip=True) if h1 else ""
        except Exception:
            h1_hint = ""

    if not invite_link:
        garbage, reason = is_garbage_context(text, url, page_title, h1_hint)
        if garbage:
            log("debug", "Garbage Context detected", url=url, reason=reason)
            mark_url_seen(url, run_id)
            return (1, [])
        
        # NEW: Analyze "wir suchen" context before blocking
        wir_suchen_classification, wir_suchen_reason = analyze_wir_suchen_context(text, url)
        if wir_suchen_classification == "job_ad":
            log("debug", "Job-Ad detected via wir suchen analysis", url=url, reason=wir_suchen_reason)
            mark_url_seen(url, run_id)
            return (1, [])
        elif wir_suchen_classification == "candidate":
            log("debug", "Candidate detected via wir suchen analysis - allowing", url=url, reason=wir_suchen_reason)
            # Continue processing - this is a candidate
        
        if not is_candidate_profile_text(text, url):
            log("debug", "Kein Kandidatenprofil - skip", url=url)
            mark_url_seen(url, run_id)
            return (1, [])
        if is_job_advertisement(text, page_title or title_text, snippet_hint):
            log("debug", "Job-Ad detected (hard block)", url=url)
            mark_url_seen(url, run_id)
            return (1, [])

    text_lower_global = (text or "").lower()
    ai_result = await analyze_content_async(text, url)
    ai_score_raw = (ai_result.get("score") if isinstance(ai_result, dict) else 100)
    try:
        ai_score = int(ai_score_raw)
    except (TypeError, ValueError):
        ai_score = 100
    ai_score = max(0, min(100, ai_score))
    ai_lead_type = (ai_result.get("lead_type") or "").strip().upper() if isinstance(ai_result, dict) else ""
    ai_category = ai_lead_type or ""
    ai_summary = (ai_result.get("reason") or "").strip() if isinstance(ai_result, dict) else ""
    ai_is_rel = bool(ai_result.get("is_relevant", True)) if isinstance(ai_result, dict) else True
    if not ai_is_rel:
        log("info", "[AI-FILTER] Irrelevant (recruiting)", url=url, score=ai_score, lead_type=ai_lead_type, reason=ai_summary)
        mark_url_seen(url, run_id)
        return (1 + extra_checked, [])
    if ai_score < 40:
        log("info", "Low Quality Lead", url=url, score=ai_score, lead_type=ai_lead_type, reason=ai_summary)
        mark_url_seen(url, run_id)
        return (1 + extra_checked, [])

    agent_fingerprint_hit = is_commercial_agent(text)
    private_address = extract_private_address(text)
    social_profile_url = extract_social_profile_url(soup, text)
    if linkedin_profile and not social_profile_url:
        social_profile_url = url
    role_category = classify_role(text, page_title or title_text)
    
    # NEW: Detect hidden gems (Quereinsteiger with sales potential)
    hidden_gem_type, hidden_gem_confidence, hidden_gem_reason = detect_hidden_gem(text, url)
    if hidden_gem_type:
        log("info", "Hidden gem detected", gem_type=hidden_gem_type, confidence=hidden_gem_confidence, 
            reason=hidden_gem_reason, url=url)

    # --- Job/Karriere/Jobs/stellen: nur weiter, wenn direkter Kontaktanker existiert ---
    lu = url.lower()
    if any(p in lu for p in ("/jobs", "/karriere", "/stellen")):
        if not _has_contact_anchor(soup):
            log("debug", "Jobpfad ohne Kontaktanker – skip", url=url)
            mark_url_seen(url, run_id)
            return (1, [])

    # >>> FAST-PATH: Kontakt/Impressum/Team/Ansprechpartner – Anker-Scan & Early-Return
    if any(key in lu for key in ("/kontakt", "/impressum", "/team", "/ansprechpartner")):
        fast_items = _anchor_contacts_fast(soup, url)
        if fast_items:
            # Erst normalen Score berechnen
            base_score = compute_score(text, url)
            # Wenn der Score unter der Mindest-Schwelle liegt, Fast-Path komplett verwerfen
            if base_score < CFG.min_score:
                log("debug", "Fast-Path Kontaktseite, aber Score unter Mindestwert", url=url, score=base_score)
                mark_url_seen(url, run_id)
                return (1, [])
            # Nur noch moderater Kontakt-Bonus, kein hartes Hochziehen mehr
            base_score = base_score + 10

            # Enrichment
            title_tag = soup.find('title') if soup else None
            comp_name = extract_company_name(title_tag.get_text() if title_tag else "")
            company_size = detect_company_size(text)
            company_domain = resolve_company_domain(comp_name) if comp_name else ""
            industry = detect_industry(text)
            recency = detect_recency(html)
            hiring_volume = estimate_hiring_volume(text)
            locations = extract_locations(text)
            base_tags = tags_from(text)
            role_guess = ("Jobseeker" if JOBSEEKER_RE.search(text) else
                          ("Callcenter" if CALLCENTER_HINT.search(text) else
                           ("D2D" if D2D_HINT.search(text) else "Vertrieb")))
            salary_hint = "low" if any(h in (text or "").lower() for h in LOW_PAY_HINT) else ""
            commission_hint = "yes" if PROVISION_HINT.search(text) else ("maybe" if SALES_RE.search(text) else "")
            region = "NRW" if CITY_RE.search(text) else ""
            last_updated = datetime.now(timezone.utc).isoformat(timespec="seconds")

            def _confidence(r: Dict[str, Any]) -> int:
                c = 0
                if r.get("telefon"): c += 50
                if r.get("email"): c += 30
                if "whatsapp" in (r.get("tags","") or "") or WA_LINK_RE.search(html): c += 10
                if comp_name: c += 10
                return min(c, 100)

            def _quality(r: Dict[str, Any]) -> int:
                q = 0
                if validate_contact(r, page_url=url, page_text=text): q += 50
                if hiring_volume != "niedrig": q += 10
                if recency != "unbekannt": q += 20
                if not ssl_insecure: q += 10
                return min(q, 100)

            def _mail_boost_and_label(email: str):
                q = email_quality(email or "", url)
                if q == "reject": return (None, q)
                return ({"personal":8, "team":4, "generic":2, "weak":0}.get(q,0), q)

            def _tel_wa_boost(record: Dict[str, Any], page_text: str, page_html: str):
                t = (page_text or "").lower()
                h = (page_html or "").lower()
                tel = (record.get("telefon") or "").strip()
                boost = 0; extras={}
                has_wa_word = ("whatsapp" in t) or bool(WHATSAPP_PHRASE_RE.search(t))
                has_wa_link = bool(WA_LINK_RE.search(h)) or bool(WA_LINK_RE.search(t))
                if has_wa_word or has_wa_link:
                    boost += 12
                    extras["tags_add_whatsapp"] = True
                    extras["whatsapp_link"] = bool(has_wa_link)
                has_tg_link = bool(TELEGRAM_LINK_RE.search(h)) or bool(TELEGRAM_LINK_RE.search(t))
                if has_tg_link:
                    boost += 6
                    extras["tags_add_telegram"] = True
                is_mobile = bool(re.search(r'(?<!\d)(?:\+49|0049|0)\s*1[5-7]\d(?:[\s\/\-]?\d){6,}(?!\d)', tel)) if tel else False
                if is_mobile:
                    boost += 10; extras["phone_type"] = "mobile"
                elif tel:
                    boost += 6; extras["phone_type"] = "fixed"
                return boost, extras

            out: List[Dict[str, Any]] = []
            snippet_lower = (snippet_hint or "").lower()
            for r in fast_items:
                check_title = (title_text or search_title_hint or page_title or snippet_hint or "").lower()
                combined_check = f"{check_title} {snippet_lower}"
                strict_hit = next((m for m in STRICT_JOB_AD_MARKERS if m in combined_check), None)
                if strict_hit:
                    log("debug", "Dropped: Detected Strict Job Ad", marker=strict_hit, url=url)
                    continue
                keep_intent, intent_reason = analyze_page_intent(text_lower_global, check_title)
                if not keep_intent:
                    log("debug", "Dropped: Intent filter", reason=intent_reason, url=url)
                    continue
                r = _ensure_candidate_name(r, text, soup, url, page_title, h1_hint, linkedin_profile=linkedin_profile)
                lead_type_guess = classify_lead(r, check_title, text)
                if lead_type_guess == "job_ad":
                    log("debug", "Dropped: Lead classified as Job Ad", url=url)
                    continue
                if "Sales Rep" in intent_reason and not r.get("role_guess"):
                    r["role_guess"] = "Handelsvertreter/Agency"
                # FILTER 1: Employer/Generic E-Mail aussortieren
                lead_email = (r.get("email") or "").lower()
                if lead_email and is_employer_email(lead_email):
                    log("debug", "Dropped: Employer/Generic Email", email=lead_email, url=url)
                    continue
                # FILTER 2: Strikte Kandidaten-Intention + Retail Allow
                is_gesuche_path = "/s-gesuche/" in url.lower()
                has_candidate_intent = any(k in combined_check or k in text.lower() for k in ["suche", "gesuch", "biete", "lebenslauf", "cv", "profil", "mich vor", "verfuegbar", "experience"])
                is_retail = any(rtok in combined_check for rtok in RETAIL_ROLES)
                if not (is_gesuche_path or has_candidate_intent):
                    log("debug", "Dropped: No Candidate Intent in Title", title=check_title, url=url)
                    continue
                role_guess_local = r.get("role_guess") or role_guess
                if is_retail:
                    r_tags = (r.get("tags") or "")
                    parts = [p for p in r_tags.split(",") if p]
                    if "retail_candidate" not in parts:
                        parts.append("retail_candidate")
                    r["tags"] = ",".join(parts)
                    r["role_guess"] = "Retail/Kitchen Sales"
                    r["score"] = max(r.get("score", 0), 70)
                    role_guess_local = r["role_guess"]
                boost = 0
                if r.get("email"):
                    b = _mail_boost_and_label(r.get("email"))
                    if b[0] is None:
                        continue
                    boost += b[0]
                tb, extras = _tel_wa_boost(r, text, html)
                boost += tb

                tag_local = (base_tags or "").strip()
                if extras.get("tags_add_whatsapp"):
                    parts = [p for p in tag_local.split(",") if p]
                    if "whatsapp" not in parts:
                        parts.append("whatsapp")
                    tag_local = ",".join(parts)
                if extras.get("tags_add_telegram"):
                    parts = [p for p in tag_local.split(",") if p]
                    if "telegram" not in parts:
                        parts.append("telegram")
                    tag_local = ",".join(parts)
                if role_category:
                    parts = [p for p in tag_local.split(",") if p]
                    if role_category not in parts:
                        parts.append(role_category)
                    tag_local = ",".join(parts)
                if company_domain:
                    parts = [p for p in tag_local.split(",") if p]
                    dom_tag = f"domain:{company_domain}"
                    if dom_tag not in parts:
                        parts.append(dom_tag)
                    tag_local = ",".join(parts)

                # Lead-Type Logik
                lt = "other"
                if "group_invite" in (base_tags or ""):
                    lt = "group_invite"
                elif lead_type_guess:
                    lt = lead_type_guess
                elif "/s-stellengesuche/" in url.lower():
                    lt = "candidate"
                elif CANDIDATE_TEXT_RE.search(text) or any(k in text.lower() for k in CANDIDATE_KEYWORDS):
                    lt = "candidate"
                elif EMPLOYER_TEXT_RE.search(text) or RECRUITER_RE.search(text):
                    lt = "employer"
                if lt in ("candidate", "individual") and role_category:
                    lt = role_category
                if lead_type_guess == "company":
                    lt = "company"
                if agent_fingerprint_hit and lt != "group_invite":
                    lt = "Handelsvertreter"
                    role_guess_local = "Handelsvertreter"

                score_final = min(100, base_score + boost)
                if looks_like_company(r.get("name", "")) or looks_like_company(comp_name):
                    score_final = max(0, score_final - 30)
                if is_likely_human_name(r.get("name", "")):
                    score_final = min(100, score_final + 20)

                enriched = {
                    **r,
                    "score": score_final,
                    "tags": tag_local,
                    "region": region,
                    "role_guess": role_guess_local,
                    "lead_type": lt,
                    "salary_hint": salary_hint,
                    "commission_hint": commission_hint,
                    "opening_line": opening_line({"tags": tag_local}),
                    "ssl_insecure": "yes" if ssl_insecure else "no",
                    "company_name": comp_name,
                    "company_domain": company_domain,
                    "company_size": company_size,
                    "hiring_volume": hiring_volume,
                    "industry": industry,
                    "recency_indicator": recency,
                    "location_specific": locations,
                    "confidence_score": 0,
                    "last_updated": last_updated,
                    "data_quality": 0,
            "phone_type": extras.get("phone_type",""),
            "whatsapp_link": "yes" if extras.get("whatsapp_link") else "no",
            "private_address": private_address,
            "social_profile_url": social_profile_url,
            "ai_category": ai_category,
            "ai_summary": ai_summary,
            "lead_type": ai_lead_type or lt,
        }
                enriched["confidence_score"] = _confidence(enriched)
                enriched["data_quality"] = _quality(enriched)
                drop, drop_reason = should_drop_lead(
                    enriched,
                    url,
                    text,
                    page_title or title_text or search_title_hint or snippet_hint,
                )
                if drop:
                    _record_drop(drop_reason)
                    continue
                ok, reason = is_high_quality_lead(enriched, linkedin_profile=linkedin_profile, agent_hit=agent_fingerprint_hit)
                if not ok:
                    log("debug", "Dropped: Quality filter", reason=reason, url=url)
                    continue
                out.append(enriched)

            if extra_followups:
                for fu in extra_followups:
                    try:
                        inc2, items2 = await process_link_async(fu, run_id, force=force)
                        extra_checked += inc2
                        if items2:
                            out.extend(items2)
                    except Exception:
                        pass

            mark_url_seen(url, run_id)
            return (1 + extra_checked, out)

    # Grundscore
    lead_score = compute_score(text, url, html=html)
    if private_address:
        lead_score += 15
    if social_profile_url:
        lead_score += 15
    lead_score = min(100, lead_score)
    if (lead_score < CFG.min_score) and (not using_linkedin_snippet):
        log("debug", "Lead zu schwach", url=url, score=lead_score)
        mark_url_seen(url, run_id)
        return (1, [])

    # --- Kontakte extrahieren: Regex zuerst, LLM nur wenn noetig ---
    items: List[Dict[str, Any]] = []

    # 1) Regex first (schnell & kostenlos)
    items = regex_extract_contacts(text, url)

    # 1b) LLM-Kontakte (Name/Rolle/Telefon) additiv mergen
    ai_contacts: List[Dict[str, Any]] = []
    if OPENAI_API_KEY:
        ai_contacts = await extract_contacts_with_ai(text, url)
        if ai_contacts:
            deduped: List[Dict[str, Any]] = []
            seen_e, seen_t = set(), set()
            for rec in items + ai_contacts:
                e = (rec.get("email") or "").lower()
                t = rec.get("telefon") or ""
                if (e and e in seen_e) or (t and t in seen_t):
                    continue
                deduped.append(rec)
                if e: seen_e.add(e)
                if t: seen_t.add(t)
            items = deduped

    # 2) LLM nur, wenn keine verwertbaren Kontakte ODER nur E-Mails ohne Telefon
    need_llm = (len(items) == 0) or all(not r.get("telefon") for r in items)
    if need_llm and OPENAI_API_KEY:
        loop = asyncio.get_running_loop()
        try:
            max_bytes = int(os.getenv("OPENAI_MAX_BYTES", "12000"))
            llm_items = await loop.run_in_executor(None, openai_extract_contacts, text[:max_bytes], url)
            if llm_items:
                items = llm_items
        except Exception as e:
            log("warn", "openai_extract_contacts failed", url=url, err=str(e))

    # 3) Kleinanzeigen-Extractor als letzter Fallback (nur bei Domain)
    if not items and "kleinanzeigen.de" in url:
        items = extract_kleinanzeigen(html, url)

    if items:
        items = [_ensure_candidate_name(r, text, soup, url, page_title, h1_hint, linkedin_profile=linkedin_profile) for r in items]

    if not items:
        mark_url_seen(url, run_id)
        return (1 + extra_checked, [])

    # Enrichment
    title_tag = soup.find('title') if soup else None
    comp_name = extract_company_name(title_tag.get_text() if title_tag else "")
    company_size = detect_company_size(text)
    company_domain = resolve_company_domain(comp_name) if comp_name else ""
    industry = detect_industry(text)
    recency = detect_recency(html)
    hiring_volume = estimate_hiring_volume(text)
    locations = extract_locations(text)
    base_tags = tags_from(text)
    role_guess = "Callcenter" if CALLCENTER_HINT.search(text) else ("D2D" if D2D_HINT.search(text) else "Vertrieb")
    salary_hint = "low" if any(h in (text or "").lower() for h in LOW_PAY_HINT) else ""
    commission_hint = "yes" if PROVISION_HINT.search(text) else ("maybe" if SALES_RE.search(text) else "")
    region = "NRW" if CITY_RE.search(text) else ""
    last_updated = datetime.now(timezone.utc).isoformat(timespec="seconds")

    def _confidence(r: Dict[str, Any]) -> int:
        c = 0
        if r.get("telefon"): c += 50
        if r.get("email"): c += 30
        if "whatsapp" in (r.get("tags","") or "") or WA_LINK_RE.search(html): c += 10
        if comp_name: c += 10
        return min(c, 100)

    def _quality(r: Dict[str, Any]) -> int:
        q = 0
        if validate_contact(r, page_url=url, page_text=text): q += 50
        if hiring_volume != "niedrig": q += 10
        if recency != "unbekannt": q += 20
        if not ssl_insecure: q += 10
        return min(q, 100)

    def _mail_boost_and_label(email: str):
        q = email_quality(email or "", url)
        if q == "reject":
            return (None, q)
        boost = {"personal": 20, "team": 2, "generic": -6, "weak": 0}.get(q, 0)
        return (boost, q)

    def _tel_wa_boost(record: Dict[str, Any], page_text: str, page_html: str):
        t = (page_text or "").lower()
        h = (page_html or "").lower()
        tel = (record.get("telefon") or "").strip()
        boost = 0
        extras = {}
        has_wa_word = ("whatsapp" in t) or bool(WHATSAPP_PHRASE_RE.search(t))
        has_wa_link = bool(WA_LINK_RE.search(h)) or bool(WA_LINK_RE.search(t))
        if has_wa_word or has_wa_link:
            boost += 12
            extras["tags_add_whatsapp"] = True
            extras["whatsapp_link"] = bool(has_wa_link)
        has_tg_link = bool(TELEGRAM_LINK_RE.search(h)) or bool(TELEGRAM_LINK_RE.search(t))
        if has_tg_link:
            boost += 6
            extras["tags_add_telegram"] = True
        is_mobile = bool(re.search(r'(?<!\d)(?:\+49|0049|0)\s*1[5-7]\d(?:[\s\/\-]?\d){6,}(?!\d)', tel)) if tel else False
        if is_mobile:
            boost += 20
            extras["phone_type"] = "mobile"
        elif tel:
            boost += 6
            extras["phone_type"] = "fixed"
        return boost, extras

    out: List[Dict[str, Any]] = []
    snippet_lower = (snippet_hint or "").lower()
    for r in items:
        # Reject früh, damit kein Müll reinkommt
        if not (r.get("email") or r.get("telefon")):
            continue
        lead_email = (r.get("email") or "").lower()
        if lead_email and is_employer_email(lead_email):
            log("debug", "Dropped: Employer/Generic Email", email=lead_email, url=url)
            continue
        check_title = (title_text or search_title_hint or page_title or snippet_hint or "").lower()
        combined_check = f"{check_title} {snippet_lower}"
        strict_hit = next((m for m in STRICT_JOB_AD_MARKERS if m in combined_check), None)
        if strict_hit:
            log("debug", "Dropped: Detected Strict Job Ad", marker=strict_hit, url=url)
            continue
        keep_intent, intent_reason = analyze_page_intent(text_lower_global, check_title)
        if not keep_intent:
            log("debug", "Dropped: Intent filter", reason=intent_reason, url=url)
            continue
        lead_type_guess = classify_lead(r, check_title, text)
        if lead_type_guess == "job_ad":
            log("debug", "Dropped: Lead classified as Job Ad", url=url)
            continue
        if "Sales Rep" in intent_reason and not r.get("role_guess"):
            r["role_guess"] = "Handelsvertreter/Agency"
        is_gesuche_path = "/s-gesuche/" in url.lower()
        has_candidate_intent = any(k in combined_check or k in text.lower() for k in ["suche", "gesuch", "biete", "lebenslauf", "cv", "profil", "mich vor", "verfuegbar", "experience"])
        is_retail = any(rtok in combined_check for rtok in RETAIL_ROLES)
        if not (is_gesuche_path or has_candidate_intent or linkedin_profile):
            log("debug", "Dropped: No Candidate Intent in Title", title=check_title, url=url)
            continue
        role_guess_local = r.get("role_guess") or role_guess
        if is_retail:
            r_tags = (r.get("tags") or "")
            parts = [p for p in r_tags.split(",") if p]
            if "retail_candidate" not in parts:
                parts.append("retail_candidate")
            r["tags"] = ",".join(parts)
            r["role_guess"] = "Retail/Kitchen Sales"
            r["score"] = max(r.get("score", 0), 70)
            role_guess_local = r["role_guess"]
        if not validate_contact(r, page_url=url, page_text=text):
            continue

        boost = 0
        if r.get("email"):
            boost_label = _mail_boost_and_label(r.get("email"))
            if boost_label[0] is None:
                continue
            boost += boost_label[0]

        tel_boost, extras = _tel_wa_boost(r, text, html)
        boost += tel_boost

        # Tags lokal ergänzen
        tag_local = (base_tags or "").strip()
        if extras.get("tags_add_whatsapp"):
            parts = [p for p in tag_local.split(",") if p]
            if "whatsapp" not in parts:
                parts.append("whatsapp")
            tag_local = ",".join(parts)
        if extras.get("tags_add_telegram"):
            parts = [p for p in tag_local.split(",") if p]
            if "telegram" not in parts:
                parts.append("telegram")
            tag_local = ",".join(parts)
        if role_category:
            parts = [p for p in tag_local.split(",") if p]
            if role_category not in parts:
                parts.append(role_category)
            tag_local = ",".join(parts)
        if company_domain:
            parts = [p for p in tag_local.split(",") if p]
            dom_tag = f"domain:{company_domain}"
            if dom_tag not in parts:
                parts.append(dom_tag)
            tag_local = ",".join(parts)

        # Lead-Type Logik
        lt = "other"
        if "group_invite" in (base_tags or ""):
            lt = "group_invite"
        elif lead_type_guess:
            lt = lead_type_guess
        elif "/s-stellengesuche/" in url.lower():
            lt = "candidate"
        elif CANDIDATE_TEXT_RE.search(text) or any(k in text.lower() for k in CANDIDATE_KEYWORDS):
            lt = "candidate"
        elif EMPLOYER_TEXT_RE.search(text) or RECRUITER_RE.search(text):
            lt = "employer"
        if lt in ("candidate", "individual") and role_category:
            lt = role_category
        if lead_type_guess == "company":
            lt = "company"
        if agent_fingerprint_hit and lt != "group_invite":
            lt = "Handelsvertreter"
            role_guess_local = "Handelsvertreter"

        score_final = min(100, lead_score + boost)
        if looks_like_company(r.get("name", "")) or looks_like_company(comp_name):
            score_final = max(0, score_final - 30)
        if is_likely_human_name(r.get("name", "")):
            score_final = min(100, score_final + 20)

        r["_phone_type"] = extras.get("phone_type", "")
        r["_whatsapp_link"] = "yes" if extras.get("whatsapp_link") else "no"
        r["company_domain"] = company_domain

        enriched = {
            **r,
            "score": score_final,
            "tags": tag_local,
            "region": region,
            "role_guess": role_guess_local,
            "lead_type": lt,
            "salary_hint": salary_hint,
            "commission_hint": commission_hint,
            "opening_line": opening_line({"tags": tag_local}),
            "ssl_insecure": "yes" if ssl_insecure else "no",
            "company_name": comp_name,
            "company_size": company_size,
            "hiring_volume": hiring_volume,
            "industry": industry,
            "recency_indicator": recency,
            "location_specific": locations,
            "confidence_score": 0,
            "last_updated": last_updated,
            "data_quality": 0,
            "phone_type": r.get("_phone_type", ""),
            "whatsapp_link": r.get("_whatsapp_link", ""),
            "private_address": private_address,
            "social_profile_url": social_profile_url,
            "ai_category": ai_category,
            "ai_summary": ai_summary,
            "lead_type": ai_lead_type or lt,
        }
        enriched["confidence_score"] = _confidence(enriched)
        enriched["data_quality"] = _quality(enriched)
        drop, drop_reason = should_drop_lead(
            enriched,
            url,
            text,
            page_title or title_text or search_title_hint or snippet_hint,
        )
        if drop:
            _record_drop(drop_reason)
            continue
        ok, reason = is_high_quality_lead(enriched, linkedin_profile=linkedin_profile, agent_hit=agent_fingerprint_hit)
        if not ok:
            log("debug", "Dropped: Quality filter", reason=reason, url=url)
            continue
        out.append(enriched)

    if extra_followups:
        for fu in extra_followups:
            try:
                inc2, items2 = await process_link_async(fu, run_id, force=force)
                extra_checked += inc2
                if items2:
                    out.extend(items2)
            except Exception:
                pass

    mark_url_seen(url, run_id)
    return (1 + extra_checked, out)



# =========================
# Sitemap + interne Tiefe (robust)
# =========================

def find_internal_links(html: str, base_url: str) -> List[str]:
    try:
        soup = BeautifulSoup(html, "html.parser")
        links: List[str] = []
        for a in soup.find_all("a", href=True):
            href = urllib.parse.urljoin(base_url, a["href"])
            p = urllib.parse.urlparse(href)
            if p.scheme not in ("http", "https"):
                continue
            if urllib.parse.urlparse(base_url).netloc != p.netloc:
                continue
            if is_denied(href):
                continue
            if not path_ok(href):
                continue
            links.append(href)
        # Dedupe unter Erhalt der Reihenfolge
        return list(dict.fromkeys(links))
    except Exception:
        return []


async def try_sitemaps_async(base: str) -> List[str]:
    global _SITEMAP_FAILED_HOSTS
    host = urllib.parse.urlparse(base).netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    SITEMAP_SKIP_HOSTS = {
        "kleinanzeigen.de", "ebay-kleinanzeigen.de",
        "t.me", "telegram.me", "facebook.com", "m.facebook.com",
        "instagram.com", "www.instagram.com", "staseve.eu",
        "locanto.at", "www.locanto.at", "twitter.com", "x.com"
    }
    if (not host) or host in SITEMAP_SKIP_HOSTS or host in _SITEMAP_FAILED_HOSTS:
        return []
    candidates = ["/sitemap.xml", "/sitemap_index.xml", "/sitemap-index.xml"]
    out: List[str] = []
    for c in candidates:
        url = urllib.parse.urljoin(base, c)
        r = await http_get_async(url, timeout=10)
        if not r or r.status_code != 200:
            continue
        txt = r.text or ""
        try:
            soup = BeautifulSoup(txt, "xml")
            urls = [loc.get_text(strip=True) for loc in soup.find_all("loc")]
        except Exception:
            urls = re.findall(r'<loc>(.*?)</loc>', txt, flags=re.I)
        for u in urls:
            u = (u or "").strip()
            if not u:
                continue
            if is_denied(u):
                continue
            if not path_ok(u):
                continue
            out.append(u)
        if out:
            break
    # Dedupe
    out = list(dict.fromkeys(out))
    if not out:
        _SITEMAP_FAILED_HOSTS.add(host)
    return out

# =========================
# Domain-Pivot Seeds
# =========================

def domain_pivot_queries(domain:str)->list[str]:
    return [
        f'site:{domain} (kontakt OR impressum OR ansprechpartner OR team)',
        f'site:{domain} (telesales OR outbound OR d2d OR "per WhatsApp bewerben")'
    ]

# =========================
# OpenAI
# =========================

def openai_extract_contacts(raw_text: str, src_url: str) -> List[Dict[str, Any]]:
    if not OPENAI_API_KEY:
        return []
    import requests, re

    def _preview(txt: Optional[str]) -> str:
        if not txt: return ""
        txt = re.sub(r"<[^>]+>", " ", txt)
        txt = re.sub(r"\s+", " ", txt).strip()
        return txt[:200]

    max_bytes = int(os.getenv("OPENAI_MAX_BYTES", "12000"))
    snippet = (raw_text or "")[:max_bytes]
    system = (
        "Extrahiere ausschließlich reale Vertrieb/Sales-Kontaktpersonen aus deutschem Webtext.\n"
        "Gib STRICT ein JSON-Objekt mit Schlüssel 'data' zurück:\n"
        "{\"data\":[{\"name\":str,\"rolle\":str,\"email\":str,\"telefon\":str,\"quelle\":str}]}\n"
        "Nur valide E-Mails/DE-Telefone; fehlende Felder als leere Strings. Keine Halluzinationen."
    )
    user = f"Quelle: {src_url}\n\nText:\n{snippet}"
    url = "https://api.openai.com/v1/chat/completions"
    headers = {"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": os.getenv("OPENAI_MODEL", "gpt-4.1-mini"),
        "temperature": 0.0,
        "response_format": {"type": "json_object"},
        "messages": [{"role": "system", "content": system}, {"role": "user", "content": user}],
    }
    max_attempts = 4
    backoff = 1.5
    last_err = None
    for attempt in range(1, max_attempts + 1):
        try:
            r = requests.post(url, headers=headers, json=payload, timeout=60)
            status = r.status_code
            if status == 200:
                try:
                    j = r.json()
                except Exception as je:
                    last_err = f"JSON decode error: {je}"
                    raise
                
                # Track API cost if usage data is available
                try:
                    usage = j.get("usage", {})
                    if usage:
                        from dashboard.db_schema import track_api_cost
                        con = db()
                        track_api_cost(
                            con,
                            provider='openai',
                            tokens_input=usage.get('prompt_tokens', 0),
                            tokens_output=usage.get('completion_tokens', 0),
                            model=payload.get('model', 'gpt-4o-mini'),
                            endpoint='chat/completions'
                        )
                        con.close()
                except (ImportError, AttributeError) as e:
                    # Dashboard not available or schema not initialized
                    pass
                except Exception as e:
                    # Log other errors but don't fail the API call
                    import sys
                    if '--verbose' in sys.argv or os.getenv('DEBUG'):
                        print(f"Warning: Cost tracking failed: {e}")
                
                choices = (j.get("choices") or [])
                if not choices or "message" not in choices[0] or "content" not in choices[0]["message"]:
                    log("error", "OpenAI Format unerwartet", url=src_url, raw=str(j)[:200])
                    return []
                content = choices[0]["message"]["content"]
                try:
                    obj = json.loads(content)
                except Exception:
                    log("error", "OpenAI JSON in content ungültig", url=src_url, content_preview=content[:200])
                    return []
                data = obj.get("data")
                if not isinstance(data, list):
                    log("error", "OpenAI JSON ohne 'data'[]", url=src_url, content_preview=str(obj)[:200])
                    return []
                cleaned: List[Dict[str, Any]] = []
                for item in data:
                    if not isinstance(item, dict):
                        continue
                    rec = {
                        "name": (item.get("name") or "").strip(),
                        "rolle": (item.get("rolle") or "").strip(),
                        "email": (item.get("email") or "").strip(),
                        "telefon": normalize_phone(item.get("telefon") or ""),
                        "quelle": src_url,
                    }
                    if rec.get("email") or rec.get("telefon"):
                        if validate_contact(rec, page_url=src_url, page_text=raw_text):
                            cleaned.append(rec)
                # dedupe
                dedup, seen_e, seen_t = [], set(), set()
                for x in cleaned:
                    e = (x.get("email") or "").lower()
                    t = x.get("telefon") or ""
                    if (e and e in seen_e) or (t and t in seen_t):
                        continue
                    dedup.append(x)
                    if e: seen_e.add(e)
                    if t: seen_t.add(t)
                log("info", "OpenAI-Extraktion", url=src_url, count=len(dedup))
                return dedup
            # Nicht-200 → nur kurze, bereinigte Preview loggen
            log("warn", "OpenAI Antwort != 200", status=status, body=_preview(r.text), url=src_url)
            last_err = f"HTTP {status}"
            if status in (429, 500, 502, 503, 504):
                time.sleep(backoff * attempt)
                continue
            return []
        except Exception as e:
            last_err = str(e)
            time.sleep(backoff * attempt)
    log("error", "OpenAI-Extraktion fehlgeschlagen", url=src_url, error=(last_err or "")[:200])
    return []

async def generate_smart_dorks(industry: str, count: int = 5) -> List[str]:
    """
    LLM-generierte Dorks für die angegebene Branche.
    CRITICAL FIX: Special prompt for candidates/recruiter mode.
    """
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return []
    
    # CRITICAL FIX: Different prompt for candidates mode
    if industry.lower() in ("candidates", "recruiter"):
        prompt = (
            "Generate Google Dorks to find JOB SEEKERS (not companies hiring). "
            "Target: People actively looking for sales/vertrieb jobs. "
            "Required patterns: "
            'site:kleinanzeigen.de/s-stellengesuche "vertrieb"; '
            'site:xing.com/profile "offen für angebote" "sales"; '
            'site:linkedin.com/in "#opentowork" "vertrieb"; '
            '"ich suche job" "vertrieb" "NRW" "mobil"; '
            'site:facebook.com "suche arbeit" "verkauf". '
            "Return ONLY the dorks, one per line."
        )
        system_content = "You create Google search dorks to find job seekers and candidates."
    else:
        # Standard B2B prompt
        prompt = (
            "You are a Headhunter. Generate Google Dorks that find lists of employees, PDF CVs, 'Unser Team' pages, or Freelancer profiles. "
            "forbidden: Do NOT generate generic B2B searches like 'Händler' or 'Hersteller'. "
            "Required Patterns (mix these): "
            f'intitle:\"Team\" \"Sales Manager\" {industry}; '
            f'filetype:pdf \"Lebenslauf\" {industry} -job -anzeige; '
            f'site:linkedin.com/in/ \"{industry}\" \"open to work\"; '
            f'\"stellengesuch\" {industry} \"verfügbar ab\"; '
            f'\"Ansprechpartner\" \"Vertrieb\" {industry} -intitle:Jobs. '
            "Return ONLY the dorks, one per line."
        )
        system_content = "You create targeted Google search dorks for B2B lead generation."
    
    payload = {
        "model": os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
        "messages": [
            {"role": "system", "content": system_content},
            {"role": "user", "content": prompt}
        ]
    }
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post("https://api.openai.com/v1/chat/completions", headers=headers, json=payload, timeout=HTTP_TIMEOUT) as resp:
                if resp.status != 200:
                    log("warn", "Smart dorks HTTP error", status=resp.status)
                    return []
                data = await resp.json()
                content = ""
                choices = data.get("choices") or []
                if choices and isinstance(choices, list):
                    content = ((choices[0] or {}).get("message") or {}).get("content", "")  # type: ignore[index]
                if not content:
                    return []
                lines = [ln.strip(" -*\t") for ln in content.splitlines() if ln.strip()]
                uniq = []
                seen = set()
                for ln in lines:
                    if ln.lower() in seen:
                        continue
                    seen.add(ln.lower())
                    uniq.append(ln)
                    if len(uniq) >= max(1, count):
                        break
                return uniq
    except Exception as e:
        log("warn", "Smart dorks generation failed", error=str(e))
        return []
    return []

# =========================
# Export (CSV/XLSX append)
# =========================

def append_csv(path: str, rows: List[Dict[str, Any]], fieldnames: List[str]):
    if not rows:
        return
    import csv, os
    rewrite_existing = False
    existing_rows: List[Dict[str, Any]] = []
    if os.path.exists(path) and os.path.getsize(path) > 0:
        try:
            with open(path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                header = reader.fieldnames or []
                if header != fieldnames:
                    rewrite_existing = True
                    for row in reader:
                        existing_rows.append({k: row.get(k, "") for k in fieldnames})
        except Exception:
            pass
    write_header = (not os.path.exists(path)) or (os.path.getsize(path) == 0) or rewrite_existing
    mode = "w" if rewrite_existing else "a"
    def _coerce(row: Dict[str, Any]) -> Dict[str, Any]:
        safe = {k: ("" if row.get(k) is None else row.get(k)) for k in fieldnames}
        if "whatsapp_link" in safe:
            safe["whatsapp_link"] = int(bool(safe["whatsapp_link"]))
        return safe
    try:
        with open(path, mode, newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            if write_header:
                w.writeheader()
            if rewrite_existing and existing_rows:
                for r in existing_rows:
                    w.writerow(_coerce(r))
            for r in rows:
                w.writerow(_coerce(r))
    except PermissionError:
        alt = os.path.join(os.path.expanduser("~"), f"vertrieb_kontakte_{int(time.time())}.csv")
        with open(alt, "a", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            if (not os.path.exists(alt)) or (os.path.getsize(alt) == 0):
                w.writeheader()
            for r in rows:
                w.writerow(_coerce(r))
        log("warn", "CSV war gelockt – in Home geschrieben", file=alt)

def append_xlsx(path: str, rows: List[Dict[str, Any]], fieldnames: List[str]):
    if not rows:
        return
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        log("warn", "openpyxl nicht installiert – XLSX übersprungen")
        return
    def _coerce_row(row: Dict[str, Any]) -> List[Any]:
        safe = [("" if row.get(k) is None else row.get(k)) for k in fieldnames]
        try:
            idx = fieldnames.index("whatsapp_link")
            safe[idx] = int(bool(safe[idx]))
        except ValueError:
            pass
        return safe

    # Robuste Header-Behandlung (fehlende Spalten am Ende ergänzen)
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        if ws.max_row < 1 or ws["A1"].value is None:
            ws.append(fieldnames)
        else:
            current = [c.value for c in ws[1] if c.value is not None]
            if current != fieldnames:
                for f in fieldnames[len(current):]:
                    ws.cell(row=1, column=len(current) + 1, value=f)
                    current.append(f)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(fieldnames)

    for r in rows:
        ws.append(_coerce_row(r))

    for i, col in enumerate(fieldnames, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(col) + 2, 12), 40)
    wb.save(path)

# =========================
# Hauptlauf (ASYNC)
# =========================

class _Rate:
    def __init__(self, max_global:int=ASYNC_LIMIT, max_per_host:int=ASYNC_PER_HOST):
        self.sem_global = asyncio.Semaphore(max(1, max_global))
        self.per_host: Dict[str, asyncio.Semaphore] = {}
        self.max_per_host = max(1, max_per_host)
        self.lock = asyncio.Lock()

    async def acquire(self, url:str):
        host = _host_from(url)
        async with self.lock:
            if host not in self.per_host:
                self.per_host[host] = asyncio.Semaphore(self.max_per_host)
        await self.sem_global.acquire()
        await self.per_host[host].acquire()
        return host

    def release(self, host:str):
        try:
            self.sem_global.release()
            self.per_host.get(host, asyncio.Semaphore(1)).release()
        except Exception:
            pass

async def _bounded_process(urls: List[UrlLike], run_id:int, *, rate:_Rate, force:bool=False):
    """Prozessiert URLs mit globalem/per-Host-Limit. Liefert (links_checked, leads)."""
    links_checked = 0
    collected: List[Dict[str, Any]] = []
    SNIPPET_EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", re.I)
    SNIPPET_PHONE_RE = re.compile(r"(?:\+49|0)[1-9][0-9 \-/]{6,}")

    def _name_from_title(title: str) -> str:
        if not title:
            return ""
        part = title.split("|", 1)[0]
        part = part.split(" - ", 1)[0]
        part = part.strip(" -|")
        part = re.sub(r"\s+", " ", part).strip()
        # Mindestlänge: Vor- + Nachname
        if len(part.split()) >= 2:
            return part
        return ""

    async def _one(item: UrlLike, url: str):
        nonlocal links_checked, collected
        host = await rate.acquire(url)
        try:
            inc, items = await process_link_async(item, run_id, force=force)
            links_checked += int(inc)
            if items:
                collected.extend(items)
        finally:
            rate.release(host)

    candidates = []
    for entry in urls:
        raw_url = _extract_url(entry)
        if not raw_url:
            continue
        norm_url = _normalize_for_dedupe(raw_url)
        candidates.append((entry, raw_url, norm_url))

    if not candidates:
        return links_checked, collected

    ordered = prioritize_urls([c[2] for c in candidates])
    # Verarbeitung von Google-Snippets (Snippet-Harvesting, z.B. Facebook blockt robots)
    seed_leads = []
    snippet_leads: List[Dict[str, Any]] = []
    for item in urls:
        if isinstance(item, dict):
            u = item.get("url", "")
            title = item.get("title", "") or ""
            snip = item.get("snippet", "") or ""
            snippet_text = f"{title} {snip}"
            name = extract_name_enhanced(snippet_text)
            title_name = _name_from_title(title)
            if title_name:
                name = title_name
            rolle, _ = extract_role_with_context(snippet_text, u)
            company = extract_company_name(title)
            if name or rolle or company:
                seed_leads.append({
                    "name": name or "",
                    "rolle": rolle or "",
                    "company_guess": company or "",
                    "quelle": u,
                    "score": 0,
                    "telefon": "",
                    "email": ""
                })
            emails = [m.group(0) for m in SNIPPET_EMAIL_RE.finditer(snippet_text)]
            phones_raw = [normalize_phone(m.group(0)) for m in SNIPPET_PHONE_RE.finditer(snippet_text)]
            emails = list(dict.fromkeys([e for e in emails if e]))
            # Phone hardfilter: validate phones early (fast-path)
            phones_validated = []
            for p in phones_raw:
                if p:
                    is_valid, _ = validate_phone(p)
                    if is_valid:
                        phones_validated.append(p)
            phones = list(dict.fromkeys(phones_validated))
            if emails or phones:
                try:
                    log("info", f"💰 SNIPPET-JACKPOT: {u} -> Mail: {len(emails)}, Tel: {len(phones)}")
                except Exception:
                    pass
                
                # CRITICAL: Check for job postings first - NEVER save as lead
                if is_job_posting(url=u, title=title, snippet=snip):
                    log("debug", "Snippet jackpot dropped (job posting)", url=u)
                    continue
                
                # Check if we have a mobile number
                tel = ""
                has_mobile = False
                for phone in phones:
                    if is_mobile_number(phone):
                        tel = phone
                        has_mobile = True
                        break
                
                # If no mobile number but has contact, still process for deep crawl
                # (might find mobile on the actual page)
                if not has_mobile and phones:
                    log("debug", "Snippet jackpot has landline only - may deep crawl", url=u)
                    # Don't save as snippet lead, let it go through normal processing
                    continue
                
                # Only create snippet lead if we have mobile number
                if not has_mobile and not emails:
                    continue
                
                tags_list = [t for t in (tags_from(snippet_text) or "").split(",") if t]
                tags_list.append("snippet")
                if has_mobile:
                    tags_list.append("mobile")
                tags_local = ",".join(dict.fromkeys(tags_list))
                
                phone_type = "mobile" if has_mobile else ""
                score = 95 if has_mobile else 80
                snippet_leads.append({
                    "name": name or "",
                    "rolle": rolle or "",
                    "email": emails[0] if emails else "",
                    "telefon": tel,
                    "quelle": u,
                    "score": score,
                    "tags": tags_local,
                    "region": "",
                    "role_guess": rolle or "Snippet",
                    "lead_type": "candidate",
                    "salary_hint": "",
                    "commission_hint": "",
                    "opening_line": opening_line({"tags": tags_local}),
                    "ssl_insecure": "no",
                    "company_name": company or "",
                    "company_size": "unbekannt",
                    "hiring_volume": "niedrig",
                    "industry": "unbekannt",
                    "recency_indicator": "unbekannt",
                    "location_specific": "",
                    "confidence_score": 80 if (tel or emails) else 50,
                    "last_updated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    "data_quality": 50,
                    "phone_type": phone_type,
                    "whatsapp_link": "no",
                    "private_address": "",
                    "social_profile_url": "",
                    "source": "snippet"
                })
    # Seed-Leads direkt in 'collected' einspeisen (nur wenn sinnvoll)
    if seed_leads:
        collected.extend(seed_leads)
    if snippet_leads:
        collected.extend(snippet_leads)
    order_map = {u: i for i, u in enumerate(ordered)}
    candidates.sort(key=lambda tpl: order_map.get(tpl[2], len(order_map)))

    await asyncio.gather(*[_one(item, raw) for item, raw, _ in candidates])
    return links_checked, collected

async def process_retry_urls(run_id: int, rate: _Rate) -> Tuple[int, int]:
    """Verarbeitet die gesammelten Retry-URLs mit Budget/Backoff."""
    if not _RETRY_URLS:
        return (0, 0)
    retry_urls = list(_RETRY_URLS.keys())
    retries_total = 0
    retries_exhausted = 0
    for url in retry_urls:
        state = _RETRY_URLS.get(url) or {}
        tries = int(state.get("retries", 0))
        if tries >= RETRY_MAX_PER_URL:
            _RETRY_URLS.pop(url, None)
            retries_exhausted += 1
            continue
        delay = RETRY_BACKOFF_BASE * max(1, tries + 1) * _jitter(0.7, 1.3)
        await asyncio.sleep(delay)
        state["retries"] = tries + 1
        state["last_ts"] = time.time()
        _RETRY_URLS[url] = state
        try:
            inc, items = await _bounded_process([url], run_id, rate=rate, force=True)
        except Exception:
            inc, items = (0, [])
        retries_total += 1
        last_status = _LAST_STATUS.get(url, 0)
        if (items and inc >= 0) or last_status == 200:
            _RETRY_URLS.pop(url, None)
        elif state["retries"] >= RETRY_MAX_PER_URL:
            retries_exhausted += 1
            _RETRY_URLS.pop(url, None)
    if retries_total or retries_exhausted:
        log("info", "Retry wave completed", retries_total=retries_total, retries_exhausted=retries_exhausted, pending=len(_RETRY_URLS))
    return retries_total, retries_exhausted

def emergency_save(run_id, links_checked=None, leads_new_total=None):
    """
    Not-Speicherung:
    - kompatibel zu alten Aufrufen mit 1 oder 3 Parametern
    - versucht Export + finish_run, fällt bei Fehlern auf Minimaldump zurück
    - darf niemals eine Exception durchreichen
    """
    try:
        log("warn", "Not-Speicherung läuft ...", run_id=run_id,
            links_checked=links_checked, leads_new_total=leads_new_total)
    except Exception:
        pass

    # 1) Bestmöglicher Export/Run-Abschluss (nicht abbrechen, nur versuchen)
    try:
        try:
            export_xlsx(DEFAULT_XLSX)
        except Exception as e:
            try:
                log("error", "Export fehlgeschlagen (Ignoriert)", error=str(e))
            except Exception:
                pass

        try:
            # finish_run kann Links/Leads-Argumente optional verarbeiten
            finish_run(run_id, links_checked, leads_new_total, status="aborted", metrics=dict(RUN_METRICS))
        except TypeError:
            # Fallback für alte Signaturen
            finish_run(run_id, status="aborted", metrics=dict(RUN_METRICS))
        except Exception as e:
            try:
                log("error", "finish_run fehlgeschlagen (Ignoriert)", error=str(e))
            except Exception:
                pass
    except Exception:
        pass

    # 2) Minimaler Dump auf Platte (falls alles andere scheitert)
    try:
        import os, time, json
        stamp = time.strftime("%Y%m%d-%H%M%S")
        payload = {
            "run_id": run_id,
            "links_checked": links_checked,
            "leads_new_total": leads_new_total,
            "timestamp": stamp
        }
        path = os.path.join(os.getcwd(), f"emergency_run_{run_id}_{stamp}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception:
        # Letzter Fallback: nichts tun, niemals hochwerfen
        pass



async def run_scrape_once_async(run_flag: Optional[dict] = None, ui_log=None, force: bool = False, date_restrict: Optional[str] = None):
    def _uilog(msg, **k):
        if ui_log:
            ui_log(msg)
        else:
            log("info", msg, **k)

    init_db()
    global _seen_urls_cache
    _seen_urls_cache = set()
    con = None
    try:
        con = db(); cur = con.cursor()
        cur.execute("SELECT url FROM urls_seen")
        _seen_urls_cache = {_normalize_for_dedupe(row[0]) for row in cur.fetchall()}
    except Exception as e:
        log("warn", "Konnte URL-Cache nicht laden", error=str(e))
    else:
        log("info", "URL-Cache geladen", count=len(_seen_urls_cache))
    finally:
        if con:
            con.close()

    # Get initial performance params from dashboard
    perf_params = get_performance_params()
    current_async_limit = perf_params.get('async_limit', ASYNC_LIMIT)
    current_request_delay = perf_params.get('request_delay', SLEEP_BETWEEN_QUERIES)
    last_perf_check = time.time()
    
    rate = _Rate(max_global=current_async_limit, max_per_host=ASYNC_PER_HOST)

    total_links_checked = 0
    leads_new_total = 0
    run_id = start_run()
    _reset_metrics()
    _uilog(f"Run #{run_id} gestartet (Performance Mode: {perf_params.get('async_limit', 'N/A')} async)")

    # Direct crawling from multiple sources (only in candidates/recruiter mode)
    if _is_candidates_mode() and ENABLE_KLEINANZEIGEN:
        _uilog("Starte direktes Multi-Portal-Crawling (Stellengesuche)...")
        direct_crawl_leads = []
        
        # Kleinanzeigen.de crawling
        if DIRECT_CRAWL_SOURCES.get("kleinanzeigen", True):
            _uilog("Crawle Kleinanzeigen.de...")
            for crawl_url in DIRECT_CRAWL_URLS:
                if run_flag and not run_flag.get("running", True):
                    _uilog("STOP erkannt – breche Direct-Crawl ab")
                    break
                
                try:
                    log("info", "Direct crawl: Listing-Seite", url=crawl_url)
                    
                    # Step 1: Crawl listing page to get ad links
                    ad_links = await crawl_kleinanzeigen_listings_async(crawl_url, max_pages=5)
                    
                    if not ad_links:
                        log("info", "Direct crawl: Keine Anzeigen gefunden", url=crawl_url)
                        continue
                    
                    log("info", "Direct crawl: Anzeigen gefunden", url=crawl_url, count=len(ad_links))
                    
                    # Step 2: Extract details from each ad (with rate limiting)
                    for i, ad_url in enumerate(ad_links):
                        if run_flag and not run_flag.get("running", True):
                            break
                        
                        # Skip if already seen
                        if url_seen(ad_url):
                            log("debug", "Direct crawl: URL bereits gesehen (skip)", url=ad_url)
                            continue
                        
                        # Rate limiting between detail page fetches
                        if i > 0:
                            await asyncio.sleep(2.5 + _jitter(0.5, 1.0))
                        
                        # Extract lead data from ad detail page
                        lead_data = await extract_kleinanzeigen_detail_async(ad_url)
                        
                        if lead_data:
                            direct_crawl_leads.append(lead_data)
                            
                            # Mark URL as seen
                            try:
                                con = db(); cur = con.cursor()
                                cur.execute("INSERT OR IGNORE INTO urls_seen (url) VALUES (?)", (ad_url,))
                                con.commit()
                                con.close()
                                _seen_urls_cache.add(_normalize_for_dedupe(ad_url))
                            except Exception as e:
                                log("warn", "Konnte URL nicht als gesehen markieren", url=ad_url, error=str(e))
                    
                    # Rate limiting between listing pages
                    await asyncio.sleep(3.0 + _jitter(0.5, 1.0))
                    
                except Exception as e:
                    log("error", "Direct crawl: Fehler beim Crawlen", url=crawl_url, error=str(e))
                    continue
        
        # Markt.de crawling
        if DIRECT_CRAWL_SOURCES.get("markt_de", True):
            if run_flag and run_flag.get("running", True):
                _uilog("Crawle Markt.de...")
                try:
                    markt_leads = await crawl_markt_de_listings_async()
                    direct_crawl_leads.extend(markt_leads)
                    log("info", "Markt.de crawl complete", count=len(markt_leads))
                except Exception as e:
                    log("error", "Markt.de crawl failed", error=str(e))
        
        # Quoka.de crawling
        if DIRECT_CRAWL_SOURCES.get("quoka", True):
            if run_flag and run_flag.get("running", True):
                _uilog("Crawle Quoka.de...")
                try:
                    quoka_leads = await crawl_quoka_listings_async()
                    direct_crawl_leads.extend(quoka_leads)
                    log("info", "Quoka crawl complete", count=len(quoka_leads))
                except Exception as e:
                    log("error", "Quoka crawl failed", error=str(e))
        
        # Kalaydo.de crawling
        if DIRECT_CRAWL_SOURCES.get("kalaydo", True):
            if run_flag and run_flag.get("running", True):
                _uilog("Crawle Kalaydo.de...")
                try:
                    kalaydo_leads = await crawl_kalaydo_listings_async()
                    direct_crawl_leads.extend(kalaydo_leads)
                    log("info", "Kalaydo crawl complete", count=len(kalaydo_leads))
                except Exception as e:
                    log("error", "Kalaydo crawl failed", error=str(e))
        
        # Meinestadt.de crawling
        if DIRECT_CRAWL_SOURCES.get("meinestadt", True):
            if run_flag and run_flag.get("running", True):
                _uilog("Crawle Meinestadt.de...")
                try:
                    meinestadt_leads = await crawl_meinestadt_listings_async()
                    direct_crawl_leads.extend(meinestadt_leads)
                    log("info", "Meinestadt crawl complete", count=len(meinestadt_leads))
                except Exception as e:
                    log("error", "Meinestadt crawl failed", error=str(e))
        
        # Insert collected leads from all sources
        if direct_crawl_leads:
            log("info", "Direct crawl: Leads gefunden (alle Quellen)", count=len(direct_crawl_leads))
            _uilog(f"Direct crawl: {len(direct_crawl_leads)} Leads extrahiert (alle Portale)")
            
            # Insert into database
            new_leads = insert_leads(direct_crawl_leads)
            leads_new_total += len(new_leads)
            
            log("info", "Direct crawl: Neue Leads gespeichert", count=len(new_leads))
            _uilog(f"Direct crawl: {len(new_leads)} neue Leads gespeichert")
        else:
            log("info", "Direct crawl: Keine Leads gefunden")
            _uilog("Direct crawl: Keine Leads gefunden")

    try:
        for q in QUERIES:
            # Periodically refresh performance params (every 30 seconds)
            if time.time() - last_perf_check > 30:
                perf_params = get_performance_params()
                new_async_limit = perf_params.get('async_limit', ASYNC_LIMIT)
                current_request_delay = perf_params.get('request_delay', SLEEP_BETWEEN_QUERIES)
                if new_async_limit != current_async_limit:
                    current_async_limit = new_async_limit
                    rate = _Rate(max_global=current_async_limit, max_per_host=ASYNC_PER_HOST)
                    log("info", "Performance params updated", async_limit=current_async_limit, request_delay=current_request_delay)
                last_perf_check = time.time()
            
            if run_flag and not run_flag.get("running", True):
                _uilog("STOP erkannt – breche ab")
                break
            if (not force) and is_query_done(q):
                log("info", "Query bereits erledigt (skip)", q=q)
                await asyncio.sleep(current_request_delay)
                continue

            log("info", "Starte Query", q=q)
            had_429_flag = False
            collected_rows = []
            links: List[UrlLike] = []

            try:
                g_links, had_429 = await google_cse_search_async(q, max_results=60, date_restrict=date_restrict)
                links.extend(g_links)
                had_429_flag |= had_429
            except Exception as e:
                log("error", "Google-Suche explodiert", q=q, error=str(e))

            if had_429_flag or not links:
                try:
                    log("info", "Nutze DuckDuckGo (Fallback)...", q=q)
                    ddg_links = await duckduckgo_search_async(q, max_results=30, date_restrict=date_restrict)
                    links.extend(ddg_links)
                except Exception as e:
                    log("error", "DuckDuckGo-Suche explodiert", q=q, error=str(e))

            if had_429_flag or len(links) < 3:
                try:
                    log("info", "Nutze Perplexity (sonar)...", q=q)
                    pplx_links = await search_perplexity_async(q)
                    links.extend(pplx_links)
                except Exception as e:
                    log("error", "Perplexity-Suche explodiert", q=q, error=str(e))

            if not links:
                try:
                    ddg_links = await duckduckgo_search_async(q, max_results=30, date_restrict=date_restrict)
                    links.extend(ddg_links)
                except Exception as e:
                    log("error", "DuckDuckGo-Suche explodiert", q=q, error=str(e))

            if not links:
                log("warn", "Alle Suchmaschinen erschöpft (Google, Perplexity, DDG). Mache eine längere Pause.", q=q)
                await asyncio.sleep(current_request_delay + _jitter(1.5,2.5))

            try:
                ka_links = await kleinanzeigen_search_async(q, max_results=KLEINANZEIGEN_MAX_RESULTS)
                if ka_links:
                    links.extend(ka_links)
            except Exception as e:
                log("warn", "Kleinanzeigen-Suche explodiert", q=q, error=str(e))

            if links:
                uniq_links: List[UrlLike] = []
                seen_links = set()
                for item in links:
                    raw_url = _extract_url(item)
                    if not raw_url:
                        continue
                    nu = _normalize_for_dedupe(raw_url)
                    if nu in seen_links:
                        continue
                    seen_links.add(nu)
                    if isinstance(item, dict):
                        uniq_links.append({**item, "url": nu})
                    else:
                        uniq_links.append(nu)
                links = uniq_links

            if not links:
                if had_429_flag:
                    log("warn", "Keine Links (429) - Query NICHT als erledigt markieren", q=q)
                await asyncio.sleep(SLEEP_BETWEEN_QUERIES + _jitter(0.4,1.2))
                continue

            per_domain_count = {}
            prim: List[UrlLike] = []
            skipped_by_learning = 0
            for link in links:
                url = _extract_url(link)
                if not url:
                    continue
                dom = urllib.parse.urlparse(url).netloc.lower()
                
                # Learning mode: Skip domains with poor performance
                if ACTIVE_MODE_CONFIG and ACTIVE_MODE_CONFIG.get("learning_enabled") and _LEARNING_ENGINE:
                    domain_clean = dom[4:] if dom.startswith("www.") else dom
                    if _LEARNING_ENGINE.should_skip_domain(domain_clean):
                        log("debug", "Learning: Domain übersprungen (schlechte Historie)", domain=domain_clean)
                        skipped_by_learning += 1
                        continue
                
                extra = any(k in url.lower() for k in ("/jobs","/karriere","/stellen","/bewerb"))
                limit = CFG.max_results_per_domain + (1 if extra else 0)
                if per_domain_count.get(dom,0) >= limit:
                    continue
                per_domain_count[dom] = per_domain_count.get(dom,0)+1
                if _url_seen_fast(url):
                    continue
                if not url_seen(url):
                    prim.append(link)
            
            if skipped_by_learning > 0:
                log("info", "Learning: Domains gefiltert", skipped=skipped_by_learning)

            chk, rows = await _bounded_process(prim, run_id, rate=rate, force=False)
            total_links_checked += chk
            collected_rows.extend(rows)

            pivot_seeds: List[UrlLike] = []
            pivot_tasks = []
            for dom in per_domain_count.keys():
                for dq in domain_pivot_queries(dom):
                    pivot_tasks.append(google_cse_search_async(dq, max_results=10, date_restrict=date_restrict))
            if pivot_tasks:
                pivot_results = await asyncio.gather(*pivot_tasks, return_exceptions=True)
                for result in pivot_results:
                    if isinstance(result, Exception):
                        continue
                    if isinstance(result, tuple) and result[0]:
                        pivot_seeds.extend(result[0])
            if pivot_seeds:
                uniq_pivots = []
                seen_p = set()
                for item in pivot_seeds:
                    p_url = _extract_url(item)
                    if not p_url:
                        continue
                    norm = _normalize_for_dedupe(p_url)
                    if norm in seen_p:
                        continue
                    seen_p.add(norm)
                    uniq_pivots.append((norm, item))
                ordered_p = prioritize_urls([u for u, _ in uniq_pivots]) if uniq_pivots else []
                order_map_p = {u: i for i, u in enumerate(ordered_p)}
                uniq_pivots.sort(key=lambda tpl: order_map_p.get(tpl[0], len(order_map_p)))
                pivot_batch = [it for _, it in uniq_pivots][:CFG.internal_depth_per_domain]
                if pivot_batch:
                    chk_p, rows_p = await _bounded_process(pivot_batch, run_id, rate=rate, force=False)
                    total_links_checked += chk_p
                    collected_rows.extend(rows_p)

            for dom in list(per_domain_count.keys()):
                base = f"https://{dom}"
                internal = []
                try:
                    sm = await try_sitemaps_async(base)
                except:
                    sm = []
                if sm:
                    internal.extend(sm[:CFG.internal_depth_per_domain])
                if not internal:
                    for pl in prim:
                        pl_url = _extract_url(pl)
                        if not pl_url:
                            continue
                        if urllib.parse.urlparse(pl_url).netloc.lower() != dom:
                            continue
                        r = await http_get_async(pl_url, timeout=10)
                        if not r or r.status_code != 200:
                            continue
                        more = find_internal_links(r.text, pl_url)
                        for u in more:
                            if _url_seen_fast(u) or url_seen(u): continue
                            if is_denied(u): continue
                            if not path_ok(u): continue
                            if urllib.parse.urlparse(u).netloc.lower() != dom: continue
                            internal.append(u)
                            if len(internal)>=CFG.internal_depth_per_domain:
                                break
                        if len(internal)>=CFG.internal_depth_per_domain:
                            break
                if internal:
                    internal = prioritize_urls(internal)
                    chk2, rows2 = await _bounded_process(internal, run_id, rate=rate, force=False)
                    total_links_checked += chk2
                    collected_rows.extend(rows2)

            mark_query_done(q, run_id)

            MIN_SCORE_TARGET = MIN_SCORE_ENV
            found = len(collected_rows)
            avg = int(sum(r.get("score",0) for r in collected_rows)/found) if found else 0
            if found >= 20 and avg < MIN_SCORE_ENV:
                MIN_SCORE_TARGET=min(80,MIN_SCORE_ENV+5)
            elif found < 10 and MIN_SCORE_ENV>=45:
                MIN_SCORE_TARGET=MIN_SCORE_ENV-10
            elif found <5 and MIN_SCORE_ENV>=35:
                MIN_SCORE_TARGET=MIN_SCORE_ENV-20

            def _is_offtarget_lead(r: Dict[str, Any]) -> bool:
                lead_type = (r.get("lead_type") or "").lower()
                if lead_type == "employer":
                    return True
                role = (r.get("rolle") or r.get("role_guess") or "").lower()
                company = (r.get("company_name") or "").lower()
                src_url = (r.get("quelle") or "").lower()

                hr_tokens = (
                    "personalreferent",
                    "personalreferentin",
                    "sachbearbeiter personal",
                    "sachbearbeiterin personal",
                    "hr-manager",
                    "hr manager",
                    "human resources",
                    "bewerbungen richten",
                    "bewerbung richten"
                )
                press_tokens = (
                    "pressesprecher",
                    "pressesprecherin",
                    "unternehmenskommunikation",
                    "pressekontakt",
                    "events",
                    "veranstaltungen",
                    "seminar",
                    "seminare"
                )
                public_tokens = (
                    "rathaus",
                    "verwaltung",
                    "gleichstellungsstelle",
                    "grundsicherung",
                    "bundestag",
                    "/stadtwerke",
                    "die-partei.de",
                    "oberhausen.de"
                )

                role_hit = any(tok in role for tok in hr_tokens + press_tokens)
                company_hit = any(tok in company for tok in ("stadtwerke", "bundestag", "die partei"))
                url_hit = any(tok in src_url for tok in public_tokens)

                return role_hit or company_hit or url_hit

            # NUR Kandidaten exportieren, wenn wir im Recruiter/Candidates-Modus sind
            # CRITICAL FIX: Also check for "candidates" in addition to "recruiter"
            industry_env = os.getenv("INDUSTRY", "").lower()
            if "recruiter" in str(QUERIES).lower() or "candidates" in str(QUERIES).lower() or industry_env in ("recruiter", "candidates"):
                collected_rows = [r for r in collected_rows if r.get("lead_type") in ("candidate", "group_invite")]
                log("info", "Filter aktiv: Nur Candidates/Gruppen behalten", remaining=len(collected_rows))

            filtered = _dedup_run(
                [
                    r for r in collected_rows
                    if r.get("score", 0) >= MIN_SCORE_TARGET and not _is_offtarget_lead(r)
                ]
            )
            if filtered:
                # Enrich leads without phone numbers using telefonbuch
                filtered = await enrich_leads_with_telefonbuch(filtered)
                inserted = insert_leads(filtered)
                if inserted:
                    append_csv(DEFAULT_CSV, inserted, ENH_FIELDS)
                    append_xlsx(DEFAULT_XLSX, inserted, ENH_FIELDS)
                    _uilog(f"Export: +{len(inserted)} neue Leads")
                    leads_new_total += len(inserted)
                    
                    # Learning mode: Track domain and query performance
                    if ACTIVE_MODE_CONFIG and ACTIVE_MODE_CONFIG.get("learning_enabled") and _LEARNING_ENGINE:
                        try:
                            # Track query performance
                            _LEARNING_ENGINE.record_query_performance(q, len(inserted))
                            
                            # Track domain success for each lead
                            domains_tracked = set()
                            for lead in inserted:
                                source_url = lead.get("quelle", "")
                                if source_url:
                                    parsed = urllib.parse.urlparse(source_url)
                                    domain = parsed.netloc.lower()
                                    if domain.startswith("www."):
                                        domain = domain[4:]
                                    if domain and domain not in domains_tracked:
                                        # Calculate quality based on score and confidence (both expected to be 0-100)
                                        # Normalize to 0.0-1.0 range and average them
                                        score_val = max(0, min(100, lead.get("score", 0)))
                                        confidence_val = max(0, min(100, lead.get("confidence_score", 0)))
                                        quality = min(1.0, (score_val / 100.0 + confidence_val / 100.0) / 2)
                                        _LEARNING_ENGINE.record_domain_success(domain, 1, quality)
                                        domains_tracked.add(domain)
                            
                            if domains_tracked:
                                log("info", "Learning: Domain-Erfolge gespeichert", domains=len(domains_tracked), query_leads=len(inserted))
                        except Exception as e:
                            log("debug", "Learning tracking failed", error=str(e))

            if _RETRY_URLS:
                try:
                    await process_retry_urls(run_id, rate)
                except Exception as e:
                    log("warn", "Retry wave failed", error=str(e))

            await asyncio.sleep(current_request_delay + _jitter(0.4,1.2))

        finish_run(run_id, total_links_checked, leads_new_total, "ok", metrics=dict(RUN_METRICS))
        _uilog(f"Run #{run_id} beendet")

    except Exception as e:
        emergency_save(run_id, total_links_checked, leads_new_total)
        log("error", "Run abgebrochen", error=str(e), tb=traceback.format_exc())
        raise

    finally:
        global _CLIENT_SECURE, _CLIENT_INSECURE
        for cl in (_CLIENT_SECURE,_CLIENT_INSECURE):
            if cl:
                try: await cl.aclose()
                except: pass
        _CLIENT_SECURE=None
        _CLIENT_INSECURE=None



# =========================
# UI (Flask + SSE)
# =========================

def start_ui(host="127.0.0.1", port=5055):
    from flask import Flask, render_template_string, Response
    app = Flask(__name__)

    TEMPLATE = """
<!doctype html><meta charset="utf-8">
<title>Leads-Scraper</title>
<style>
  html,body{font:14px system-ui;margin:0;background:#0b0f0c;color:#e8efe9}
  header{padding:16px 20px;border-bottom:1px solid #1c241e;background:#0f1511}
  main{padding:20px}
  button{padding:10px 16px;margin:6px 8px 6px 0;border:0;border-radius:8px;background:#1b5e20;color:#fff;cursor:pointer}
  button.stop{background:#a82b2b}
  .warn{background:#a8742b}
  .muted{background:#2b3a2f}
  .card{background:#0f1511;border:1px solid #1c241e;border-radius:12px;padding:12px;margin-top:16px}
  pre{white-space:pre-wrap;max-height:60vh;overflow:auto;margin:0}
</style>
<header><h2>NRW Leads-Scraper</h2></header>
<main>
  <button onclick="fetch('/start',{method:'POST'})">Start</button>
  <button class="warn" onclick="fetch('/start_force',{method:'POST'})">Start (Force)</button>
  <button class="stop" onclick="fetch('/stop',{method:'POST'})">Stop</button>
  <button class="warn" onclick="fetch('/reset',{method:'POST'})">Reset (History)</button>
  <button class="muted" onclick="fetch('/reset_seen',{method:'POST'})">Seen (24h) löschen</button>
  <button class="muted" onclick="fetch('/reset_queries',{method:'POST'})">Queries zurücksetzen</button>
  <a style="margin-left:12px;color:#9fe29f" href="/logs" target="_blank">Logs Stream</a>
  <div class="card"><pre id="p"></pre></div>
</main>
<script>
const p=document.getElementById('p');
const es=new EventSource('/stream');
es.onmessage=(e)=>{ if(e.data!==undefined){ p.textContent+=e.data+"\\n"; p.scrollTop=p.scrollHeight; } };
</script>
"""

    @app.route("/")
    def index():
        return render_template_string(TEMPLATE)

    @app.route("/start", methods=["POST"])
    def start():
        if RUN_FLAG["running"]: return "already running", 200
        RUN_FLAG["running"]=True
        RUN_FLAG["force"]=False
        threading.Thread(target=_ui_controller, daemon=True).start()
        return "ok", 200

    @app.route("/start_force", methods=["POST"])
    def start_force():
        if RUN_FLAG["running"]: return "already running", 200
        RUN_FLAG["running"]=True
        RUN_FLAG["force"]=True
        threading.Thread(target=_ui_controller, daemon=True).start()
        return "ok", 200

    @app.route("/stop", methods=["POST"])
    def stop():
        RUN_FLAG["running"]=False
        if UILOGQ: UILOGQ.put("STOP angefordert")
        return "ok", 200

    @app.route("/reset", methods=["POST"])
    def reset():
        a,b = reset_history()
        msg = f"Reset: queries_done={a}, urls_seen={b} gelöscht"
        if UILOGQ: UILOGQ.put(msg)
        return "ok", 200

    @app.route("/reset_seen", methods=["POST"])
    def reset_seen():
        con=db(); cur=con.cursor()
        cur.execute("DELETE FROM urls_seen WHERE ts < datetime('now','-1 day')")
        con.commit(); con.close()
        if UILOGQ: UILOGQ.put("urls_seen (älter 24h) bereinigt")
        return "ok", 200

    @app.route("/reset_queries", methods=["POST"])
    def reset_queries():
        con=db(); cur=con.cursor()
        cur.execute("DELETE FROM queries_done")
        con.commit(); con.close()
        if UILOGQ: UILOGQ.put("queries_done geleert")
        return "ok", 200

    @app.route("/stream")
    def stream():
        def gen():
            while True:
                try:
                    msg = UILOGQ.get(timeout=1)
                    yield f"data: {msg}\n\n"
                except queue.Empty:
                    time.sleep(0.5)  # entlastet CPU; kein leeres Event
        return Response(gen(), mimetype='text/event-stream')

    @app.route("/logs")
    def logs():
        return "<pre>Live-Logs im /stream-Event-Stream.</pre>"

    app.run(host=host, port=port, debug=False, threaded=True)

def _ui_controller():
    try:
        force = bool(RUN_FLAG.get("force"))
        # UI-Run: DATE_RESTRICT (falls gesetzt) aus ENV übernehmen
        dr = os.getenv("DATE_RESTRICT", "").strip() or None
        asyncio.run(
            run_scrape_once_async(
                RUN_FLAG,
                ui_log=lambda s: UILOGQ.put(s),
                force=force,
                date_restrict=dr
            )
        )
    except Exception as e:
        if UILOGQ: UILOGQ.put(f"Fehler: {e}")
    finally:
        RUN_FLAG["running"]=False
        RUN_FLAG["force"]=False
        if UILOGQ: UILOGQ.put("Run beendet")


# =========================
# CLI / Entry
# =========================

def validate_config():
    errs=[]
    if not OPENAI_API_KEY or len(OPENAI_API_KEY) < 40:
        errs.append("OPENAI_API_KEY zu kurz/leer (für KI-Extraktion). Regex-Fallback läuft dennoch.")
    if (GCS_KEYS and not GCS_CXS) or (GCS_CXS and not GCS_KEYS):
        errs.append("GCS_KEYS/GCS_CXS unvollständig (beide listenbasiert setzen oder deaktivieren).")
    if not (GCS_KEYS and GCS_CXS) and (GCS_API_KEY and not GCS_CX):
        errs.append("GCS_CX fehlt trotz GCS_API_KEY (Legacy-Single-Config).")
    if errs: log("warn","Konfiguration Hinweise", errors=errs)

def parse_args():
    ap = argparse.ArgumentParser(description="NRW Vertrieb-Leads Scraper (inkrementell + UI)")
    ap.add_argument("--ui", action="store_true", help="Web-UI starten (Start/Stop/Logs)")
    ap.add_argument("--once", action="store_true", help="Einmaliger Lauf im CLI")
    ap.add_argument("--interval", type=int, default=0, help="Pause in Sekunden zwischen den Durchläufen im Loop-Modus")
    ap.add_argument("--force", action="store_true", help="Ignoriere History (queries_done)")
    ap.add_argument("--tor", action="store_true", help="Leite Traffic über Tor (SOCKS5 127.0.0.1:9050)")
    ap.add_argument("--reset", action="store_true", help="Lösche queries_done und urls_seen vor dem Lauf")
    ap.add_argument("--industry", choices=["all","recruiter","candidates"] + list(INDUSTRY_ORDER),
                default=os.getenv("INDUSTRY","all"),
                help="Branche für diesen Run (Standard: all)")
    ap.add_argument("--qpi", type=int, default=int(os.getenv("QPI","6")),
                    help="Queries pro Branche in diesem Run (Standard: 6)")
    ap.add_argument("--daterestrict", type=str, default=os.getenv("DATE_RESTRICT",""),
                    help="Google CSE dateRestrict, z.B. d30, w8, m3")
    ap.add_argument("--smart", action="store_true", help="AI-generierte Dorks (selbstlernend) aktivieren")
    ap.add_argument("--no-google", action="store_true", help="Google CSE deaktivieren")
    ap.add_argument(
        "--mode",
        choices=["standard", "learning", "aggressive", "snippet_only"],
        default="standard",
        help="Betriebsmodus: standard, learning (lernt aus Erfolgen), aggressive (mehr Requests), snippet_only (nur Snippets)"
    )
    return ap.parse_args()



if __name__ == "__main__":
    try:
        args = parse_args()
        USE_TOR = bool(getattr(args, "tor", False))
        if getattr(args, "no_google", False):
            os.environ["DISABLE_GOOGLE"] = "1"
        
        # === TASK 1: Global Proxy Reset (Nuclear Option) ===
        # When NOT using Tor, aggressively clear all proxy environment variables
        # to prevent WinError 10060 ConnectTimeout issues from inherited proxy settings
        if not USE_TOR:
            for var in PROXY_ENV_VARS:
                if var in os.environ:
                    del os.environ[var]
                    log("debug", f"Cleared proxy environment variable: {var}")
            # Set no_proxy to wildcard to bypass any remaining system defaults
            os.environ["no_proxy"] = "*"
            os.environ["NO_PROXY"] = "*"
            log("info", "Proxy environment cleaned: Direct connection mode enforced")
        
        if USE_TOR:
            log("info", "ANONYMITY MODE: Running over TOR Network")
            try:
                tor_resp = asyncio.run(http_get_async("https://check.torproject.org/api/ip", timeout=15))
                if tor_resp and getattr(tor_resp, "status_code", 0) == 200:
                    log("info", "TOR check ok", status=tor_resp.status_code, body=(getattr(tor_resp, "text", "") or "")[:200])
                else:
                    log("warn", "TOR check failed", status=getattr(tor_resp, "status_code", None))
            except Exception as e:
                log("warn", "TOR check failed", error=str(e))
        validate_config()
        init_db()
        migrate_db_unique_indexes()
        
        # Initialize mode
        mode = getattr(args, "mode", "standard")
        mode_config = init_mode(mode)

        if args.reset:
            a,b = reset_history()
            log("info","Reset durchgeführt", queries_done=a, urls_seen=b)

        def _run_cycle():
            global QUERIES
            selected_industry = getattr(args, "industry", "all")
            # CRITICAL FIX: Set INDUSTRY env variable for _is_candidates_mode()
            os.environ["INDUSTRY"] = selected_industry
            per_industry_limit = max(1, getattr(args, "qpi", 2))
            QUERIES = build_queries(selected_industry, per_industry_limit)
            log("info", "Query-Set gebaut", industry=selected_industry,
                per_industry_limit=per_industry_limit, count=len(QUERIES))
            
            # Optimize query order if learning mode is active
            if mode_config.get("query_optimization") and _LEARNING_ENGINE:
                original_count = len(QUERIES)
                QUERIES = _LEARNING_ENGINE.optimize_query_order(QUERIES)
                log("info", "Learning: Query-Reihenfolge optimiert", count=original_count)
            
            if getattr(args, "smart", False):
                if not OPENAI_API_KEY:
                    log("warn", "AI-Smart Dorks aktiviert, aber kein OPENAI_API_KEY gesetzt")
                else:
                    log("info", f"[AI] Generiere neue, intelligente Suchanfragen fuer {selected_industry}...")
                    try:
                        smart_extra = asyncio.run(generate_smart_dorks(selected_industry, count=7))
                    except Exception as e:
                        log("warn", "Smart Dorks generation crashed", error=str(e))
                        smart_extra = []
                    if smart_extra:
                        merged = []
                        seen = set()
                        for q in QUERIES + smart_extra:
                            if q in seen:
                                continue
                            seen.add(q)
                            merged.append(q)
                        QUERIES = merged
                        log("info", "Smart Dorks hinzugefuegt", added=len(smart_extra), total=len(QUERIES))

            RUN_FLAG["running"]=True
            RUN_FLAG["force"]=bool(args.force)
            asyncio.run(
                run_scrape_once_async(
                    RUN_FLAG,
                    force=bool(args.force),
                    date_restrict=(args.daterestrict or None)
                )
            )

        if args.ui:
            from flask import Flask
            UILOGQ = queue.Queue(maxsize=1000)
            RUN_FLAG["running"]=False
            RUN_FLAG["force"]=False
            start_ui()
        elif args.once:
            _run_cycle()
        else:
            while True:
                try:
                    _run_cycle()
                    if args.interval > 0:
                        log("info", f"Schlafe {args.interval} Sekunden...")
                        time.sleep(args.interval)
                except Exception as e:
                    log("error", "Crash im Loop, warte 60 Sekunden und starte neu...", error=str(e), tb=traceback.format_exc())
                    time.sleep(60)
                    continue

    except KeyboardInterrupt:
        die("Abgebrochen (Strg+C)")
    except Exception as e:
        die("Unerwarteter Fehler", error=str(e), tb=traceback.format_exc())

# --- NOTFALL-PATCH fuer fehlende Konstanten ---
DENY_DOMAINS = {
    'stepstone.de', 'indeed.com', 'monster.de', 'arbeitsagentur.de',
    'xing.com/jobs', 'linkedin.com/jobs', 'meinestadt.de', 'kimeta.de',
    'jobware.de', 'stellenanzeigen.de', 'absolventa.de', 'glassdoor.de',
    'kununu.com', 'azubiyo.de', 'ausbildung.de', 'gehalt.de', 'lehrstellen-radar.de',
    'wikipedia.org', 'youtube.com', 'amazon.de', 'ebay.de',
    'heyjobs.de', 'heyjobs.co', 'softgarden.io', 'jobijoba.de', 'jobijoba.com',
    'ok.ru', 'tiktok.com', 'patents.google.com'
}

DENY_DOMAINS.update({
    "stellenonline.de", "stellenmarkt.de", "stepstone.de", "indeed.com",
    "meinestadt.de", "kimeta.de", "jobware.de", "monster.de",
    "arbeitsagentur.de", "nadann.de", "freshplaza.de", "kikxxl.de",
    "xing.com/jobs", "linkedin.com/jobs", "gehalt.de", "kununu.com",
    "ausbildung.de", "azubiyo.de", "lehrstellen-radar.de",
    "softgarden.io", "jobijoba.de", "jobijoba.com", "heyjobs.de", "heyjobs.co"
})

PORTAL_DOMAINS = {
    't-online.de', 'gmx.net', 'web.de', 'yahoo.com', 'msn.com',
    'freenet.de', '1und1.de', 'vodafone.de', 'telekom.de', 'o2online.de',
    'stepstone.de', 'indeed.com', 'monster.de', 'arbeitsagentur.de',
    'meinestadt.de', 'kimeta.de', 'jobware.de', 'stellenanzeigen.de', 'glassdoor.de',
    'kununu.com', 'azubiyo.de', 'ausbildung.de', 'gehalt.de', 'lehrstellen-radar.de',
    'heyjobs.de', 'heyjobs.co', 'softgarden.io', 'jobijoba.de', 'jobijoba.com',
    'ok.ru', 'tiktok.com', 'patents.google.com'
}
